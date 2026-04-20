"""Excel Budget vs Actual variance — adds analysis sheets + AI commentary."""
from __future__ import annotations

import io
import json
import logging
from typing import Any

logger = logging.getLogger(__name__)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from app.services import llm_service
from app.services.json_llm_extract import parse_llm_json_dict
from app.services.excel_formatter import (
    BORDER,
    INR_NUM_FMT,
    auto_width,
    expense_like_from_account,
    rag_fill_for_variance,
    style_header_row,
)


def _parse_num(val: Any) -> float:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("₹", "").replace("$", "").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _find_sheet(xl: pd.ExcelFile, candidates: tuple[str, ...]) -> str | None:
    names = {s.lower(): s for s in xl.sheet_names}
    for c in candidates:
        if c.lower() in names:
            return names[c.lower()]
    for s in xl.sheet_names:
        sl = s.lower().strip()
        for c in candidates:
            if c.lower() in sl:
                return s
    return None


def _account_column(df: pd.DataFrame) -> str:
    cols = list(df.columns.astype(str))
    for c in cols:
        cl = c.lower()
        if "account" in cl or "particular" in cl or "description" in cl or "gl" in cl or "line" in cl:
            return c
    return cols[0]


def _value_total(row: pd.Series, df: pd.DataFrame) -> float:
    """Prefer YTD / Total / FY column; else sum numeric columns except account."""
    cols = [str(c) for c in df.columns]
    acc = _account_column(df)
    priority = ("ytd", "total", "fy", "full year", "annual")
    for p in priority:
        for c in cols:
            if p in c.lower() and c != acc:
                return _parse_num(row.get(c, 0))
    total = 0.0
    for c in cols:
        if c == acc:
            continue
        if str(row.get(c, "")).strip() == "":
            continue
        total += _parse_num(row.get(c, 0))
    return total


def _sheet_to_totals(xl: pd.ExcelFile, sheet: str) -> pd.DataFrame:
    df = pd.read_excel(xl, sheet_name=sheet, engine="openpyxl")
    df = df.dropna(how="all")
    acc_col = _account_column(df)
    out = []
    for _, row in df.iterrows():
        name = str(row.get(acc_col, "")).strip()
        if not name or name.lower() in ("nan", "account", "total", "grand total"):
            continue
        out.append({"Account": name, "Amount": _value_total(row, df)})
    return pd.DataFrame(out)


def analyse_variance(file_bytes: bytes) -> bytes:
    bio = io.BytesIO(file_bytes)
    xl = pd.ExcelFile(bio, engine="openpyxl")
    actual_name = _find_sheet(xl, ("actual", "actuals", "act"))
    budget_name = _find_sheet(xl, ("budget", "bud", "plan"))
    if not actual_name or not budget_name:
        sheets = xl.sheet_names
        if len(sheets) >= 2:
            actual_name = actual_name or sheets[0]
            budget_name = budget_name or sheets[1]
        else:
            raise ValueError("Need two sheets (Actual and Budget) or named Actual/Budget.")

    act = _sheet_to_totals(xl, actual_name)
    bud = _sheet_to_totals(xl, budget_name)
    merged = act.merge(bud, on="Account", how="outer", suffixes=("_Actual", "_Budget")).fillna(0)
    if "Amount_Actual" in merged.columns:
        merged = merged.rename(columns={"Amount_Actual": "Actual", "Amount_Budget": "Budget"})
    elif "Amount_x" in merged.columns:
        merged = merged.rename(columns={"Amount_x": "Actual", "Amount_y": "Budget"})
    else:
        merged["Actual"] = merged.get("Actual", 0)
        merged["Budget"] = merged.get("Budget", 0)

    rows = []
    for _, r in merged.iterrows():
        acct = str(r["Account"])
        a = float(r["Actual"])
        b = float(r["Budget"])
        var = a - b
        var_pct = (var / abs(b) * 100) if b else (100.0 if var else 0.0)
        exp_like = expense_like_from_account(acct)
        flag = ""
        if exp_like and var > 0 and b and var / abs(b) > 0.10:
            flag = "Over budget (>10%)"
        elif not exp_like and var < 0 and b and abs(var / b) > 0.10:
            flag = "Below budget (>10%)"
        elif abs(var_pct) >= 5:
            flag = "Watch"
        else:
            flag = "OK"
        rows.append(
            {
                "Account": acct,
                "Actual": a,
                "Budget": b,
                "Variance": var,
                "Var%": var_pct,
                "Flag": flag,
                "AI Commentary": "",
            }
        )

    df_va = pd.DataFrame(rows).sort_values(by="Account").reset_index(drop=True)

    material = df_va.copy()
    material["abs_pct"] = material["Var%"].abs()
    material = material.nlargest(15, "abs_pct")

    summary_json = {
        "accounts": len(df_va),
        "total_actual": float(df_va["Actual"].sum()),
        "total_budget": float(df_va["Budget"].sum()),
        "top_variances": material[["Account", "Actual", "Budget", "Variance", "Var%"]].head(12).to_dict("records"),
    }

    narrative = ""
    top_issues: list[str] = []
    actions: list[str] = []
    if llm_service.is_configured():
        prompt = f"""You are a CFO. Given this variance summary JSON, write:
1) Exactly 3 short paragraphs (executive narrative) for the board.
2) JSON key "top_issues": array of exactly 5 strings (top issues to address).
3) JSON key "recommended_actions": array of exactly 5 strings (concrete actions).

Summary:
{json.dumps(summary_json, indent=2)}

Reply with ONLY valid JSON (no markdown fences, no text before or after the object). Escape any double quotes inside strings. Use \\n for line breaks inside "narrative".
Schema: {{"narrative": "paragraph1\\n\\nparagraph2\\n\\nparagraph3", "top_issues": [], "recommended_actions": []}}"""
        try:
            raw = llm_service.invoke(prompt, max_tokens=1800, temperature=0.25)
            data = parse_llm_json_dict(raw)
            if data:
                narrative = str(data.get("narrative", "")).strip()
                top_issues = [str(x) for x in data.get("top_issues", [])][:5]
                actions = [str(x) for x in data.get("recommended_actions", [])][:5]
            else:
                logger.warning("Variance AI: could not parse JSON from model output (preview=%s)", raw[:400])
                narrative = "AI commentary unavailable (parse error — model did not return valid JSON). Review Variance_Analysis sheet manually."
        except Exception as e:
            logger.warning("Variance AI: invoke or parse failed: %s", e, exc_info=True)
            narrative = "AI commentary unavailable (API or network error). Confirm ANTHROPIC_API_KEY and backend can reach Anthropic; then re-run. Review Variance_Analysis sheet manually."
    else:
        narrative = (
            "ANTHROPIC_API_KEY is not set on the server. Numeric variance analysis is complete; "
            "configure the API key for AI narrative."
        )
        top_issues = ["Enable Claude API for AI-generated issues list."]
        actions = ["Set ANTHROPIC_API_KEY in backend environment and re-run."]

    # Per-row short commentary for material rows (batched)
    commentary_map: dict[str, str] = {}
    mat_accounts = material.head(8)["Account"].tolist()
    if mat_accounts and llm_service.is_configured():
        compact = df_va[df_va["Account"].isin(mat_accounts)][["Account", "Actual", "Budget", "Variance", "Var%"]]
        p2 = f"""For each account in this table, one sentence of CFO commentary (why it matters). Return JSON object mapping account name to string.
Table CSV:
{compact.to_csv(index=False)}
Return JSON only (no markdown), keys exactly as Account names."""
        try:
            raw2 = llm_service.invoke(p2, max_tokens=1200, temperature=0.2)
            parsed = parse_llm_json_dict(raw2)
            if parsed:
                commentary_map = {str(k): str(v) for k, v in parsed.items()}
        except Exception:
            pass

    for idx in range(len(df_va)):
        acct = str(df_va.at[idx, "Account"])
        if acct in commentary_map:
            df_va.at[idx, "AI Commentary"] = str(commentary_map[acct])[:500]

    # Executive summary KPIs (simple totals — treat all as P&L lines; user refines in Excel)
    total_a = float(df_va["Actual"].sum())
    total_b = float(df_va["Budget"].sum())
    ebit_actual = total_a  # placeholder label; real split would need mapping
    ebit_budget = total_b

    out_bio = io.BytesIO()
    out_bio.write(file_bytes)
    out_bio.seek(0)
    wb = load_workbook(out_bio)

    if "Variance_Analysis" in wb.sheetnames:
        del wb["Variance_Analysis"]
    if "AI_Commentary" in wb.sheetnames:
        del wb["AI_Commentary"]
    if "Executive_Summary" in wb.sheetnames:
        del wb["Executive_Summary"]

    ws_v = wb.create_sheet("Variance_Analysis", 0)
    headers = ["Account", "Actual", "Budget", "Variance", "Var%", "Flag", "AI Commentary"]
    ws_v.append(headers)
    style_header_row(ws_v, 1, len(headers))
    for _, r in df_va.iterrows():
        ws_v.append(
            [
                r["Account"],
                r["Actual"],
                r["Budget"],
                r["Variance"],
                r["Var%"] / 100.0,
                r["Flag"],
                r["AI Commentary"],
            ]
        )
    for row_idx in range(2, ws_v.max_row + 1):
        c = ws_v.cell(row=row_idx, column=5)
        pct = float(df_va.iloc[row_idx - 2]["Var%"])
        c.value = pct / 100.0
        c.number_format = "0.0%"
        for col in (2, 3, 4):
            ws_v.cell(row=row_idx, column=col).number_format = INR_NUM_FMT
        exp_like = expense_like_from_account(str(ws_v.cell(row=row_idx, column=1).value))
        var = float(ws_v.cell(row=row_idx, column=4).value or 0)
        bud = float(ws_v.cell(row=row_idx, column=3).value or 0)
        fill = rag_fill_for_variance(exp_like, var, bud)
        if fill:
            for col in range(1, 6):
                ws_v.cell(row=row_idx, column=col).fill = fill
        for col in range(1, 8):
            ws_v.cell(row=row_idx, column=col).border = BORDER
        ws_v.cell(row=row_idx, column=7).alignment = Alignment(wrap_text=True, vertical="top")

    auto_width(ws_v)

    ws_ai = wb.create_sheet("AI_Commentary")
    ws_ai["A1"] = "CFO Narrative"
    ws_ai["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws_ai["A2"] = narrative
    ws_ai["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_ai.merge_cells("A2:F12")
    r0 = 14
    ws_ai.cell(row=r0, column=1, value="Top 5 issues").font = Font(bold=True)
    for i, issue in enumerate(top_issues[:5], start=1):
        ws_ai.cell(row=r0 + i, column=1, value=f"{i}. {issue}")
    r1 = r0 + 7
    ws_ai.cell(row=r1, column=1, value="Recommended actions").font = Font(bold=True)
    for i, act in enumerate(actions[:5], start=1):
        ws_ai.cell(row=r1 + i, column=1, value=f"{i}. {act}")
    ws_ai.column_dimensions["A"].width = 100

    ws_e = wb.create_sheet("Executive_Summary")
    ws_e.merge_cells("A1:F1")
    ws_e["A1"] = "Executive Summary — KPI snapshot"
    ws_e["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws_e.append([])
    hdr = ["KPI", "Actual", "Budget", "Variance", "Var%", "Note"]
    ws_e.append(hdr)
    style_header_row(ws_e, 3, len(hdr))
    kpi_rows = [
        ("Total P&L movement (sum of lines)", total_a, total_b, total_a - total_b),
        ("EBIT (proxy: same as total lines)", ebit_actual, ebit_budget, ebit_actual - ebit_budget),
    ]
    for label, a, b, v in kpi_rows:
        pct = (v / abs(b) * 100) if b else 0
        ws_e.append([label, a, b, v, pct / 100.0, "Paste into board pack"])
    for row_idx in range(4, ws_e.max_row + 1):
        for col in (2, 3, 4):
            ws_e.cell(row=row_idx, column=col).number_format = INR_NUM_FMT
        ws_e.cell(row=row_idx, column=5).number_format = "0.0%"
    auto_width(ws_e)

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final.getvalue()
