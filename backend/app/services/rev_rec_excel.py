"""IFRS 15 Rev Rec — period close workbook (openpyxl)."""
from __future__ import annotations

import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# FinReportAI palette (ARGB without alpha prefix works in openpyxl 3.1)
NAVY = "0F2D5E"
BLUE = "1D4ED8"
ORANGE = "EA580C"
GREEN = "15803D"
RED = "DC2626"
GREY = "F3F4F6"
WHITE = "FFFFFF"
LIGHT_BLUE = "DBEAFE"
LIGHT_GREEN = "DCFCE7"
LIGHT_RED = "FEE2E2"
LIGHT_ORANGE = "FFEDD5"

REV_REC_EXCEL_OUTPUT_DIR = Path(__file__).resolve().parent.parent.parent / "outputs"


def thin_border() -> Border:
    t = Side(style="thin", color="D1D5DB")
    return Border(left=t, right=t, top=t, bottom=t)


def title_row(ws, row: int, col_start: int, col_end: int, text: str, fill: str = NAVY) -> None:
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = PatternFill("solid", fgColor=fill)
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22


def hdr_row(ws, row: int, cols: list[str], fill: str = BLUE) -> None:
    for col, text in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = PatternFill("solid", fgColor=fill)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 16


def data_cell(
    ws,
    row: int,
    col: int,
    value: Any,
    *,
    fill: str = WHITE,
    bold: bool = False,
    color: str = "000000",
    align: str = "left",
    num_fmt: str | None = None,
):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=fill)
    c.font = Font(name="Arial", size=10, bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = thin_border()
    if num_fmt:
        c.number_format = num_fmt
    return c


def spacer(ws, row: int) -> None:
    ws.row_dimensions[row].height = 8


def generate_period_close_pack(data: dict) -> dict:
    """
    data keys (all optional except period):
    period, customer_name, roll_forward_result, three_way_match_result,
    anomaly_result, rpo_result, commission_result, period_close_result
    """
    wb = Workbook()
    period = data.get("period") or "Unknown Period"
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── Sheet 1 ───────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Period Close Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 24
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 20
    ws1.sheet_properties.tabColor = NAVY

    title_row(ws1, 1, 1, 4, "REVENUE RECOGNITION RECONCILIATION — PERIOD CLOSE SUMMARY")
    ws1.merge_cells("A2:D2")
    sub = ws1["A2"]
    sub.value = f"Period: {period} | Generated: {generated} | IFRS AI — FinReportAI"
    sub.fill = PatternFill("solid", fgColor=BLUE)
    sub.font = Font(name="Arial", size=9, color=WHITE, italic=True)
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[2].height = 14
    spacer(ws1, 3)

    pcs = data.get("period_close_result") or {}
    overall = str(pcs.get("overall_status", "—"))
    total_exc = pcs.get("total_exceptions", 0)
    high_risk = pcs.get("high_risk_exceptions", 0)
    modules_run = pcs.get("modules_run", 0)

    status_fill_map = {"Clean": LIGHT_GREEN, "Exceptions": LIGHT_ORANGE, "High Risk": LIGHT_RED}
    status_color_map = {"Clean": GREEN, "Exceptions": "9A3412", "High Risk": "991B1B"}

    ws1.merge_cells("A4:D4")
    sc = ws1["A4"]
    sc.value = f"OVERALL STATUS: {overall.upper()}"
    sc.fill = PatternFill("solid", fgColor=status_fill_map.get(overall, GREY))
    sc.font = Font(name="Arial", size=13, bold=True, color=status_color_map.get(overall, "000000"))
    sc.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[4].height = 24
    spacer(ws1, 5)

    hdr_row(ws1, 6, ["Metric", "Value", "", ""])
    summary_rows = [
        ("Modules Completed", f"{modules_run} / 5"),
        ("Total Exceptions", str(total_exc)),
        ("High Risk Exceptions", str(high_risk)),
        ("Period", str(period)),
        ("Report Generated", generated),
    ]
    fills = [GREY, WHITE]
    for i, (k, v) in enumerate(summary_rows):
        row = 7 + i
        ws1.row_dimensions[row].height = 15
        data_cell(ws1, row, 1, k, fill=fills[i % 2], bold=True, color="1E3A5F")
        ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        data_cell(ws1, row, 2, v, fill=fills[i % 2])

    last_summary_row = 7 + len(summary_rows) - 1
    spacer(ws1, last_summary_row + 1)
    mod_row = last_summary_row + 2
    title_row(ws1, mod_row, 1, 4, "MODULE STATUS", NAVY)

    hdr_row(ws1, mod_row + 1, ["Module", "Status", "Detail", "Risk"])
    module_statuses = pcs.get("module_statuses") or []
    fills2 = [LIGHT_BLUE, WHITE]
    for i, ms in enumerate(module_statuses):
        row = mod_row + 2 + i
        ws1.row_dimensions[row].height = 15
        st = str(ms.get("status", ""))
        st_l = st.lower()
        sf = LIGHT_GREEN if "clean" in st_l else LIGHT_RED if "high" in st_l else LIGHT_ORANGE
        font_c = GREEN if "clean" in st_l else "991B1B" if "high" in st_l else "9A3412"
        vals = [ms.get("module", ""), st.upper(), ms.get("detail", ""), st.upper()]
        for col, val in enumerate(vals, 1):
            row_fill = sf if col in (2, 4) else fills2[i % 2]
            data_cell(ws1, row, col, val, fill=row_fill, bold=(col == 2), color=font_c if col in (2, 4) else "000000")

    action_start = mod_row + 2 + len(module_statuses) + 1
    spacer(ws1, action_start - 1)
    title_row(ws1, action_start, 1, 4, "ACTION ITEMS", ORANGE)
    action_items = pcs.get("action_items") or []
    if action_items:
        hdr_row(ws1, action_start + 1, ["Priority", "Action", "Owner", "Due Date"])
        for i, ai in enumerate(action_items):
            row = action_start + 2 + i
            ws1.row_dimensions[row].height = 16
            pri = str(ai.get("priority", ""))
            pf = LIGHT_RED if pri == "HIGH" else LIGHT_ORANGE if pri == "MEDIUM" else WHITE
            for col, val in enumerate(
                [pri, ai.get("description", ""), ai.get("owner", ""), ai.get("due_date", "")], 1
            ):
                data_cell(ws1, row, col, val, fill=pf, bold=(col == 1))
        last_action_row = action_start + 1 + len(action_items)
    else:
        ws1.merge_cells(start_row=action_start + 1, start_column=1, end_row=action_start + 1, end_column=4)
        data_cell(
            ws1,
            action_start + 1,
            1,
            "No action items — period is clean.",
            fill=LIGHT_GREEN,
            color=GREEN,
            bold=True,
        )
        last_action_row = action_start + 1

    narrative_start = last_action_row + 2
    title_row(ws1, narrative_start, 1, 4, "EXECUTIVE SUMMARY (AI Generated)", NAVY)
    narrative_text = str(pcs.get("nova_executive_summary", "") or "")
    ws1.merge_cells(start_row=narrative_start + 1, start_column=1, end_row=narrative_start + 5, end_column=4)
    nc = ws1.cell(row=narrative_start + 1, column=1, value=narrative_text)
    nc.font = Font(name="Times New Roman", size=10, color="1F2937")
    nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    nc.fill = PatternFill("solid", fgColor=GREY)
    for r in range(narrative_start + 1, narrative_start + 6):
        ws1.row_dimensions[r].height = 24

    # ── Sheet 2 ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Deferred Rev Roll-Forward")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 20
    ws2.sheet_properties.tabColor = BLUE

    title_row(ws2, 1, 1, 3, f"DEFERRED REVENUE ROLL-FORWARD — {period}")
    ws2.merge_cells("A2:C2")
    sub2 = ws2["A2"]
    sub2.value = "IFRS 15 — Contract Liability Movement | Reconciled to GL Closing Balance"
    sub2.fill = PatternFill("solid", fgColor=BLUE)
    sub2.font = Font(name="Arial", size=9, color=WHITE, italic=True)
    sub2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[2].height = 14
    spacer(ws2, 3)

    hdr_row(ws2, 4, ["Line Item", "Amount ($)", "Status"])
    rf = data.get("roll_forward_result") or {}
    lines = rf.get("roll_forward_lines") or []
    rf_fills = [GREY, WHITE]

    for i, line in enumerate(lines):
        row = 5 + i
        ws2.row_dimensions[row].height = 15
        label = str(line.get("label", ""))
        amount = line.get("amount", 0)
        if amount is None:
            amount = 0
        try:
            amt_f = float(amount)
        except (TypeError, ValueError):
            amt_f = 0.0
        direction = str(line.get("direction", ""))
        is_diff = direction == "difference"
        is_total = direction in ("total", "gl", "difference", "opening")

        cell_fill = (
            LIGHT_RED
            if (is_diff and abs(amt_f) >= 1)
            else LIGHT_GREEN
            if (is_diff and abs(amt_f) < 1)
            else NAVY
            if (is_total and not is_diff)
            else rf_fills[i % 2]
        )
        font_color = (
            WHITE
            if (is_total and not is_diff)
            else "991B1B"
            if (is_diff and abs(amt_f) >= 1)
            else GREEN
            if (is_diff and abs(amt_f) < 1)
            else "000000"
        )

        data_cell(ws2, row, 1, label, fill=cell_fill, bold=is_total, color=font_color)
        data_cell(
            ws2,
            row,
            2,
            amt_f,
            fill=cell_fill,
            bold=is_total,
            color=font_color,
            align="right",
            num_fmt="#,##0.00",
        )
        status_text = ""
        if is_diff:
            status_text = "✓ Reconciled" if abs(amt_f) < 1 else f"▲ Difference: ${abs(amt_f):,.2f}"
        data_cell(ws2, row, 3, status_text, fill=cell_fill, bold=is_diff, color=font_color)

    exc_row = 5 + len(lines) + 1
    spacer(ws2, exc_row - 1)
    exceptions = rf.get("exceptions") or []
    if exceptions:
        title_row(ws2, exc_row, 1, 3, "EXCEPTIONS", ORANGE)
        hdr_row(ws2, exc_row + 1, ["Contract ID", "Difference ($)", "AI Explanation"])
        for i, exc in enumerate(exceptions):
            row = exc_row + 2 + i
            ws2.row_dimensions[row].height = 20
            vals = [exc.get("contract_id", ""), exc.get("difference", 0), exc.get("nova_explanation", "")]
            for col, val in enumerate(vals, 1):
                data_cell(
                    ws2,
                    row,
                    col,
                    val,
                    fill=LIGHT_RED if i % 2 == 0 else WHITE,
                    num_fmt=("#,##0.00" if col == 2 else None),
                )
        commentary_row = exc_row + 2 + len(exceptions) + 1
    else:
        ws2.merge_cells(start_row=exc_row, start_column=1, end_row=exc_row, end_column=3)
        data_cell(
            ws2,
            exc_row,
            1,
            "✓ No exceptions — roll-forward reconciles to GL.",
            fill=LIGHT_GREEN,
            color=GREEN,
            bold=True,
        )
        commentary_row = exc_row + 2

    title_row(ws2, commentary_row, 1, 3, "AI INSIGHT", BLUE)
    ws2.merge_cells(start_row=commentary_row + 1, start_column=1, end_row=commentary_row + 3, end_column=3)
    cc = ws2.cell(row=commentary_row + 1, column=1, value=str(rf.get("nova_commentary", "") or ""))
    cc.font = Font(name="Times New Roman", size=10, italic=True, color="1F2937")
    cc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cc.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    for r in range(commentary_row + 1, commentary_row + 4):
        ws2.row_dimensions[r].height = 20

    # ── Sheet 3 ───────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Three-Way Match")
    ws3.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [18, 22, 16, 16, 16, 18, 12]):
        ws3.column_dimensions[col].width = w
    ws3.sheet_properties.tabColor = BLUE

    title_row(ws3, 1, 1, 7, f"THREE-WAY MATCH — {period}")
    spacer(ws3, 2)
    tm = data.get("three_way_match_result") or {}
    match_rate = float(tm.get("match_rate_pct", 0) or 0)
    total_c = tm.get("total_contracts", 0)
    matched = tm.get("matched", 0)
    unmatched = tm.get("unmatched", 0)

    hdr_row(ws3, 3, ["Total Contracts", "Matched", "Unmatched", "Match Rate %", "", "", ""])
    for col, val in enumerate([total_c, matched, unmatched, f"{match_rate:.1f}%", "", "", ""], 1):
        fill = (
            LIGHT_GREEN
            if col == 2
            else LIGHT_RED
            if col == 3
            else (
                LIGHT_GREEN
                if match_rate >= 95
                else LIGHT_ORANGE
                if match_rate >= 85
                else LIGHT_RED
            )
            if col == 4
            else WHITE
        )
        data_cell(ws3, 4, col, val, fill=fill, bold=(col <= 4), align="center")
    ws3.row_dimensions[4].height = 18
    spacer(ws3, 5)

    hdr_row(ws3, 6, ["Contract ID", "Customer", "Billing ($)", "GL ($)", "Difference ($)", "Status", "Risk"])
    items = tm.get("items") or []
    if not items:
        ws3.merge_cells(start_row=7, start_column=1, end_row=7, end_column=7)
        data_cell(ws3, 7, 1, "No three-way match data available.", fill=GREY, color="6B7280")
        nova_tm_row = 10
    else:
        for i, item in enumerate(items):
            row = 7 + i
            ws3.row_dimensions[row].height = 14
            status = str(item.get("status", ""))
            risk = str(item.get("risk", ""))
            row_fill = (
                LIGHT_GREEN
                if status == "matched"
                else LIGHT_RED
                if risk == "high"
                else LIGHT_ORANGE
                if risk == "medium"
                else WHITE
            )
            vals = [
                item.get("contract_id", ""),
                item.get("customer", ""),
                item.get("billing_amount") or 0,
                item.get("gl_amount") or 0,
                item.get("difference") or 0,
                status.replace("_", " ").upper(),
                risk.upper(),
            ]
            for col, val in enumerate(vals, 1):
                fmt = "#,##0.00" if col in (3, 4, 5) else None
                fc = GREEN if status == "matched" else "991B1B" if risk == "high" else "000000"
                data_cell(
                    ws3,
                    row,
                    col,
                    val,
                    fill=row_fill,
                    num_fmt=fmt,
                    align=("right" if col in (3, 4, 5) else "left"),
                    color=fc,
                )
        nova_tm_row = 7 + len(items) + 2

    title_row(ws3, nova_tm_row, 1, 7, "AI SUMMARY", BLUE)
    ws3.merge_cells(start_row=nova_tm_row + 1, start_column=1, end_row=nova_tm_row + 3, end_column=7)
    nc3 = ws3.cell(row=nova_tm_row + 1, column=1, value=str(tm.get("nova_summary", "") or ""))
    nc3.font = Font(name="Times New Roman", size=10, italic=True, color="1F2937")
    nc3.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    nc3.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    for r in range(nova_tm_row + 1, nova_tm_row + 4):
        ws3.row_dimensions[r].height = 20

    # ── Sheet 4 ───────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Revenue Anomalies")
    ws4.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [16, 16, 20, 28, 12, 38]):
        ws4.column_dimensions[col].width = w
    ws4.sheet_properties.tabColor = ORANGE

    title_row(ws4, 1, 1, 6, f"REVENUE ANOMALY DETECTION — {period}", ORANGE)
    an = data.get("anomaly_result") or {}
    spacer(ws4, 2)
    hdr_row(ws4, 3, ["Total Entries", "Flagged", "Flag Rate %", "High Risk", "Benford Dev.", ""])
    fr = float(an.get("flag_rate_pct", 0) or 0)
    for col, val in enumerate(
        [
            an.get("total_entries", 0),
            an.get("flagged_count", 0),
            f"{fr:.1f}%",
            an.get("high_risk_entries", 0),
            f"{float(an.get('benford_deviation', 0) or 0):.3f}",
            "",
        ],
        1,
    ):
        fill = (
            LIGHT_RED
            if (col == 2 and int(an.get("flagged_count", 0) or 0) > 0)
            else LIGHT_ORANGE
            if (col == 3 and fr > 5)
            else LIGHT_RED
            if (col == 4 and int(an.get("high_risk_entries", 0) or 0) > 0)
            else WHITE
        )
        data_cell(ws4, 4, col, val, fill=fill, bold=True, align="center")
    ws4.row_dimensions[4].height = 18
    spacer(ws4, 5)

    hdr_row(ws4, 6, ["Account", "Amount ($)", "Posted By", "Flag Types", "Risk", "AI Assessment"])
    flags = an.get("flags") or []
    if not flags:
        ws4.merge_cells(start_row=7, start_column=1, end_row=7, end_column=6)
        data_cell(
            ws4,
            7,
            1,
            "✓ No anomalies detected in revenue journal entries.",
            fill=LIGHT_GREEN,
            color=GREEN,
            bold=True,
        )
        nova_an_row = 10
    else:
        for i, flag in enumerate(flags):
            row = 7 + i
            ws4.row_dimensions[row].height = 20
            entry = flag.get("entry") or {}
            risk = str(flag.get("risk", ""))
            row_fill = LIGHT_RED if risk == "high" else LIGHT_ORANGE if risk == "medium" else WHITE
            debit = float(entry.get("debit", 0) or 0)
            credit = float(entry.get("credit", 0) or 0)
            amount = entry.get("amount")
            if amount is not None:
                try:
                    amt_val = abs(float(amount))
                except (TypeError, ValueError):
                    amt_val = abs(debit - credit)
            else:
                amt_val = abs(debit - credit)
            vals = [
                entry.get("account_code", ""),
                amt_val,
                entry.get("posted_by", ""),
                ", ".join(flag.get("flag_types") or []),
                risk.upper(),
                flag.get("nova_assessment", ""),
            ]
            for col, val in enumerate(vals, 1):
                data_cell(
                    ws4,
                    row,
                    col,
                    val,
                    fill=row_fill,
                    num_fmt=("#,##0.00" if col == 2 else None),
                    align=("right" if col == 2 else "left"),
                )
        nova_an_row = 7 + len(flags) + 2

    title_row(ws4, nova_an_row, 1, 6, "AI BATCH SUMMARY", BLUE)
    ws4.merge_cells(start_row=nova_an_row + 1, start_column=1, end_row=nova_an_row + 3, end_column=6)
    nc4 = ws4.cell(row=nova_an_row + 1, column=1, value=str(an.get("nova_batch_summary", "") or ""))
    nc4.font = Font(name="Times New Roman", size=10, italic=True, color="1F2937")
    nc4.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    nc4.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for r in range(nova_an_row + 1, nova_an_row + 4):
        ws4.row_dimensions[r].height = 20

    # ── Sheet 5 ───────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("RPO & Commission")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 36
    ws5.column_dimensions["B"].width = 20
    ws5.column_dimensions["C"].width = 20
    ws5.sheet_properties.tabColor = GREEN

    title_row(ws5, 1, 1, 3, f"RPO & COMMISSION RECONCILIATION — {period}")
    spacer(ws5, 2)
    title_row(ws5, 3, 1, 3, "REMAINING PERFORMANCE OBLIGATIONS", BLUE)

    rpo = data.get("rpo_result") or {}
    rpo_lines = rpo.get("movement_lines") or []
    hdr_row(ws5, 4, ["Line Item", "Amount ($)", "Status"])
    for i, line in enumerate(rpo_lines):
        row = 5 + i
        ws5.row_dimensions[row].height = 15
        amount = line.get("amount", 0)
        try:
            amt_f = float(amount)
        except (TypeError, ValueError):
            amt_f = 0.0
        label = str(line.get("label", ""))
        is_diff = "Difference" in label
        is_total = "Expected" in label or "Disclosed" in label

        fill = (
            LIGHT_RED
            if (is_diff and abs(amt_f) >= 1)
            else LIGHT_GREEN
            if (is_diff and abs(amt_f) < 1)
            else NAVY
            if (is_total and not is_diff)
            else (GREY if i % 2 == 0 else WHITE)
        )
        font_c = (
            WHITE
            if (is_total and not is_diff)
            else "991B1B"
            if (is_diff and abs(amt_f) >= 1)
            else GREEN
            if (is_diff and abs(amt_f) < 1)
            else "000000"
        )
        data_cell(ws5, row, 1, label, fill=fill, bold=is_total, color=font_c)
        data_cell(ws5, row, 2, amt_f, fill=fill, bold=is_total, color=font_c, align="right", num_fmt="#,##0.00")
        status_txt = ""
        if is_diff:
            status_txt = "✓ Reconciled" if abs(amt_f) < 1 else f"▲ ${abs(amt_f):,.2f}"
        data_cell(ws5, row, 3, status_txt, fill=fill, bold=is_diff, color=font_c)

    rpo_comm = str(rpo.get("nova_commentary", "") or "").strip()
    rpo_nova_row = 5 + len(rpo_lines) + 1
    if rpo_comm:
        ws5.merge_cells(start_row=rpo_nova_row, start_column=1, end_row=rpo_nova_row + 1, end_column=3)
        nc5a = ws5.cell(row=rpo_nova_row, column=1, value=rpo_comm)
        nc5a.font = Font(name="Times New Roman", size=10, italic=True, color="1F2937")
        nc5a.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        nc5a.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for r in range(rpo_nova_row, rpo_nova_row + 2):
            ws5.row_dimensions[r].height = 20
        comm_start = rpo_nova_row + 3
    else:
        comm_start = rpo_nova_row + 1

    spacer(ws5, comm_start - 1)
    title_row(ws5, comm_start, 1, 3, "CONTRACT COST ASSET (COMMISSION)", BLUE)
    cm = data.get("commission_result") or {}
    comm_lines = [
        ("Opening Asset", cm.get("opening_asset", 0)),
        ("+ New Commissions Capitalised", cm.get("new_commissions", 0)),
        ("- Monthly Amortisation", -float(cm.get("amortisation", 0) or 0)),
        ("Expected Closing Balance", cm.get("expected_closing", 0)),
        ("GL Closing Balance", cm.get("gl_closing_balance", 0)),
        ("Difference", cm.get("difference", 0)),
    ]
    hdr_row(ws5, comm_start + 1, ["Line Item", "Amount ($)", "Status"])
    for i, (label, amount) in enumerate(comm_lines):
        row = comm_start + 2 + i
        ws5.row_dimensions[row].height = 15
        try:
            amt_f = float(amount)
        except (TypeError, ValueError):
            amt_f = 0.0
        is_diff = label == "Difference"
        is_total = "Expected" in label or "GL Closing" in label

        fill = (
            LIGHT_RED
            if (is_diff and abs(amt_f) > 0.01)
            else LIGHT_GREEN
            if (is_diff and abs(amt_f) <= 0.01)
            else NAVY
            if (is_total and not is_diff)
            else (GREY if i % 2 == 0 else WHITE)
        )
        font_c = (
            WHITE
            if (is_total and not is_diff)
            else "991B1B"
            if (is_diff and abs(amt_f) > 0.01)
            else GREEN
            if (is_diff and abs(amt_f) <= 0.01)
            else "000000"
        )
        data_cell(ws5, row, 1, label, fill=fill, bold=is_total, color=font_c)
        data_cell(ws5, row, 2, amt_f, fill=fill, bold=is_total, color=font_c, align="right", num_fmt="#,##0.00")
        status_txt = ""
        if is_diff:
            status_txt = "✓ Reconciled" if abs(amt_f) <= 0.01 else f"▲ ${abs(amt_f):,.2f}"
        data_cell(ws5, row, 3, status_txt, fill=fill, bold=is_diff, color=font_c)

    # ── Sheet 6 ───────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("AI Audit Commentary")
    ws6.sheet_view.showGridLines = False
    ws6.column_dimensions["A"].width = 14
    ws6.column_dimensions["B"].width = 68
    ws6.sheet_properties.tabColor = NAVY

    title_row(ws6, 1, 1, 2, f"RECONCILIATION AUDIT MEMO — {period}")
    ws6.merge_cells("A2:B2")
    sub6 = ws6["A2"]
    sub6.value = "AI-generated commentary for audit file. Review and approve before sign-off."
    sub6.fill = PatternFill("solid", fgColor=BLUE)
    sub6.font = Font(name="Arial", size=9, color=WHITE, italic=True)
    sub6.alignment = Alignment(horizontal="left", vertical="center")
    ws6.row_dimensions[2].height = 14
    spacer(ws6, 3)

    current_row = 4
    all_commentary: list[tuple[str, str]] = []
    rf2 = data.get("roll_forward_result") or {}
    if rf2.get("nova_commentary"):
        all_commentary.append(("Deferred Revenue Roll-Forward", str(rf2["nova_commentary"])))
    tm2 = data.get("three_way_match_result") or {}
    if tm2.get("nova_summary"):
        all_commentary.append(("Three-Way Match", str(tm2["nova_summary"])))
    an2 = data.get("anomaly_result") or {}
    if an2.get("nova_batch_summary"):
        all_commentary.append(("Revenue Anomaly Detection", str(an2["nova_batch_summary"])))
    rpo2 = data.get("rpo_result") or {}
    if rpo2.get("nova_commentary"):
        all_commentary.append(("RPO Movement", str(rpo2["nova_commentary"])))
    cm2 = data.get("commission_result") or {}
    if cm2.get("nova_commentary"):
        all_commentary.append(("Commission Asset", str(cm2["nova_commentary"])))
    pc2 = data.get("period_close_result") or {}
    if pc2.get("nova_executive_summary"):
        all_commentary.append(("Period Close — Executive Summary", str(pc2["nova_executive_summary"])))

    if not all_commentary:
        ws6.merge_cells(start_row=4, start_column=1, end_row=6, end_column=2)
        data_cell(
            ws6,
            4,
            1,
            "No commentary generated yet. Run modules and generate period close report to populate.",
            fill=GREY,
            color="6B7280",
        )
        current_row = 7
    else:
        for section_label, commentary_text in all_commentary:
            title_row(ws6, current_row, 1, 2, section_label, NAVY)
            current_row += 1
            ws6.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 3, end_column=2)
            tc = ws6.cell(row=current_row, column=1, value=commentary_text)
            tc.font = Font(name="Times New Roman", size=11, color="1F2937")
            tc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            tc.fill = PatternFill("solid", fgColor=GREY)
            for r in range(current_row, current_row + 4):
                ws6.row_dimensions[r].height = 22
            current_row += 5
            spacer(ws6, current_row)
            current_row += 1

    disclaimer_row = current_row + 1
    ws6.merge_cells(start_row=disclaimer_row, start_column=1, end_row=disclaimer_row, end_column=2)
    dc = ws6.cell(
        row=disclaimer_row,
        column=1,
        value=(
            "This commentary was generated by IFRS AI — FinReportAI. All content must be reviewed "
            "and approved by a qualified accountant before use in audit files or financial statements."
        ),
    )
    dc.font = Font(name="Arial", size=9, italic=True, color="6B7280")
    dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws6.row_dimensions[disclaimer_row].height = 24

    # ── Save ──────────────────────────────────────────────────────────────
    file_id = str(uuid.uuid4())[:8]
    safe_period = str(period).replace(" ", "_").replace("/", "-")
    filename = f"RevRec_PeriodClose_{safe_period}_{file_id}.xlsx"
    REV_REC_EXCEL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    filepath = os.path.join(REV_REC_EXCEL_OUTPUT_DIR, filename)
    wb.save(filepath)

    return {"file_id": file_id, "filename": filename, "sheets": 6, "path": filepath}
