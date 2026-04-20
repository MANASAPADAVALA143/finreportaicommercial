"""Shared Excel styling for Excel AI Suite (openpyxl)."""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Theme: dark blue + orange (board / professional)
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ACCENT_FILL = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
ACCENT_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RAG_RED = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
RAG_AMBER = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
RAG_GREEN = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

# Indian grouping / currency (Excel locale may still control separators)
INR_NUM_FMT = "₹#,##0.00"


def style_header_row(ws, row: int = 1, max_col: int | None = None) -> None:
    last = max_col or ws.max_column
    for c in range(1, last + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def auto_width(ws, min_w: int = 10, max_w: int = 48) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        maxlen = min_w
        for row in range(1, min(ws.max_row + 1, 200)):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                maxlen = max(maxlen, min(max_w, len(str(v)) + 2))
        ws.column_dimensions[letter].width = maxlen


def set_title_banner(ws, title: str, subtitle: str = "", rows: int = 2) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(size=16, bold=True, color="1E3A5F")
    t.alignment = Alignment(horizontal="left", vertical="center")
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        s = ws.cell(row=2, column=1, value=subtitle)
        s.font = Font(size=10, color="64748B")
        s.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def rag_fill_for_variance(is_expense_like: bool, variance: float, budget: float) -> PatternFill | None:
    """Return RAG fill for Actual vs Budget variance. Unfavorable depends on account type."""
    if budget == 0 and variance == 0:
        return None
    pct = (variance / abs(budget)) * 100 if budget else 0.0
    if is_expense_like:
        unfavorable = variance > 0  # spent more than budget
    else:
        unfavorable = variance < 0  # revenue below budget
    mag = abs(pct)
    if not unfavorable:
        return RAG_GREEN if mag >= 5 else None
    if mag > 10:
        return RAG_RED
    if mag >= 5:
        return RAG_AMBER
    return None


def expense_like_from_account(name: str) -> bool:
    n = (name or "").lower()
    revenue_hints = ("revenue", "sales", "income", "turnover", "topline", "other income")
    expense_hints = (
        "cost",
        "expense",
        "cogs",
        "opex",
        "payroll",
        "salary",
        "rent",
        "depreciation",
        "interest",
        "tax",
        "overhead",
        "material",
        "freight",
        "marketing",
        "admin",
    )
    if any(h in n for h in revenue_hints) and not any(x in n for x in ("cost of", "cogs")):
        return False
    if any(h in n for h in expense_hints):
        return True
    return True  # default P&L below-the-line as cost-like for conservative RAG
