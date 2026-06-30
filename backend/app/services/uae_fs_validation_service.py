"""Financial statement validation and Excel export."""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from sqlalchemy.orm import Session

from app.models.uae_accounting_full import UAEAccount, UAEJournalEntry, UAEJournalLine
from app.services.uae_journal_service import get_trial_balance


def _period_from_dates(period_start: str, period_end: str) -> str:
    """Use YYYY-MM from period_end for trial balance."""
    return period_end[:7] if period_end else period_start[:7]


def _prior_period(period: str) -> str:
    y, m = int(period[:4]), int(period[5:7])
    if m == 1:
        return f"{y - 1}-12"
    return f"{y:04d}-{m - 1:02d}"


def _sum_by_prefix(tb: dict, prefixes: tuple[str, ...], account_type: str | None = None) -> float:
    total = 0.0
    for line in tb.get("lines", []):
        code = line["account_code"]
        if any(code.startswith(p) for p in prefixes):
            total += line.get("net_balance", 0)
        elif account_type and code:
            pass
    return total


def _sum_accounts(tb: dict, codes: set[str]) -> float:
    return sum(
        line.get("net_balance", 0)
        for line in tb.get("lines", [])
        if line["account_code"] in codes
    )


def _sum_type(tb: dict, key: str) -> float:
    return float(tb.get("totals", {}).get(key, 0))


def validate_financial_statements(
    db: Session,
    workspace_id: str,
    company_id: str | None,
    period_start: str,
    period_end: str,
) -> dict[str, Any]:
    period = _period_from_dates(period_start, period_end)
    prior = _prior_period(period)
    tb = get_trial_balance(workspace_id, period, db, company_id=company_id)
    tb_prior = get_trial_balance(workspace_id, prior, db, company_id=company_id)

    total_assets = _sum_type(tb, "asset")
    total_liabilities = _sum_type(tb, "liability")
    total_equity = _sum_type(tb, "equity")
    diff_bs = total_assets - (total_liabilities + total_equity)
    bs_passed = abs(diff_bs) < 1.0
    check1 = {
        "check": "balance_sheet",
        "passed": bs_passed,
        "total_assets": round(total_assets, 2),
        "total_liabilities": round(total_liabilities, 2),
        "total_equity": round(total_equity, 2),
        "difference": round(diff_bs, 2),
        "message": "Balance Sheet Validated ✓" if bs_passed else f"Difference of AED {diff_bs:,.2f}",
    }

    bs_cash = sum(
        line.get("net_balance", 0)
        for line in tb.get("lines", [])
        if line["account_code"].startswith("100") or line["account_code"] in ("1001", "1002", "1003", "1004", "1005")
    )
    prior_cash = sum(
        line.get("net_balance", 0)
        for line in tb_prior.get("lines", [])
        if line["account_code"].startswith("100")
    )
    revenue = _sum_type(tb, "revenue")
    expense = _sum_type(tb, "expense")
    net_movement = revenue - expense
    cf_closing = prior_cash + net_movement
    diff_cf = cf_closing - bs_cash
    cf_passed = abs(diff_cf) < 1.0
    check2 = {
        "check": "cash_flow",
        "passed": cf_passed,
        "bs_cash": round(bs_cash, 2),
        "cf_closing": round(cf_closing, 2),
        "cf_opening": round(prior_cash, 2),
        "cf_net_movement": round(net_movement, 2),
        "difference": round(diff_cf, 2),
        "message": "Cash Flow Validated ✓" if cf_passed else f"Cash difference AED {diff_cf:,.2f}",
    }

    re_codes = {"5002", "5003"}
    current_re = sum(
        line.get("net_balance", 0)
        for line in tb.get("lines", [])
        if line["account_code"] in re_codes
    )
    prior_re = sum(
        line.get("net_balance", 0)
        for line in tb_prior.get("lines", [])
        if line["account_code"] in re_codes
    )
    re_movement = current_re - prior_re
    net_profit = revenue - expense
    diff_pl = re_movement - net_profit
    pl_passed = abs(diff_pl) < 1.0 or abs(net_profit) < 1.0
    check3 = {
        "check": "pl_equity",
        "passed": pl_passed,
        "retained_earnings_movement": round(re_movement, 2),
        "net_profit": round(net_profit, 2),
        "difference": round(diff_pl, 2),
        "message": "P&L to Equity Validated ✓" if pl_passed else f"P&L/equity difference AED {diff_pl:,.2f}",
    }

    checks = [check1, check2, check3]
    return {
        "all_passed": all(c["passed"] for c in checks),
        "checks": checks,
        "validated_at": datetime.utcnow().isoformat() + "Z",
        "period": period,
    }


def export_fs_excel(
    db: Session,
    workspace_id: str,
    company_id: str | None,
    period_start: str,
    period_end: str,
) -> bytes:
    from openpyxl import Workbook

    period = _period_from_dates(period_start, period_end)
    tb = get_trial_balance(workspace_id, period, db, company_id=company_id)
    totals = tb.get("totals", {})
    wb = Workbook()

    ws_pl = wb.active
    ws_pl.title = "Income Statement"
    ws_pl.append(["Line Item", "Amount AED"])
    ws_pl.append(["Revenue", totals.get("revenue", 0)])
    ws_pl.append(["Expenses", totals.get("expense", 0)])
    ws_pl.append(["Net Profit", totals.get("revenue", 0) - totals.get("expense", 0)])

    ws_bs = wb.create_sheet("Balance Sheet")
    ws_bs.append(["Line Item", "Amount AED"])
    ws_bs.append(["Total Assets", totals.get("asset", 0)])
    ws_bs.append(["Total Liabilities", totals.get("liability", 0)])
    ws_bs.append(["Total Equity", totals.get("equity", 0)])

    ws_cf = wb.create_sheet("Cash Flow")
    ws_cf.append(["Category", "Amount AED"])
    ws_cf.append(["Operating (approx)", totals.get("revenue", 0) - totals.get("expense", 0)])
    ws_cf.append(["Cash balance", totals.get("cash", 0)])

    ws_notes = wb.create_sheet("Notes")
    ws_notes.append(["Account Code", "Account Name", "Note #", "Balance"])
    for line in tb.get("lines", []):
        ws_notes.append([line["account_code"], line["account_name"], "", line.get("net_balance", 0)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
