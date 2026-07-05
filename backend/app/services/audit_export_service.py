"""Audit-ready tax period export — RDS artifacts only (Excel + ZIP manifest)."""
from __future__ import annotations

import hashlib
import io
import json
import zipfile
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.models.client_data import (
    BadDebtReliefClaim,
    CtReturn,
    DesignatedZoneTransaction,
    GulftaxTransaction,
    PartialExemptionCalculation,
)
from app.models.workspace_audit import WorkspaceAuditLog
from app.modules.gulftax.vat_return_service import fetch_all_vat_return_boxes, parse_period
from app.services.gulftax_sync_service import _fetch_company_config
from app.services.vat_recon_service import _latest_recon

TAX_RELEVANT_ACTIONS = frozenset(
    {
        "invoice_approved",
        "invoice_rejected",
        "je_posted",
        "je_approved",
        "je_pending_approval",
        "je_deleted",
        "ar_invoice_created",
        "ar_payment_received",
        "company_setup_completed",
        "period_locked",
        "fx_revaluation_posted",
    }
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)


def _company_meta(company_id: str) -> tuple[str, str]:
    cfg = _fetch_company_config(company_id)
    name = str(cfg.get("name") or cfg.get("company_name") or company_id)
    trn = str(cfg.get("trn") or cfg.get("gstin") or cfg.get("vat_trn") or "")
    return name, trn


def _serialize_ct(row: CtReturn | None) -> dict[str, Any] | None:
    if not row:
        return None
    return {
        "id": row.id,
        "period_start": row.period_start.isoformat() if row.period_start else None,
        "period_end": row.period_end.isoformat() if row.period_end else None,
        "revenue": float(row.revenue or 0),
        "accounting_profit": float(row.accounting_profit or 0),
        "non_deductible_expenses": float(row.non_deductible_expenses or 0),
        "taxable_income": float(row.taxable_income or 0),
        "ct_payable_aed": float(row.ct_payable_aed or 0),
        "sbr_eligible": bool(row.sbr_eligible),
        "qfzp_eligible": bool(row.qfzp_eligible),
        "free_zone_status": row.free_zone_status,
        "status": row.status,
        "approved_at": row.approved_at.isoformat() if row.approved_at else None,
        "filed_at": row.filed_at.isoformat() if row.filed_at else None,
        "override_reason": row.override_reason,
        "breakdown": row.breakdown,
    }


def _latest_ct_return(
    db: Session,
    *,
    tenant_id: str,
    company_id: str,
    period_start: date,
    period_end: date,
) -> CtReturn | None:
    return (
        db.query(CtReturn)
        .filter(
            CtReturn.tenant_id == tenant_id,
            CtReturn.company_id == company_id,
            CtReturn.period_start <= period_end,
            CtReturn.period_end >= period_start,
        )
        .order_by(CtReturn.created_at.desc())
        .first()
    )


def _gather_transactions(
    db: Session,
    *,
    tenant_id: str,
    company_id: str,
    tax_period: str,
) -> list[GulftaxTransaction]:
    return (
        db.query(GulftaxTransaction)
        .filter(
            GulftaxTransaction.tenant_id == tenant_id,
            GulftaxTransaction.company_id == company_id,
            GulftaxTransaction.tax_period == tax_period,
            GulftaxTransaction.status == "posted",
        )
        .order_by(GulftaxTransaction.transaction_date.asc())
        .all()
    )


def _gather_audit_trail(
    db: Session,
    *,
    tenant_id: str,
    company_id: str,
    period_start: date,
    period_end: date,
) -> list[WorkspaceAuditLog]:
    start_dt = datetime.combine(period_start, datetime.min.time())
    end_dt = datetime.combine(period_end, datetime.max.time())
    try:
        return (
            db.query(WorkspaceAuditLog)
            .filter(
                WorkspaceAuditLog.workspace_id == tenant_id,
                WorkspaceAuditLog.company_id == company_id,
                WorkspaceAuditLog.created_at >= start_dt,
                WorkspaceAuditLog.created_at <= end_dt,
                WorkspaceAuditLog.action.in_(TAX_RELEVANT_ACTIONS),
            )
            .order_by(WorkspaceAuditLog.created_at.asc())
            .all()
        )
    except Exception:
        db.rollback()
        return []


def _style_header_row(ws, row: int, ncol: int) -> None:
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _append_table(ws, headers: list[str], rows: list[list[Any]]) -> int:
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)
    return len(rows)


def _build_excel(
    *,
    company_name: str,
    trn: str,
    tax_period: str,
    period_start: date,
    period_end: date,
    generated_at: datetime,
    generated_by: str | None,
    vat_boxes: dict[str, Any],
    transactions: list[GulftaxTransaction],
    recon_row: Any | None,
    ct_row: CtReturn | None,
    partial_exemption: list[PartialExemptionCalculation],
    bad_debt: list[BadDebtReliefClaim],
    dz_transactions: list[DesignatedZoneTransaction],
    audit_entries: list[WorkspaceAuditLog],
) -> tuple[bytes, dict[str, int]]:
    wb = Workbook()
    sheet_counts: dict[str, int] = {}

    # Cover
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "UAE Tax Audit Pack"
    cover["A1"].font = _TITLE_FONT
    cover_rows = [
        ("Company", company_name),
        ("TRN", trn or "—"),
        ("Tax Period", tax_period),
        ("Period Start", period_start.isoformat()),
        ("Period End", period_end.isoformat()),
        ("Generated At (UTC)", generated_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Generated By", generated_by or "system"),
        ("Data Source", "RDS — gulftax_transactions, reconciliation_results, gulftax_ct_returns"),
    ]
    for idx, (label, value) in enumerate(cover_rows, start=3):
        cover[f"A{idx}"] = label
        cover[f"A{idx}"].font = Font(bold=True)
        cover[f"B{idx}"] = value
    cover.column_dimensions["A"].width = 22
    cover.column_dimensions["B"].width = 52
    sheet_counts["Cover"] = len(cover_rows)

    dz_count = sum(1 for t in transactions if t.designated_zone)
    dz_vat = round(sum(float(t.vat_amount or 0) for t in transactions if t.designated_zone), 2)

    # Sheet 1 — VAT Return Summary
    ws1 = wb.create_sheet("VAT Return Summary")
    vat_fields: list[tuple[str, Any]] = [
        ("Box 1 — Standard rated sales (net)", vat_boxes.get("box1_standard_rated_sales_net")),
        ("Box 1 — Standard rated sales (VAT)", vat_boxes.get("box1_standard_rated_sales_vat")),
        ("Box 2 — Tourist refunds", vat_boxes.get("box2_tourist_refunds")),
        ("Box 2 — Advance payment VAT", vat_boxes.get("box2_advance_payment_vat")),
        ("Box 3 — Reverse charge supplies (net)", vat_boxes.get("box3_reverse_charge_supplies_net")),
        ("Box 3 — Reverse charge supplies (VAT)", vat_boxes.get("box3_reverse_charge_supplies_vat")),
        ("Box 4 — Zero-rated supplies", vat_boxes.get("box4_zero_rated_supplies")),
        ("Box 5 — Exempt supplies", vat_boxes.get("box5_exempt_supplies")),
        ("Box 6 — Imports VAT", vat_boxes.get("box6_imports_vat")),
        ("Box 7 — Output adjustments", vat_boxes.get("box7_output_adjustments")),
        ("Box 8 — Total output VAT", vat_boxes.get("box8_total_output_vat")),
        ("Box 9 — Standard rated expenses", vat_boxes.get("box9_standard_rated_expenses")),
        ("Box 10 — Reverse charge expenses", vat_boxes.get("box10_reverse_charge_expenses")),
        ("Box 11 — Total input VAT (adjusted)", vat_boxes.get("box11_total_input_vat")),
        ("Box 11 — Total input VAT (raw)", vat_boxes.get("box11_total_input_vat_raw")),
        ("Box 12 — Net VAT payable / refundable", vat_boxes.get("box12_net_vat_payable_or_refundable")),
        ("Partial exemption applied", vat_boxes.get("partial_exemption_applied")),
        ("Recovery percentage", vat_boxes.get("recovery_percentage")),
        ("Bad debt relief applied (AED)", vat_boxes.get("bad_debt_relief_applied")),
        ("Designated zone transactions (count)", dz_count),
        ("Designated zone VAT (reference)", dz_vat),
        ("Sales invoice count", vat_boxes.get("sales_invoice_count")),
        ("Purchase entry count", vat_boxes.get("purchase_entry_count")),
        ("Source", vat_boxes.get("source")),
    ]
    sheet_counts["VAT Return Summary"] = _append_table(
        ws1,
        ["Field", "Amount / Value"],
        [[k, v] for k, v in vat_fields],
    )

    # Sheet 2 — Transaction Listing (TAF)
    ws2 = wb.create_sheet("Transaction Listing")
    taf_rows = [
        [
            tx.invoice_number,
            tx.vendor_name,
            tx.vendor_trn,
            float(tx.gross_amount or 0),
            float(tx.vat_amount or 0),
            tx.fta_box,
            tx.direction,
            tx.source,
            tx.ap_invoice_id,
            tx.transaction_date.isoformat() if tx.transaction_date else "",
            tx.vat_category,
            bool(tx.designated_zone),
        ]
        for tx in transactions
    ]
    sheet_counts["Transaction Listing"] = _append_table(
        ws2,
        [
            "Invoice Number",
            "Vendor / Customer",
            "TRN",
            "Gross Amount",
            "VAT Amount",
            "FTA Box",
            "Direction",
            "Source",
            "Source Invoice ID",
            "Transaction Date",
            "VAT Category",
            "Designated Zone",
        ],
        taf_rows,
    )

    # Sheet 3 — VAT Reconciliation
    ws3 = wb.create_sheet("VAT Reconciliation")
    if recon_row:
        recon_rows = [
            ["Status", recon_row.status],
            ["Difference (AED)", float(recon_row.difference_aed or 0)],
            ["Source", recon_row.source],
            ["Override Reason", recon_row.override_reason or ""],
            ["Run At", recon_row.created_at.isoformat() if recon_row.created_at else ""],
        ]
        breakdown = recon_row.box_breakdown or {}
        if isinstance(breakdown, dict):
            for key, val in breakdown.items():
                recon_rows.append([f"Box breakdown — {key}", val])
        sheet_counts["VAT Reconciliation"] = _append_table(ws3, ["Field", "Value"], recon_rows)
    else:
        sheet_counts["VAT Reconciliation"] = _append_table(
            ws3,
            ["Field", "Value"],
            [["Status", "never_run"], ["Note", "No reconciliation_results row for this period"]],
        )

    # Sheet 4 — CT Return
    ws4 = wb.create_sheet("CT Return")
    ct = _serialize_ct(ct_row)
    if ct:
        ct_rows = [
            ["ID", ct["id"]],
            ["Period Start", ct["period_start"]],
            ["Period End", ct["period_end"]],
            ["Taxable Income (AED)", ct["taxable_income"]],
            ["CT Payable (AED)", ct["ct_payable_aed"]],
            ["Revenue (AED)", ct["revenue"]],
            ["Accounting Profit (AED)", ct["accounting_profit"]],
            ["Non-deductible Expenses (AED)", ct["non_deductible_expenses"]],
            ["Status", ct["status"]],
            ["SBR Eligible", ct["sbr_eligible"]],
            ["QFZP Eligible", ct["qfzp_eligible"]],
            ["Free Zone Status", ct["free_zone_status"]],
            ["Approved At", ct["approved_at"] or ""],
            ["Filed At", ct["filed_at"] or ""],
            ["Override Reason", ct["override_reason"] or ""],
            ["Breakdown (JSON)", json.dumps(ct["breakdown"], default=str) if ct["breakdown"] else ""],
        ]
        sheet_counts["CT Return"] = _append_table(ws4, ["Field", "Value"], ct_rows)
    else:
        sheet_counts["CT Return"] = _append_table(
            ws4,
            ["Field", "Value"],
            [["Status", "not_found"], ["Note", "No gulftax_ct_returns row overlapping this period"]],
        )

    # Sheet 5 — Advanced VAT
    ws5 = wb.create_sheet("Advanced VAT")

    def _append_section(ws, title: str, headers: list[str], rows: list[list[Any]]) -> int:
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
        ws.append(headers)
        _style_header_row(ws, ws.max_row, len(headers))
        for row in rows:
            ws.append(row)
        ws.append([])
        return len(rows)

    pe_count = _append_section(
        ws5,
        "Partial Exemption Calculations",
        [
            "Period",
            "Taxable Supplies",
            "Exempt Supplies",
            "Input VAT Paid",
            "Recovery %",
            "Recoverable VAT",
            "Irrecoverable VAT",
            "Status",
        ],
        [
            [
                pe.period,
                float(pe.taxable_supplies or 0),
                float(pe.exempt_supplies or 0),
                float(pe.input_vat_paid or 0),
                float(pe.recovery_pct or 0),
                float(pe.recoverable_vat or 0),
                float(pe.irrecoverable_vat or 0),
                pe.status,
            ]
            for pe in partial_exemption
        ],
    )
    bd_count = _append_section(
        ws5,
        "Bad Debt Relief Claims",
        [
            "Invoice Number",
            "Invoice Date",
            "Invoice Amount",
            "VAT Amount",
            "Claim Period",
            "Status",
            "Eligible",
            "Reason",
        ],
        [
            [
                bd.invoice_number,
                bd.invoice_date.isoformat() if bd.invoice_date else "",
                float(bd.invoice_amount or 0),
                float(bd.vat_amount or 0),
                bd.claim_period,
                bd.status,
                bool(bd.eligible),
                bd.eligibility_reason or "",
            ]
            for bd in bad_debt
        ],
    )
    dz_count_adv = _append_section(
        ws5,
        "Designated Zone Transactions",
        [
            "Supplier Location",
            "Customer Location",
            "Transaction Type",
            "VAT Treatment",
            "VAT Rate",
            "Explanation",
            "Created At",
        ],
        [
            [
                dz.supplier_location,
                dz.customer_location,
                dz.transaction_type,
                dz.vat_treatment,
                float(dz.vat_rate or 0),
                dz.explanation,
                dz.created_at.isoformat() if dz.created_at else "",
            ]
            for dz in dz_transactions
        ],
    )
    sheet_counts["Advanced VAT"] = pe_count + bd_count + dz_count_adv

    # Sheet 6 — Audit Trail
    ws6 = wb.create_sheet("Audit Trail")
    audit_rows = [
        [
            e.created_at.isoformat() if e.created_at else "",
            e.action,
            e.entity_type,
            e.entity_id,
            e.user_email or "",
            json.dumps(e.details, default=str) if e.details else "",
        ]
        for e in audit_entries
    ]
    sheet_counts["Audit Trail"] = _append_table(
        ws6,
        ["Timestamp", "Action", "Entity Type", "Entity ID", "User", "Details"],
        audit_rows,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), sheet_counts


def _build_manifest(
    *,
    company_name: str,
    trn: str,
    tax_period: str,
    period_start: date,
    period_end: date,
    generated_at: datetime,
    generated_by: str | None,
    sheet_counts: dict[str, int],
    excel_bytes: bytes | None,
    preview: bool = False,
) -> dict[str, Any]:
    excel_filename = f"audit_pack_{tax_period.replace('/', '-')}.xlsx"
    artifacts = [
        {"sheet": "Cover", "description": "Pack metadata", "row_count": sheet_counts.get("Cover", 0)},
        {
            "sheet": "VAT Return Summary",
            "description": "FTA 12-box VAT return with adjustments",
            "row_count": sheet_counts.get("VAT Return Summary", 0),
        },
        {
            "sheet": "Transaction Listing",
            "description": "RDS gulftax_transactions (TAF)",
            "row_count": sheet_counts.get("Transaction Listing", 0),
        },
        {
            "sheet": "VAT Reconciliation",
            "description": "Latest reconciliation_results for period",
            "row_count": sheet_counts.get("VAT Reconciliation", 0),
        },
        {
            "sheet": "CT Return",
            "description": "Latest gulftax_ct_returns overlapping period",
            "row_count": sheet_counts.get("CT Return", 0),
        },
        {
            "sheet": "Advanced VAT",
            "description": "Partial exemption, bad debt, designated zones",
            "row_count": sheet_counts.get("Advanced VAT", 0),
        },
        {
            "sheet": "Audit Trail",
            "description": "Workspace audit log (tax-relevant actions)",
            "row_count": sheet_counts.get("Audit Trail", 0),
        },
    ]
    manifest: dict[str, Any] = {
        "company_name": company_name,
        "trn": trn,
        "tax_period": tax_period,
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "generated_at": generated_at.isoformat() + "Z",
        "generated_by": generated_by or "system",
        "excel_filename": excel_filename,
        "artifacts": artifacts,
        "preview": preview,
    }
    if excel_bytes is not None:
        manifest["excel_sha256"] = hashlib.sha256(excel_bytes).hexdigest()
    return manifest


def _build_zip(excel_bytes: bytes, manifest: dict[str, Any], excel_filename: str) -> bytes:
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(excel_filename, excel_bytes)
        zf.writestr("manifest.json", json.dumps(manifest, indent=2, default=str))
    return zip_buf.getvalue()


def _query_period_advanced_vat(
    db: Session,
    *,
    tenant_id: str,
    company_id: str,
    tax_period: str,
    period_start: date,
    period_end: date,
) -> tuple[list[PartialExemptionCalculation], list[BadDebtReliefClaim], list[DesignatedZoneTransaction]]:
    start_dt = datetime.combine(period_start, datetime.min.time())
    end_dt = datetime.combine(period_end, datetime.max.time())
    try:
        partial_exemption = (
            db.query(PartialExemptionCalculation)
            .filter(
                PartialExemptionCalculation.tenant_id == tenant_id,
                PartialExemptionCalculation.company_id == company_id,
                PartialExemptionCalculation.period == tax_period,
            )
            .all()
        )
        bad_debt = (
            db.query(BadDebtReliefClaim)
            .filter(
                BadDebtReliefClaim.tenant_id == tenant_id,
                BadDebtReliefClaim.company_id == company_id,
                BadDebtReliefClaim.claim_period == tax_period,
            )
            .all()
        )
        dz_transactions = (
            db.query(DesignatedZoneTransaction)
            .filter(
                DesignatedZoneTransaction.tenant_id == tenant_id,
                DesignatedZoneTransaction.company_id == company_id,
                DesignatedZoneTransaction.created_at >= start_dt,
                DesignatedZoneTransaction.created_at <= end_dt,
            )
            .all()
        )
        return partial_exemption, bad_debt, dz_transactions
    except Exception:
        db.rollback()
        return [], [], []


def preview_period_manifest(
    db: Session,
    ported_db: Session,
    *,
    tenant_id: str,
    company_id: str,
    tax_period: str,
    generated_by: str | None = None,
) -> dict[str, Any]:
    """Build manifest preview with row counts (no Excel hash)."""
    period_start, period_end = parse_period(tax_period)
    company_name, trn = _company_meta(company_id)
    generated_at = datetime.utcnow()

    transactions = _gather_transactions(db, tenant_id=tenant_id, company_id=company_id, tax_period=tax_period)
    recon_row = _latest_recon(ported_db, company_id=company_id, tax_period=tax_period)
    ct_row = _latest_ct_return(
        db, tenant_id=tenant_id, company_id=company_id, period_start=period_start, period_end=period_end
    )
    partial_exemption, bad_debt, dz_transactions = _query_period_advanced_vat(
        db,
        tenant_id=tenant_id,
        company_id=company_id,
        tax_period=tax_period,
        period_start=period_start,
        period_end=period_end,
    )
    audit_entries = _gather_audit_trail(
        db, tenant_id=tenant_id, company_id=company_id, period_start=period_start, period_end=period_end
    )

    sheet_counts = {
        "Cover": 8,
        "VAT Return Summary": 24,
        "Transaction Listing": len(transactions),
        "VAT Reconciliation": 6 if recon_row else 2,
        "CT Return": 16 if ct_row else 2,
        "Advanced VAT": len(partial_exemption) + len(bad_debt) + len(dz_transactions),
        "Audit Trail": len(audit_entries),
    }

    return _build_manifest(
        company_name=company_name,
        trn=trn,
        tax_period=tax_period,
        period_start=period_start,
        period_end=period_end,
        generated_at=generated_at,
        generated_by=generated_by,
        sheet_counts=sheet_counts,
        excel_bytes=None,
        preview=True,
    )


def generate_period_audit_pack(
    db: Session,
    ported_db: Session,
    *,
    tenant_id: str,
    company_id: str,
    tax_period: str,
    generated_by: str | None = None,
) -> dict[str, Any]:
    """Generate Excel workbook + ZIP with manifest for a tax period."""
    period_start, period_end = parse_period(tax_period)
    company_name, trn = _company_meta(company_id)
    generated_at = datetime.utcnow()

    vat_boxes = fetch_all_vat_return_boxes(
        db,
        workspace_id=tenant_id,
        company_id=company_id,
        period=tax_period,
    )
    transactions = _gather_transactions(db, tenant_id=tenant_id, company_id=company_id, tax_period=tax_period)
    recon_row = _latest_recon(ported_db, company_id=company_id, tax_period=tax_period)
    ct_row = _latest_ct_return(
        db, tenant_id=tenant_id, company_id=company_id, period_start=period_start, period_end=period_end
    )
    partial_exemption, bad_debt, dz_transactions = _query_period_advanced_vat(
        db,
        tenant_id=tenant_id,
        company_id=company_id,
        tax_period=tax_period,
        period_start=period_start,
        period_end=period_end,
    )
    audit_entries = _gather_audit_trail(
        db, tenant_id=tenant_id, company_id=company_id, period_start=period_start, period_end=period_end
    )

    excel_bytes, sheet_counts = _build_excel(
        company_name=company_name,
        trn=trn,
        tax_period=tax_period,
        period_start=period_start,
        period_end=period_end,
        generated_at=generated_at,
        generated_by=generated_by,
        vat_boxes=vat_boxes,
        transactions=transactions,
        recon_row=recon_row,
        ct_row=ct_row,
        partial_exemption=partial_exemption,
        bad_debt=bad_debt,
        dz_transactions=dz_transactions,
        audit_entries=audit_entries,
    )

    manifest = _build_manifest(
        company_name=company_name,
        trn=trn,
        tax_period=tax_period,
        period_start=period_start,
        period_end=period_end,
        generated_at=generated_at,
        generated_by=generated_by,
        sheet_counts=sheet_counts,
        excel_bytes=excel_bytes,
        preview=False,
    )
    excel_filename = manifest["excel_filename"]
    zip_bytes = _build_zip(excel_bytes, manifest, excel_filename)

    return {
        "excel_bytes": excel_bytes,
        "zip_bytes": zip_bytes,
        "manifest": manifest,
        "sheet_row_counts": sheet_counts,
        "zip_filename": f"audit_pack_{tax_period.replace('/', '-')}.zip",
    }
