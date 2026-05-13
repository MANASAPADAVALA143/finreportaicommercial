"""Excel export for 3-statement model (openpyxl)."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

_GREY = PatternFill("solid", fgColor="E2E8F0")
_BLUE = PatternFill("solid", fgColor="DBEAFE")
_YELLOW = PatternFill("solid", fgColor="FEF9C3")
_GREEN = PatternFill("solid", fgColor="DCFCE7")
_BOLD = Font(bold=True)
_ITALIC = Font(italic=True)


def _row_values(stmt: dict[str, Any], line_name: str) -> list[float]:
    for row in stmt.get("rows", []):
        if row.get("is_header"):
            continue
        if str(row.get("line", "")).strip() == line_name:
            return [float(x) for x in row.get("values", [])]
    return []


def build_model_excel_bytes(
    *,
    company_name: str,
    currency: str,
    base_model: dict[str, Any],
    scenarios: dict[str, dict[str, Any]],
    assumptions: dict[str, Any],
    n_hist: int,
) -> bytes:
    wb = Workbook()
    # Sheet 1 — Assumptions
    wa = wb.active
    wa.title = "Assumptions"
    wa["A1"] = "Model assumptions (inputs in yellow)"
    wa["A1"].font = Font(bold=True, size=14)
    r = 3
    wa.cell(r, 1, "Company")
    wa.cell(r, 2, company_name)
    r += 1
    wa.cell(r, 1, "Currency")
    wa.cell(r, 2, currency)
    r += 2
    keys = [
        ("revenue_growth", "Revenue growth (list)"),
        ("gross_margin", "Gross margin (list)"),
        ("ebitda_margin", "EBITDA margin (list)"),
        ("tax_rate", "Tax rate"),
        ("capex_pct_revenue", "Capex % revenue"),
        ("da_pct_revenue", "D&A % revenue"),
        ("interest_rate", "Interest rate"),
        ("dividend_payout", "Dividend payout ratio"),
    ]
    for k, lab in keys:
        wa.cell(r, 1, lab)
        c = wa.cell(r, 2, str(assumptions.get(k, "")))
        c.fill = _YELLOW
        r += 1
    nwc = assumptions.get("nwc_days") or {}
    for nk, lab in [("ar_days", "AR days"), ("inventory_days", "Inventory days"), ("ap_days", "AP days")]:
        wa.cell(r, 1, lab)
        c = wa.cell(r, 2, nwc.get(nk, ""))
        c.fill = _YELLOW
        r += 1
    wa.cell(r, 1, "Debt repayment (list)")
    c = wa.cell(r, 2, str(assumptions.get("debt_repayment", "")))
    c.fill = _YELLOW
    wa.freeze_panes = "B4"
    for col in range(1, 4):
        wa.column_dimensions[get_column_letter(col)].width = 28 if col == 1 else 22

    def write_statement_sheet(name: str, stmt: dict[str, Any], hist_cols: int) -> None:
        w = wb.create_sheet(name)
        labels = stmt.get("labels", [])
        w.cell(1, 1, name)
        w.cell(1, 1).font = _BOLD
        for j, lab in enumerate(labels, start=2):
            w.cell(2, j, lab)
            w.cell(2, j).font = Font(bold=True, color="FFFFFF")
            w.cell(2, j).fill = PatternFill("solid", fgColor="1E40AF")
            if j - 2 < hist_cols:
                w.cell(2, j).fill = PatternFill("solid", fgColor="64748B")
        rr = 3
        for row in stmt.get("rows", []):
            if row.get("is_header"):
                w.cell(rr, 1, row.get("line", ""))
                w.cell(rr, 1).font = Font(bold=True, size=11)
                rr += 1
                continue
            w.cell(rr, 1, row.get("line", ""))
            if row.get("is_bold"):
                w.cell(rr, 1).font = _BOLD
            if row.get("is_percent"):
                w.cell(rr, 1).font = _ITALIC
            vals = row.get("values", [])
            for j in range(len(labels)):
                v = float(vals[j]) if j < len(vals) else 0.0
                c = w.cell(rr, j + 2, v)
                if row.get("is_percent"):
                    c.number_format = "0.0%"
                    c.font = _ITALIC
                if j < hist_cols:
                    c.fill = _GREY
                else:
                    c.fill = _BLUE
            rr += 1
        w.freeze_panes = "C3"
        for col in range(1, len(labels) + 2):
            w.column_dimensions[get_column_letter(col)].width = 28 if col == 1 else 14

    pl = base_model.get("statements", {}).get("income_statement", {})
    bs = base_model.get("statements", {}).get("balance_sheet", {})
    cfs = base_model.get("statements", {}).get("cash_flow", {})
    write_statement_sheet("Income Statement", pl, n_hist)
    write_statement_sheet("Balance Sheet", bs, n_hist)
    # Balance check row on BS sheet
    wbs = wb["Balance Sheet"]
    max_r = wbs.max_row + 2
    wbs.cell(max_r, 1, "Check: Assets - (L+E)")
    wbs.cell(max_r, 1).font = _BOLD
    labels_b = bs.get("labels", [])
    ta_row = _row_values(bs, "Total Assets")
    tle_row = _row_values(bs, "Total Liabilities + Equity")
    for j, _lab in enumerate(labels_b):
        ta = float(ta_row[j]) if j < len(ta_row) else 0.0
        tle = float(tle_row[j]) if j < len(tle_row) else 0.0
        d = ta - tle
        c = wbs.cell(max_r, j + 2, d)
        c.fill = _GREEN if abs(d) < 1 else PatternFill("solid", fgColor="FEE2E2")

    # CFS
    wc = wb.create_sheet("Cash Flow")
    wc.cell(1, 1, "Cash flow (forecast)")
    wc.cell(1, 1).font = _BOLD
    labels_c = cfs.get("labels", [])
    for j, lab in enumerate(labels_c, start=2):
        wc.cell(2, j, lab)
        wc.cell(2, j).font = Font(bold=True, color="FFFFFF")
        wc.cell(2, j).fill = PatternFill("solid", fgColor="1E40AF")
    rr = 3
    for row in cfs.get("rows", []):
        if row.get("is_header"):
            wc.cell(rr, 1, row.get("line", ""))
            wc.cell(rr, 1).font = Font(bold=True)
            rr += 1
            continue
        wc.cell(rr, 1, row.get("line", ""))
        if row.get("is_bold"):
            wc.cell(rr, 1).font = _BOLD
        vals = row.get("values", [])
        for j in range(len(labels_c)):
            v = float(vals[j]) if j < len(vals) else 0.0
            cell = wc.cell(rr, j + 2, v)
            if v < 0 and "Net" in str(row.get("line", "")):
                cell.font = Font(color="B91C1C")
        rr += 1
    wc.freeze_panes = "C3"
    for col in range(1, len(labels_c) + 2):
        wc.column_dimensions[get_column_letter(col)].width = 28 if col == 1 else 14

    # Scenarios sheet — Base | Upside | Downside for each forecast year
    ws5 = wb.create_sheet("Scenarios")
    ws5["A1"] = "Scenario comparison (all forecast years)"
    ws5["A1"].font = _BOLD
    fy = base_model.get("meta", {}).get("forecast_year_list", [])
    rr = 3
    hdr = ["Metric", "Base", "Upside", "Downside"]
    last_ni_base_row: int | None = None

    def _scen_val(mod: dict[str, Any], pl_bs: str, yi: int, key: str) -> float:
        seq = (mod.get("forecast") or {}).get(pl_bs) or []
        if not seq or yi < 0 or yi >= len(seq):
            return 0.0
        return float(seq[yi].get(key, 0))

    for yi, y_year in enumerate(fy):
        ws5.cell(rr, 1, f"FY{y_year}E")
        ws5.cell(rr, 1).font = Font(bold=True, size=11)
        rr += 1
        for i, h in enumerate(hdr, start=1):
            c = ws5.cell(rr, i, h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0F172A")
        rr += 1
        for label, pl_bs, key in [
            ("Revenue", "pl", "revenue"),
            ("EBITDA", "pl", "ebitda"),
            ("Net income", "pl", "net_income"),
            ("Closing cash", "bs", "cash"),
            ("Total debt", "bs", "total_debt"),
        ]:
            ws5.cell(rr, 1, label)
            col = 2
            for sc in ("base", "upside", "downside"):
                mod = scenarios.get(sc) or {}
                v = _scen_val(mod, pl_bs, yi, key)
                c = ws5.cell(rr, col, v)
                if label in ("Revenue", "EBITDA", "Net income"):
                    c.font = Font(bold=True)
                col += 1
            if label == "Net income":
                last_ni_base_row = rr
            rr += 1
        rr += 1
    ws5.freeze_panes = "B4"
    for col in range(1, 6):
        ws5.column_dimensions[get_column_letter(col)].width = 26 if col == 1 else 18

    # Named ranges (key cells for linking / what-if)
    rev_col = get_column_letter(2 + max(0, n_hist))
    # Assumptions: row 3 company, 4 currency, +2 gap → keys start row 6; tax_rate is 4th key → row 9
    tax_row = 9
    try:
        wb.defined_names.add(
            DefinedName("Model_TaxRate", attr_text=f"'Assumptions'!$B${tax_row}")
        )
        wb.defined_names.add(
            DefinedName("Model_Revenue_FY1", attr_text=f"'Income Statement'!${rev_col}$3")
        )
        if last_ni_base_row:
            wb.defined_names.add(
                DefinedName("Model_NetIncome_LastFY", attr_text=f"'Scenarios'!$B${last_ni_base_row}")
            )
    except Exception:
        pass

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()
