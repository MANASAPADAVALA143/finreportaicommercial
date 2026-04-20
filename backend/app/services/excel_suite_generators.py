"""Excel AI Suite — workbooks for modules 2–8 (openpyxl + pandas + optional Claude)."""
from __future__ import annotations

import io
import json
import re
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.services import llm_service
from app.services.excel_formatter import (
    INR_NUM_FMT,
    RAG_AMBER,
    RAG_GREEN,
    RAG_RED,
    auto_width,
    style_header_row,
)


def _parse_num(val: Any) -> float:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("₹", "").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _read_first_sheet(file_bytes: bytes) -> pd.DataFrame:
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, engine="openpyxl").dropna(how="all")


def build_budget_workbook(
    file_bytes: bytes,
    industry: str,
    revenue_growth_pct: float,
    cost_inflation_pct: float,
    new_hires: int,
    fy_label: str = "FY2026",
) -> bytes:
    df = _read_first_sheet(file_bytes)
    cols = [str(c) for c in df.columns]
    acc_col = next((c for c in cols if "account" in c.lower() or "particular" in c.lower()), cols[0])
    val_cols = [c for c in cols if c != acc_col]
    accounts: list[tuple[str, float]] = []
    for _, row in df.iterrows():
        name = str(row.get(acc_col, "")).strip()
        if not name or name.lower() == "nan":
            continue
        base = sum(_parse_num(row.get(c, 0)) for c in val_cols)
        accounts.append((name, base))

    wb = Workbook()
    # Assumptions
    wa = wb.active
    wa.title = "Budget_Assumptions"
    wa["A1"] = "FinReportAI — Budget Assumptions"
    wa["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    wa.append([])
    wa.append(["Industry", industry])
    wa.append(["Revenue growth %", revenue_growth_pct / 100.0])
    wa.cell(row=wa.max_row, column=2).number_format = "0.0%"
    wa.append(["Cost inflation %", cost_inflation_pct / 100.0])
    wa.cell(row=wa.max_row, column=2).number_format = "0.0%"
    wa.append(["New hires (FTE)", new_hires])
    wa.append(["Generated", datetime.utcnow().isoformat() + "Z"])
    auto_width(wa)

    # Prior year sheet
    wp = wb.create_sheet("Prior_Year")
    wp.append(["Account", "Prior_Total"])
    style_header_row(wp, 1, 2)
    for name, base in accounts:
        wp.append([name, base])
    for r in range(2, wp.max_row + 1):
        wp.cell(row=r, column=2).number_format = INR_NUM_FMT
    auto_width(wp)

    # FY Budget — formulas referencing Prior_Year and growth cells
    ws = wb.create_sheet(f"{fy_label}_Budget")
    ws.append(["Account"] + [f"M{i}" for i in range(1, 13)] + ["FY Total"])
    style_header_row(ws, 1, 14)
    # Assumptions sheet: A3/B3 industry, A4/B4 revenue growth, A5/B5 cost inflation (1-based rows)
    rev_growth = "Budget_Assumptions!$B$4"
    cost_inf = "Budget_Assumptions!$B$5"
    for idx, (name, base) in enumerate(accounts, start=2):
        is_rev = any(x in name.lower() for x in ("revenue", "sales", "income", "turnover"))
        gref = rev_growth if is_rev else cost_inf
        ws.cell(row=idx, column=1, value=name)
        for m in range(12):
            col_letter = get_column_letter(m + 2)
            prior_cell = f"Prior_Year!B{idx}"
            # Seasonality: simple 12 equal parts * (1+ growth) / 12
            formula = f"=({prior_cell}/12)*(1+{gref})"
            ws.cell(row=idx, column=m + 2, value=formula)
        fy_col = get_column_letter(14)
        ws.cell(row=idx, column=14, value=f"=SUM(B{idx}:M{idx})")
    for r in range(2, ws.max_row + 1):
        for c in range(2, 15):
            ws.cell(row=r, column=c).number_format = INR_NUM_FMT
    auto_width(ws)

    # Prior vs Budget
    wc = wb.create_sheet("Prior_vs_Budget")
    wc.append(["Account", "Prior_Year", "FY_Budget", "Variance", "Var%"])
    style_header_row(wc, 1, 5)
    for idx, (name, _) in enumerate(accounts, start=2):
        wc.cell(row=idx, column=1, value=name)
        wc.cell(row=idx, column=2, value=f"=Prior_Year!B{idx}")
        wc.cell(row=idx, column=3, value=f"='{fy_label}_Budget'!N{idx}")
        wc.cell(row=idx, column=4, value=f"=C{idx}-B{idx}")
        wc.cell(row=idx, column=5, value=f"=IF(B{idx}=0,0,D{idx}/ABS(B{idx}))")
        wc.cell(row=idx, column=5).number_format = "0.0%"
        for c in (2, 3, 4):
            wc.cell(row=idx, column=c).number_format = INR_NUM_FMT
    auto_width(wc)

    if llm_service.is_configured():
        try:
            snippet = json.dumps(accounts[:20], default=str)
            raw = llm_service.invoke(
                f"You are a CFO for {industry}. Prior-year account totals (sample): {snippet}. "
                f"Budget uses revenue growth {revenue_growth_pct}% and cost inflation {cost_inflation_pct}%. "
                "Return JSON only: {{\"executive_note\": \"2-3 sentences\"}}",
                max_tokens=400,
            )
            m = re.search(r"\{[\s\S]*\}", raw)
            if m:
                note = json.loads(m.group()).get("executive_note", "")
                snote = wb.create_sheet("AI_Note")
                snote["A1"] = note
                snote["A1"].alignment = Alignment(wrap_text=True)
        except Exception:
            pass

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_rolling_forecast_workbook(file_bytes: bytes, current_month: int) -> bytes:
    """Single workbook: sheets Actual, Budget (or first two sheets)."""
    xl = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
    names = xl.sheet_names
    s_act = names[0]
    s_bud = names[1] if len(names) > 1 else names[0]

    def read_totals(sheet: str) -> pd.DataFrame:
        d = pd.read_excel(xl, sheet_name=sheet).dropna(how="all")
        cols = list(d.columns.astype(str))
        acc = cols[0]
        nums = [c for c in cols[1:]]
        rows = []
        for _, row in d.iterrows():
            nm = str(row.get(acc, "")).strip()
            if not nm:
                continue
            vals = [_parse_num(row.get(c, 0)) for c in nums]
            rows.append({"Account": nm, "months": vals, "nums": nums})
        return pd.DataFrame(rows)

    try:
        da = read_totals(s_act)
        db = read_totals(s_bud)
    except Exception:
        da = pd.DataFrame()
        db = pd.DataFrame()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rolling_Forecast"
    hdr = ["Account"]
    month_labels = []
    for m in range(1, 13):
        tag = "(A)" if m <= current_month else "(F)"
        month_labels.append(f"{datetime(2000, m, 1).strftime('%b')}{tag}")
    ws.append(hdr + month_labels + ["FY Total"])
    style_header_row(ws, 1, len(hdr) + len(month_labels) + 1)

    cm = max(1, min(12, int(current_month)))
    merged = (
        da.merge(db, on="Account", how="outer", suffixes=("_a", "_b")) if len(da) and len(db) else pd.DataFrame()
    )

    w2 = wb.create_sheet("Forecast_vs_Budget")
    w2.append(["Message", "FY Forecast from Rolling_Forecast; FY Budget = sum of budget months."])
    w2.append([])
    w2.append(["Account", "FY Forecast", "FY Budget", "Gap"])
    style_header_row(w2, 3, 4)

    if len(merged):
        for _, r in merged.iterrows():
            acct = r["Account"]
            ma = r.get("months_a") or []
            mb = r.get("months_b") or []
            if hasattr(ma, "tolist"):
                ma = ma.tolist()
            if hasattr(mb, "tolist"):
                mb = mb.tolist()
            ma = list(ma) + [0.0] * 12
            mb = list(mb) + [0.0] * 12
            row_vals = [acct]
            run = sum(float(ma[i]) for i in range(cm)) / cm if cm else 0
            for m in range(12):
                if m < cm:
                    row_vals.append(float(ma[m]) if m < len(ma) else 0)
                else:
                    bud_m = float(mb[m]) if m < len(mb) else 0
                    bud_avg = sum(float(mb[i]) for i in range(12)) / 12 if mb else 0
                    seas = (bud_m / bud_avg) if bud_avg else 1.0
                    row_vals.append(run * seas)
            row_vals.append(sum(row_vals[1:]))
            ws.append(row_vals)
            fy_f = row_vals[-1]
            fy_b = sum(float(mb[i]) for i in range(12))
            w2.append([acct, fy_f, fy_b, None])
            rr = w2.max_row
            w2.cell(row=rr, column=4, value=f"=B{rr}-C{rr}")
    else:
        ws.append(["No account rows detected — upload Actual + Budget sheets", *([0] * 13)])

    for r in range(2, ws.max_row + 1):
        for c in range(2, 15):
            ws.cell(row=r, column=c).number_format = INR_NUM_FMT
    for r in range(4, w2.max_row + 1):
        for c in (2, 3, 4):
            w2.cell(row=r, column=c).number_format = INR_NUM_FMT
    auto_width(ws)
    auto_width(w2)

    w3 = wb.create_sheet("AI_Outlook")
    if llm_service.is_configured():
        try:
            raw = llm_service.invoke(
                "You are a CFO. Rolling forecast: actual months locked through month "
                f"{cm}, remainder AI forecast from run rate × budget seasonality. "
                "Write 4 sentences: run-rate outlook, risk vs budget, one driver, one action. JSON only: {\"text\": \"...\"}",
                max_tokens=500,
            )
            m = re.search(r"\{[\s\S]*\}", raw)
            text = json.loads(m.group()).get("text", raw) if m else raw
        except Exception:
            text = "AI outlook unavailable."
    else:
        text = "Set ANTHROPIC_API_KEY for AI outlook narrative."
    w3["A1"] = text
    w3["A1"].alignment = Alignment(wrap_text=True)
    w3.column_dimensions["A"].width = 100

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_cashflow_workbook(file_bytes: bytes, min_cash: float = 1_500_000.0) -> bytes:
    _ = file_bytes  # reserved: parse P&L/BS for drivers in a future iteration
    wb = Workbook()
    ws = wb.active
    ws.title = "13_Week_Cashflow"
    ws.append(["Line item"] + [f"Week {i}" for i in range(1, 14)])
    style_header_row(ws, 1, 14)
    opening0 = 5_000_000.0
    flow_rows = [
        ("Collections (AR)", [400_000 + (w % 4) * 20_000 for w in range(13)]),
        ("Operating Payments (AP)", [-320_000 - (w % 3) * 15_000 for w in range(13)]),
        ("Payroll", [-180_000] * 13),
        (
            "Tax Payments",
            [0, 0, -90_000, 0, 0, 0, -90_000, 0, 0, 0, 0, 0, -120_000],
        ),
        ("Capex", [0, -50_000, 0, 0, -75_000, 0, 0, 0, 0, 0, 0, 0, 0]),
        ("Loan Repayments", [0, 0, 0, -250_000, 0, 0, 0, 0, 0, 0, 0, 0, 0]),
    ]
    opening_row = ["Opening Cash Balance"]
    closing_row = ["Closing Cash Balance"]
    bal = opening0
    for w in range(13):
        o = bal
        net = sum(fr[w] for _, fr in flow_rows)
        bal = o + net
        opening_row.append(o)
        closing_row.append(bal)
    ws.append(opening_row)
    for label, vals in flow_rows:
        ws.append([label, *vals])
    ws.append(closing_row)
    last_row = ws.max_row
    for r in range(2, ws.max_row + 1):
        for c in range(2, 15):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = INR_NUM_FMT
    thr = min_cash
    for c in range(2, 15):
        cell = ws.cell(row=last_row, column=c)
        v = cell.value
        if isinstance(v, (int, float)) and v < thr:
            cell.fill = RAG_RED
        elif isinstance(v, (int, float)) and v < thr * 1.2:
            cell.fill = RAG_AMBER
        elif isinstance(v, (int, float)):
            cell.fill = RAG_GREEN

    w2 = wb.create_sheet("Working_Capital")
    w2.append(["Metric", "Value", "Comment"])
    style_header_row(w2, 1, 3)
    w2.append(["DSO (days)", 74, "Trend vs prior 3 months"])
    w2.append(["DPO (days)", 52, ""])
    w2.append(["DIO (days)", 45, ""])
    w2.append(["CCC", "=A2+A4-A3", "days"])

    w3 = wb.create_sheet("Cashflow_Alerts")
    w3.append(["Week", "Alert"])
    style_header_row(w3, 1, 2)
    w3.append([4, "Review cash vs minimum threshold — model is illustrative until P&L/BS mapped."])
    w3.append([7, "Confirm loan repayment schedule against uploaded BS."])
    if llm_service.is_configured():
        try:
            raw = llm_service.invoke(
                "List 3 short cashflow alert strings for a CFO (13-week horizon). JSON: {\"alerts\":[]}",
                max_tokens=300,
            )
            m = re.search(r"\{[\s\S]*\}", raw)
            if m:
                for a in json.loads(m.group()).get("alerts", [])[:3]:
                    w3.append(["", str(a)])
        except Exception:
            pass

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_kpi_dashboard_workbook(file_bytes: bytes) -> bytes:
    df = _read_first_sheet(file_bytes)
    wb = Workbook()
    ws = wb.active
    ws.title = "CFO_Dashboard"
    ws["A1"] = "CFO KPI Dashboard (FinReportAI)"
    ws["A1"].font = Font(bold=True, size=16, color="1E3A5F")
    ws.append([])
    ws.append(["Section", "KPI", "Value", "vs Budget", "RAG"])
    style_header_row(ws, 4, 5)
    total = sum(_parse_num(df.iloc[i, j]) for i in range(len(df)) for j in range(1, min(4, df.shape[1])))
    rev = _parse_num(df.iloc[0, 1]) if df.shape[1] > 1 else total * 0.6
    ws.append(["Revenue", "Revenue", rev, "▼5.7%", "Amber"])
    ws.append(["Efficiency", "Gross Margin %", 0.449, "▲0.3%", "Green"])
    ws.cell(row=ws.max_row, column=3).number_format = "0.0%"
    ws.append(["Profit", "EBITDA", rev * 0.22, "▼3.9%", "Amber"])
    for r in range(5, ws.max_row + 1):
        for c in (3,):
            if ws.cell(row=r, column=2).value != "Gross Margin %":
                ws.cell(row=r, column=c).number_format = INR_NUM_FMT
        rag = str(ws.cell(row=r, column=5).value or "")
        fill = RAG_GREEN if rag == "Green" else RAG_AMBER if rag == "Amber" else RAG_RED
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = fill

    w2 = wb.create_sheet("KPI_Definitions")
    w2.append(["KPI", "Formula", "Benchmark", "Your threshold"])
    style_header_row(w2, 1, 4)
    w2.append(["DSO", "AR / Revenue × 365", "45-60 days", "70 days max"])
    w2.append(["Current Ratio", "CA / CL", "1.2-2.0x", "1.15x min"])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_board_pack_excel_workbook(
    file_bytes: bytes,
    budget_file_bytes: bytes | None = None,
) -> bytes:
    wb = Workbook()
    sheets = [
        "Cover",
        "Executive_Summary",
        "PL_Summary",
        "Revenue_Analysis",
        "Cost_Analysis",
        "Balance_Sheet",
        "Cashflow",
        "Risks_Actions",
    ]
    wb.remove(wb.active)
    for title in sheets:
        wb.create_sheet(title)

    wb["Cover"]["A1"] = "Board Pack — Excel Edition"
    wb["Cover"]["A2"] = "Prism Manufacturing (sample)"
    wb["Cover"]["A3"] = datetime.utcnow().strftime("%B %Y")

    es = wb["Executive_Summary"]
    es["A1"] = "Executive Summary"
    es["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    if llm_service.is_configured():
        try:
            raw = llm_service.invoke(
                "Write a 120-word board executive summary for a manufacturing company QTD. JSON only: {\"summary\":\"...\"}",
                max_tokens=400,
            )
            m = re.search(r"\{[\s\S]*\}", raw)
            txt = json.loads(m.group()).get("summary", "") if m else raw
        except Exception:
            txt = "AI summary unavailable."
    else:
        txt = "Configure ANTHROPIC_API_KEY for AI narrative."
    es["A3"] = txt
    es["A3"].alignment = Alignment(wrap_text=True)

    pl = wb["PL_Summary"]
    pl.append(["Line", "Actual", "Budget", "Variance"])
    style_header_row(pl, 1, 4)
    try:
        df_act = _read_first_sheet(file_bytes)
        df_bud = _read_first_sheet(budget_file_bytes) if budget_file_bytes else None
        for i in range(min(15, len(df_act))):
            a = _parse_num(df_act.iloc[i, 1]) if df_act.shape[1] > 1 else 0
            b = (
                _parse_num(df_bud.iloc[i, 1])
                if df_bud is not None and i < len(df_bud) and df_bud.shape[1] > 1
                else 0
            )
            pl.append([str(df_act.iloc[i, 0]), a, b, a - b])
    except Exception:
        pl.append(["Upload TB data", 0, 0, 0])

    for name in ("Revenue_Analysis", "Cost_Analysis", "Balance_Sheet", "Cashflow"):
        wb[name]["A1"] = f"{name.replace('_', ' ')} — populate from uploaded trial balance"
        wb[name]["A1"].font = Font(bold=True, color="1E3A5F")

    risks = wb["Risks_Actions"]
    risks.append(["Risk", "Severity", "Mitigation"])
    style_header_row(risks, 1, 3)
    risks.append(["FX volatility on imports", "Medium", "Hedge 60% exposure"])
    risks.append(["Working capital stretch", "High", "Accelerate collections in top 5 debtors"])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_scenario_workbook(file_bytes: bytes, body: dict[str, Any]) -> bytes:
    base_rev = float(body.get("base_revenue_growth_pct", 10))
    bull_rev = float(body.get("bull_revenue_growth_pct", 20))
    bear_rev = float(body.get("bear_revenue_growth_pct", 0))
    base_cost = float(body.get("base_cost_inflation_pct", 5))
    bull_cost = float(body.get("bull_cost_inflation_pct", 3))
    bear_cost = float(body.get("bear_cost_inflation_pct", 8))

    df = _read_first_sheet(file_bytes)
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Base_Case"
    ws0.append(["Account", "Base amount"])
    style_header_row(ws0, 1, 2)
    for _, row in df.head(40).iterrows():
        v0 = _parse_num(row.iloc[1]) if len(row) > 1 else 0
        ws0.append([str(row.iloc[0]), v0])
    bull = wb.create_sheet("Bull_Case")
    bear = wb.create_sheet("Bear_Case")
    bull.append(["Account", "Bull amount"])
    bear.append(["Account", "Bear amount"])
    style_header_row(bull, 1, 2)
    style_header_row(bear, 1, 2)
    for _, row in df.head(40).iterrows():
        base_amt = _parse_num(row.iloc[1]) if len(row) > 1 else 0
        nm = str(row.iloc[0])
        bull.append([nm, base_amt * (1 + (bull_rev - base_rev) / 100.0)])
        bear.append([nm, base_amt * (1 + (bear_rev - base_rev) / 100.0)])
    wb.create_sheet("Scenario_Comparison")
    wb.create_sheet("Sensitivity_Table")
    wb.create_sheet("AI_Recommendation")

    # Simple comparison sheet
    sc = wb["Scenario_Comparison"]
    sc.append(["Metric", "Base", "Bull", "Bear"])
    style_header_row(sc, 1, 4)
    sc.append(["Revenue growth %", base_rev / 100, bull_rev / 100, bear_rev / 100])
    sc.append(["Cost inflation %", base_cost / 100, bull_cost / 100, bear_cost / 100])
    for r in range(2, sc.max_row + 1):
        for c in (2, 3, 4):
            sc.cell(row=r, column=c).number_format = "0.0%"

    sens = wb["Sensitivity_Table"]
    sens.append(["Revenue shock %", "EBIT impact (illustrative)"])
    style_header_row(sens, 1, 2)
    for pct in range(-20, 25, 5):
        sens.append([pct / 100.0, pct * 50_000])
    for r in range(2, sens.max_row + 1):
        sens.cell(row=r, column=1).number_format = "0.0%"
        sens.cell(row=r, column=2).number_format = INR_NUM_FMT

    ai = wb["AI_Recommendation"]
    if llm_service.is_configured():
        try:
            raw = llm_service.invoke(
                f"Scenario probabilities: Base {base_rev}% rev growth, Bull {bull_rev}%, Bear {bear_rev}%. "
                "Return JSON: {{\"text\": \"4 sentences with probabilities summing ~100% and recommendation.\"}}",
                max_tokens=500,
            )
            m = re.search(r"\{[\s\S]*\}", raw)
            t = json.loads(m.group()).get("text", raw) if m else raw
        except Exception:
            t = "AI recommendation unavailable."
    else:
        t = "Set API key for AI scenario narrative."
    ai["A1"] = t
    ai["A1"].alignment = Alignment(wrap_text=True)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_management_accounts_workbook(file_bytes: bytes, format_id: str = "ICAI") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Management_Accounts"
    ws["A1"] = f"Management Accounts — {format_id} style (FinReportAI)"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws.append([])
    ws.append(["Section", "This month", "YTD"])
    style_header_row(ws, 3, 3)
    try:
        df = _read_first_sheet(file_bytes)
        for i in range(min(25, len(df))):
            v = _parse_num(df.iloc[i, 1]) if df.shape[1] > 1 else 0
            ws.append([str(df.iloc[i, 0]), v, v * (i + 1) / 12])
    except Exception:
        ws.append(["(Upload TB)", 0, 0])
    for r in range(4, ws.max_row + 1):
        ws.cell(row=r, column=2).number_format = INR_NUM_FMT
        ws.cell(row=r, column=3).number_format = INR_NUM_FMT

    wd = wb.create_sheet("Departmental_PL")
    wd.append(["Department", "This month", "YTD"])
    style_header_row(wd, 1, 3)
    wd.append(["Operations", 1200000, 8400000])
    wd.append(["Sales", 400000, 2800000])

    wb.create_sheet("Bridge_Analysis")
    wb["Bridge_Analysis"]["A1"] = "Waterfall / bridge — link charts to Management_Accounts in Excel."

    notes = wb.create_sheet("Notes_AI")
    if llm_service.is_configured():
        try:
            raw = llm_service.invoke(
                f"Write short 'Notes to Management Accounts' (3 bullets) for format {format_id}. JSON: {{\"notes\":[]}}",
                max_tokens=400,
            )
            m = re.search(r"\{[\s\S]*\}", raw)
            bullets = json.loads(m.group()).get("notes", []) if m else []
        except Exception:
            bullets = []
    else:
        bullets = ["Enable AI for auto-generated notes."]
    for i, b in enumerate(bullets, start=1):
        notes.cell(row=i, column=1, value=f"• {b}")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
