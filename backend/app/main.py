import asyncio
import io
import json
import logging
import os
from typing import Any
import tempfile
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv

logger = logging.getLogger(__name__)

# Load .env with override=True so a fresh start always picks up the latest keys.
# Also try explicit path so it works regardless of which directory uvicorn was launched from.
_env_file = Path(__file__).resolve().parent.parent / ".env"  # backend/.env
load_dotenv(dotenv_path=_env_file if _env_file.exists() else None, override=True)

from fastapi import BackgroundTasks, Body, Depends, FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import inspect, text
from app.core.config import settings
from app.core.mcp_auth_middleware import add_mcp_api_key_middleware
from app.middleware.product_role_middleware import ProductRoleMiddleware
from app.core.database import get_db
from app.routers import consolidation as consolidation_router
from app.routers import month_end_close as month_end_close_router
from app.routers import earnings_review as earnings_review_router
from app.routers import gl_reconciliation as gl_reconciliation_router
from app.routers import model_builder as model_builder_router
from app.routers import auth as rbac_auth_router
from app.routers import users as rbac_users_router
from app.routers import gulftax_team as gulftax_team_router
from app.routers import workspaces as workspaces_router
from app.routers import company_setup as company_setup_router
from app.routers import ap_settings as ap_settings_router
from app.api.routes import (
    fpa_master_upload,
    upload_routes,
    auth_routes,
    cfo_dashboard,
    ifrs_statements,
    ifrs_week1,
    ifrs_agentic,
    nova,
    fpa_variance,
    fpa_pvm,
    fpa_three_statement,
    fpa_monte_carlo,
    fpa_arr,
    fpa_headcount,
    fpa_sensitivity,
    reports_board_pack,
    r2r_history,
    stateful_journal,
    bank_recon,
    erp_integration,
    erp_connections,
    ap_integrations,
    bookkeeping,
    r2r_pattern,
    r2r_learning_routes,
    board_pack_routes,
    excel_suite,
    excel_addon_routes,
    voice_inbound,
    chat,
    cfo_agents,
    rev_rec_recon,
    audit_intelligence,
    history_router,
    historical_analysis,
    entity_health,
    payment_calendar,
    covenant_tracker,
    ar_collections,
    ca_bank_router,
    tally_push,
    ifrs_export,
    ifrs16_routes,
    uae_accounting,
    uae_full_routes,
    uae_ar_routes,
    uae_controls_routes,
    india_routes,
    pipeline as pipeline_router,
    o2c_routes,
    crm_routes,
    training_routes,
    agent_extract,
    ap_anomaly,
    ap_insights,
    ap_aging,
    audit_log_routes,
    cit_return,
    gl_summary,
    notifications_routes,
    uae_account_classification,
    uae_fs_routes,
    uae_journals_routes,
    journal_entries,
    analytics,
    ap_invoices_rds,
    ap_companies_rds,
    vat_advanced_rds,
    system_routes,
)
from app.modules.ifrs9.router import router as ifrs9_router
from app.middleware.request_logging import RequestLoggingMiddleware
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.util import get_remote_address
from app.db import init_db
from app.agents.intelligence import generate_board_pack_content
from app.board_pack_generator import generate_pdf
from app.scheduler import run_daily_watchdog, scheduler, setup_scheduler
from app.agents.command_center.registry import list_agent_names
from app.services.cfo_orchestrator_service import create_queued_run, execute_cfo_agent_task
from excel_addin import analyze_service, chat_layer, intent_layer, legacy_shim

_BOARD_PACK_DIR = Path(__file__).resolve().parent.parent / "board_packs"
os.makedirs(_BOARD_PACK_DIR, exist_ok=True)


def _init_db_safe() -> None:
    try:
        init_db()
    except Exception:
        logger.exception("init_db failed — API will run but DB-backed routes may error until fixed")


def _is_sqlite(db) -> bool:
    try:
        return db.bind is not None and db.bind.dialect.name == "sqlite"
    except Exception:
        return False


def _has_agentic_runs_schema(db) -> bool:
    try:
        cols = {c.get("name") for c in inspect(db.bind).get_columns("agent_runs")}
        required = {"agent_name", "insight", "urgency", "created_at"}
        return required.issubset(cols)
    except Exception:
        return False


@asynccontextmanager
async def lifespan(app: FastAPI):
    # Run DB setup in a thread so startup does not block the server from accepting
    # connections (slow/unreachable DB or SQLite on a synced drive otherwise looks like a hung tab).
    asyncio.create_task(asyncio.to_thread(_init_db_safe))
    if settings.ENABLE_CFO_SCHEDULER:
        setup_scheduler()
        if not scheduler.running:
            scheduler.start()
            print("Agentic scheduler started")
            logger.info("Agentic scheduler started")
        asyncio.create_task(run_daily_watchdog())
    yield


app = FastAPI(
    title=settings.APP_NAME,
    version=settings.VERSION,
    lifespan=lifespan,
)

limiter = Limiter(key_func=get_remote_address, default_limits=["100/minute"])
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
app.add_middleware(RequestLoggingMiddleware)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://finreportai.com",
        "https://www.finreportai.com",
        "https://finreportaicommercial.vercel.app",
        "http://localhost:5173",
        "http://localhost:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(ProductRoleMiddleware)
add_mcp_api_key_middleware(app, settings.CLIENT_API_KEY)


# Routes
app.include_router(auth_routes.router)
app.include_router(fpa_master_upload.router)
app.include_router(upload_routes.router)
app.include_router(cfo_dashboard.router)
app.include_router(ifrs_statements.router)
app.include_router(ifrs_week1.router, prefix="/api/ifrs")
app.include_router(ifrs_agentic.router)
app.include_router(nova.router, prefix="/api")
app.include_router(chat.router, prefix="/api")
app.include_router(cfo_agents.agents_router)
app.include_router(cfo_agents.briefing_router)
app.include_router(fpa_variance.router)
app.include_router(fpa_pvm.router, prefix="/api/fpa")
app.include_router(fpa_three_statement.router, prefix="/api/fpa")
app.include_router(fpa_monte_carlo.router, prefix="/api/fpa")
app.include_router(fpa_arr.router, prefix="/api/fpa")
app.include_router(fpa_headcount.router, prefix="/api/fpa")
app.include_router(fpa_sensitivity.router, prefix="/api/fpa")
app.include_router(reports_board_pack.router)
app.include_router(r2r_history.router)
app.include_router(stateful_journal.router)
app.include_router(bank_recon.router)
app.include_router(erp_integration.router, prefix="/api")
app.include_router(erp_connections.router)
app.include_router(bookkeeping.router)
app.include_router(r2r_pattern.router)
app.include_router(r2r_learning_routes.router)
app.include_router(board_pack_routes.router, prefix="/api/board-pack")
app.include_router(excel_suite.router, prefix="/api/excel")
app.include_router(excel_addon_routes.router)
app.include_router(voice_inbound.router)
app.include_router(rev_rec_recon.router)
app.include_router(audit_intelligence.router)
app.include_router(history_router.router, prefix="/api/v2", tags=["History"])
app.include_router(historical_analysis.router, prefix="/api/v2", tags=["History"])
app.include_router(month_end_close_router.router)
app.include_router(earnings_review_router.router)
app.include_router(gl_reconciliation_router.router)
app.include_router(model_builder_router.router)
app.include_router(rbac_auth_router.router)
app.include_router(rbac_users_router.router)
app.include_router(gulftax_team_router.router)
app.include_router(workspaces_router.router)
app.include_router(company_setup_router.router)
app.include_router(ap_settings_router.router)
app.include_router(ap_integrations.router)
app.include_router(o2c_routes.router)
app.include_router(crm_routes.router)
app.include_router(training_routes.router)
app.include_router(entity_health.router)
app.include_router(payment_calendar.router)
app.include_router(covenant_tracker.router)
app.include_router(ar_collections.router)
app.include_router(ca_bank_router.router)
app.include_router(tally_push.router)
app.include_router(ifrs_export.router, prefix="/api/ifrs")
app.include_router(ifrs16_routes.router)
app.include_router(uae_accounting.router)
app.include_router(uae_full_routes.router)
app.include_router(uae_full_routes.fx_router)
app.include_router(uae_ar_routes.router)
app.include_router(uae_controls_routes.router)
app.include_router(india_routes.router)
app.include_router(pipeline_router.router)
app.include_router(agent_extract.router)
app.include_router(ap_anomaly.router)
app.include_router(ap_insights.router)
app.include_router(ap_aging.router)
app.include_router(audit_log_routes.router)
app.include_router(cit_return.router)
app.include_router(gl_summary.router)
app.include_router(notifications_routes.router)
app.include_router(uae_account_classification.router)
app.include_router(uae_fs_routes.router)
app.include_router(uae_journals_routes.router)
app.include_router(journal_entries.router)
app.include_router(analytics.router)
app.include_router(consolidation_router.router)
app.include_router(ifrs9_router)
app.include_router(ap_invoices_rds.router)
app.include_router(ap_companies_rds.router)
app.include_router(vat_advanced_rds.router)
app.include_router(system_routes.router)

# GulfTax AI — embedded (replaces external localhost:8000 service)
from app.modules.gulftax.router import router as gulftax_router
from app.modules.gulftax.gulftax_einvoicing import router as gulftax_einvoicing_router
from app.modules.gulftax.ported_mount import register_gulftax_ported_routers

app.include_router(gulftax_router)
app.include_router(gulftax_einvoicing_router)
register_gulftax_ported_routers(app)

if settings.ENABLE_FASTAPI_MCP:
    try:
        from fastapi_mcp import FastApiMCP

        FastApiMCP(app).mount()
    except ImportError:
        logger.warning(
            "fastapi-mcp is not installed; MCP is disabled (pip install fastapi-mcp)."
        )
    except Exception as e:
        # Do not fail the whole API if MCP mount breaks in production.
        logger.warning("fastapi-mcp mount skipped: %s", e)

@app.get("/")
async def root():
    # Local: return HTML so embedded browsers (e.g. VS Code Simple Browser) show content; redirects there are often blank.
    if settings.DEBUG and not os.environ.get("RAILWAY_ENVIRONMENT"):
        return HTMLResponse(
            "<!DOCTYPE html><html lang='en'><head><meta charset='utf-8'/>"
            f"<title>{settings.APP_NAME} API</title></head>"
            "<body style='font-family:system-ui,sans-serif;max-width:42rem;margin:2rem;line-height:1.5'>"
            f"<h1 style='font-size:1.25rem'>{settings.APP_NAME} — API</h1>"
            "<p>This port is the FastAPI backend. The React UI is started with Vite, usually at "
            "<a href='http://localhost:3006'>http://localhost:3006</a>.</p>"
            "<ul>"
            "<li><a href='/docs'>Swagger UI</a> (<code>/docs</code>)</li>"
            "<li><a href='/redoc'>ReDoc</a> (<code>/redoc</code>)</li>"
            "<li><a href='/health'>Health</a> (<code>/health</code>)</li>"
            "</ul></body></html>"
        )
    return {
        "app": settings.APP_NAME,
        "version": settings.VERSION,
        "status": "running",
    }

@app.get("/health")
async def health():
    import os as _os
    from pathlib import Path as _Path
    _k = _os.environ.get("ANTHROPIC_API_KEY", "")
    _env = _Path(__file__).resolve().parent.parent / ".env"
    return {
        "status": "healthy",
        "ai_key_set": bool(_k),
        "ai_key_prefix": _k[:12] if _k else "EMPTY",
        "cwd": _os.getcwd(),
        "env_file_path": str(_env),
        "env_file_exists": _env.exists(),
        "file": __file__,
    }


@app.get("/api/test-aws", tags=["AWS"])
async def test_aws():
    """Test connectivity to both S3 buckets (UAE + India)."""
    from app.core.aws_config import test_aws_connection
    return test_aws_connection()


@app.post("/api/setup-aws", tags=["AWS"])
async def setup_aws():
    """Create standard folder structure in both S3 buckets (idempotent)."""
    from app.core.aws_config import create_bucket_folders
    return create_bucket_folders()


class RunAgentCompatBody(BaseModel):
    agent_name: str = Field(..., min_length=1)
    test_mode: bool = False
    context: dict = Field(default_factory=dict)


@app.post("/api/agents/run")
async def run_agent_compat(
    body: RunAgentCompatBody,
    background_tasks: BackgroundTasks,
    db=Depends(get_db),
):
    """
    Compatibility endpoint for one-body agent trigger style.
    """
    agent_name = body.agent_name.strip().lower()
    valid_agents = list_agent_names()
    if agent_name not in valid_agents:
        raise HTTPException(status_code=400, detail=f"Unknown agent. Valid: {', '.join(valid_agents)}")
    row = create_queued_run(db, "default", agent_name, body.context or {})
    background_tasks.add_task(execute_cfo_agent_task, row.id)
    return {"cfo_run_id": row.run_id, "id": row.id, "agent": agent_name, "status": "queued"}


@app.get("/api/agents/latest-brief")
async def get_latest_brief(db=Depends(get_db)):
    """
    Called automatically on CFO login.
    Returns today's insights with 3-tier urgency buckets.
    """
    try:
        if not _has_agentic_runs_schema(db):
            return {
                "date": datetime.now().isoformat(),
                "alerts": {"red": [], "yellow": [], "green": []},
                "total_agents_run": 0,
                "last_run": None,
            }

        if _is_sqlite(db):
            rows = db.execute(
                text(
                    """
                    SELECT ar.agent_name, ar.insight, ar.urgency, ar.created_at
                    FROM agent_runs ar
                    INNER JOIN (
                      SELECT agent_name, MAX(created_at) AS max_created_at
                      FROM agent_runs
                      WHERE created_at > datetime('now', '-1 day')
                      GROUP BY agent_name
                    ) latest
                      ON ar.agent_name = latest.agent_name
                     AND ar.created_at = latest.max_created_at
                    ORDER BY ar.created_at DESC
                    """
                )
            ).fetchall()
        else:
            rows = db.execute(
                text(
                    """
                    SELECT DISTINCT ON (agent_name)
                      agent_name, insight, urgency, created_at
                    FROM agent_runs
                    WHERE created_at > NOW() - INTERVAL '24 hours'
                    ORDER BY agent_name, created_at DESC
                    """
                )
            ).fetchall()

        alerts: dict[str, list[dict]] = {"red": [], "yellow": [], "green": []}
        for row in rows:
            item = dict(row._mapping)
            insight = item.get("insight")
            if isinstance(insight, str):
                try:
                    insight = json.loads(insight)
                except json.JSONDecodeError:
                    insight = {"what_happened": insight}
            urgency = (item.get("urgency") or "green").lower()
            if urgency not in alerts:
                urgency = "green"
            alerts[urgency].append(
                {
                    "agent": item.get("agent_name"),
                    "insight": insight,
                    "time": str(item.get("created_at")),
                }
            )

        return {
            "date": datetime.now().isoformat(),
            "alerts": alerts,
            "total_agents_run": len(rows),
            "last_run": str(rows[0]._mapping["created_at"]) if rows else None,
        }
    except Exception as exc:
        return {
            "date": datetime.now().isoformat(),
            "alerts": {"red": [], "yellow": [], "green": []},
            "total_agents_run": 0,
            "error": str(exc),
        }


@app.post("/api/board-pack/generate")
async def generate_board_pack(db=Depends(get_db)):
    """
    Generate board pack with streaming progress updates.
    """

    async def stream_progress():
        steps = [
            "Collecting variance data",
            "Running forecast analysis",
            "Analysing cash position",
            "Generating AI commentary",
            "Compiling board pack",
            "Generating PDF",
        ]

        for step in steps:
            yield f"data: {json.dumps({'step': step, 'status': 'running'})}\n\n"
            await asyncio.sleep(1)

        rows = []
        if _has_agentic_runs_schema(db):
            if _is_sqlite(db):
                rows = db.execute(
                    text(
                        """
                        SELECT ar.agent_name, ar.insight
                        FROM agent_runs ar
                        INNER JOIN (
                          SELECT agent_name, MAX(created_at) AS max_created_at
                          FROM agent_runs
                          WHERE created_at > datetime('now', '-30 days')
                          GROUP BY agent_name
                        ) latest
                          ON ar.agent_name = latest.agent_name
                         AND ar.created_at = latest.max_created_at
                        ORDER BY ar.created_at DESC
                        """
                    )
                ).fetchall()
            else:
                rows = db.execute(
                    text(
                        """
                        SELECT DISTINCT ON (agent_name)
                          agent_name, insight
                        FROM agent_runs
                        WHERE created_at > NOW() - INTERVAL '30 days'
                        ORDER BY agent_name, created_at DESC
                        """
                    )
                ).fetchall()
        all_results = {}
        for row in rows:
            item = dict(row._mapping)
            insight = item.get("insight")
            if isinstance(insight, str):
                try:
                    insight = json.loads(insight)
                except json.JSONDecodeError:
                    insight = {"note": insight}
            all_results[item.get("agent_name", "unknown")] = insight

        content = await generate_board_pack_content(all_results)
        pdf_path = await generate_pdf(content, f"board_pack_{datetime.now().strftime('%Y_%m')}.pdf")
        yield f"data: {json.dumps({'step': 'complete', 'pdf_path': pdf_path, 'content': content}, default=str)}\n\n"

    return StreamingResponse(stream_progress(), media_type="text/event-stream")


@app.get("/api/board-pack/download-file/{filename}")
async def download_board_pack(filename: str):
    """Download generated board pack PDF."""
    path = os.path.join(tempfile.gettempdir(), filename)
    if os.path.exists(path):
        return FileResponse(path, media_type="application/pdf", filename=filename)
    return {"error": "Board pack not found"}


_EXCEL_ADDIN_CORS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "POST, OPTIONS",
    "Access-Control-Allow-Headers": "*",
}


class ExcelAddinAnalyzeBody(BaseModel):
    """Power Automate / Excel Controls sheet → one endpoint."""

    analysis_type: str = Field(..., min_length=1)
    company: str = Field(..., min_length=1)
    period: str = ""
    currency: str = "USD"
    thresholds: dict[str, Any] = Field(default_factory=dict)
    filters: dict[str, Any] = Field(default_factory=dict)
    rows: list[dict[str, Any]] = Field(default_factory=list)


class ExcelAddinAnalyzeResponse(BaseModel):
    """OpenAPI shape for /api/excel-addin/analyze (ml_engine + narrative vary by analysis_type)."""

    model_config = ConfigDict(extra="ignore")

    analysis_type: str
    company: str
    period: str
    currency: str
    thresholds: dict[str, Any] = Field(default_factory=dict)
    filters: dict[str, Any] = Field(default_factory=dict)
    ml_engine: dict[str, Any] = Field(default_factory=dict)
    narrative: dict[str, Any] = Field(default_factory=dict)
    claude_error: str | None = None


class ExcelAddinExportBody(BaseModel):
    """Build one workbook from all 6 Excel add-in module payloads."""

    company: str = Field("Unknown")
    period: str = Field("")
    currency: str = Field("USD")
    variance: dict[str, Any] = Field(default_factory=dict)
    budget: dict[str, Any] = Field(default_factory=dict)
    kpi: dict[str, Any] = Field(default_factory=dict)
    forecast: dict[str, Any] = Field(default_factory=dict)
    scenarios: dict[str, Any] = Field(default_factory=dict)
    reports: dict[str, Any] = Field(default_factory=dict)


_EXCEL_ADDIN_OPENAPI_TAGS = ["Excel Add-in", "FP&A"]

_EXCEL_ANALYZE_OPENAPI_EXAMPLES: dict[str, dict] = {
    "variance": {
        "summary": "Variance (BvA)",
        "description": "Budget vs actual with optional threshold % and department filter.",
        "value": {
            "analysis_type": "variance",
            "company": "Test Co",
            "period": "Q1 2026",
            "currency": "INR",
            "thresholds": {"variance_pct": 5},
            "filters": {"department": "All"},
            "rows": [
                {"account": "Revenue", "budget": 4000000, "actual": 3650000},
                {"account": "Salaries", "budget": 1200000, "actual": 1380000},
            ],
        },
    },
    "forecast": {
        "summary": "Forecast",
        "description": "Monthly series → linear extension (needs ≥2 months).",
        "value": {
            "analysis_type": "forecast",
            "company": "Test Co",
            "period": "",
            "currency": "USD",
            "thresholds": {},
            "filters": {},
            "rows": [
                {"month": "2025-10", "revenue": 800, "costs": 520},
                {"month": "2025-11", "revenue": 830, "costs": 540},
                {"month": "2025-12", "revenue": 900, "costs": 560},
            ],
        },
    },
}


@app.options("/api/excel-addin/analyze", tags=_EXCEL_ADDIN_OPENAPI_TAGS)
async def excel_addin_analyze_options():
    """CORS preflight for browser / Office clients."""
    return Response(status_code=204, headers=dict(_EXCEL_ADDIN_CORS))


@app.post(
    "/api/excel-addin/analyze",
    tags=_EXCEL_ADDIN_OPENAPI_TAGS,
    summary="Unified FP&A analyze (Power Automate / Controls sheet)",
    response_model=ExcelAddinAnalyzeResponse,
)
async def excel_addin_analyze_post(
    body: ExcelAddinAnalyzeBody = Body(
        ...,
        openapi_examples=_EXCEL_ANALYZE_OPENAPI_EXAMPLES,
    ),
):
    """
    Single dynamic FP&A endpoint: analysis_type + thresholds + filters + rows.
    Legacy /variance, /budget, /kpi, /forecast, /scenarios, /reports still work
    but return header X-Deprecated-Use-Instead: /api/excel-addin/analyze.
    """
    try:
        out = analyze_service.run(body.model_dump())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if isinstance(out.get("narrative"), dict) and out["narrative"].get("error"):
        out["claude_error"] = out["narrative"]["error"]

    return JSONResponse(content=out, headers=dict(_EXCEL_ADDIN_CORS))


_DEPRECATION_HEADERS = {
    **_EXCEL_ADDIN_CORS,
    "X-Deprecated-Use-Instead": "/api/excel-addin/analyze",
}


def _legacy_json_response(payload: dict) -> JSONResponse:
    return JSONResponse(content=payload, headers=dict(_DEPRECATION_HEADERS))


def _to_num(v: Any) -> float:
    try:
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        return float(str(v).replace(",", "").strip())
    except Exception:
        return 0.0


def _title(ws, text: str) -> None:
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0F2D5E")
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Generated: {datetime.utcnow().isoformat()}Z"
    ws["A2"].font = Font(size=10, color="334155")


def _header_row(ws, row: int, labels: list[str]) -> None:
    for i, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E3A8A")
        c.alignment = Alignment(horizontal="center")


def _status_fill(flag: str) -> PatternFill:
    f = str(flag or "").upper()
    if f in {"FAV", "GREEN", "GOOD"}:
        return PatternFill("solid", fgColor="DCFCE7")
    if f in {"ADV", "RED", "CRITICAL"}:
        return PatternFill("solid", fgColor="FEE2E2")
    return PatternFill("solid", fgColor="FEF9C3")


def _build_executive_summary_sheet(wb: Workbook, payloads: dict[str, dict[str, Any]], results: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    _title(ws, "Executive Summary")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    kpi = payloads.get("kpi", {}).get("actuals", {}) if isinstance(payloads.get("kpi"), dict) else {}
    rev = _to_num(kpi.get("revenue"))
    cogs = _to_num(kpi.get("cogs"))
    opex = _to_num(kpi.get("opex"))
    ebitda = rev - cogs - opex
    net_profit = ebitda
    health_score = _to_num(results.get("reports", {}).get("health_score"))
    gross_margin = ((rev - cogs) / rev * 100.0) if rev else 0.0
    net_margin = (net_profit / rev * 100.0) if rev else 0.0

    _header_row(ws, 4, ["KPI", "Value", "Status"])
    kpi_rows = [
        ("Revenue", rev, "GOOD" if rev > 0 else "WATCH"),
        ("Gross Margin %", gross_margin, "GOOD" if gross_margin >= 50 else "WATCH"),
        ("EBITDA", ebitda, "GOOD" if ebitda >= 0 else "CRITICAL"),
        ("Net Margin %", net_margin, "GOOD" if net_margin >= 10 else "WATCH"),
        ("Health Score", health_score, "GOOD" if health_score >= 70 else "WATCH"),
    ]
    r = 5
    for name, value, status in kpi_rows:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=status)
        ws.cell(row=r, column=3).fill = _status_fill(status)
        r += 1

    _header_row(ws, 12, ["P&L Line", "Actual", "Budget", "Variance", "Variance %", "Flag"])
    variance_rows = payloads.get("variance", {}).get("rows", []) if isinstance(payloads.get("variance"), dict) else []
    if not isinstance(variance_rows, list):
        variance_rows = []
    rr = 13
    for item in variance_rows:
        if not isinstance(item, dict):
            continue
        account = str(item.get("account", "Line Item"))
        actual = _to_num(item.get("actual"))
        budget = _to_num(item.get("budget"))
        variance = actual - budget
        v_pct = ((variance / budget) * 100.0) if budget else 0.0
        flag = "FAV" if variance >= 0 else "ADV"
        ws.cell(row=rr, column=1, value=account)
        ws.cell(row=rr, column=2, value=actual)
        ws.cell(row=rr, column=3, value=budget)
        ws.cell(row=rr, column=4, value=variance)
        ws.cell(row=rr, column=5, value=v_pct / 100.0)
        ws.cell(row=rr, column=5).number_format = "0.0%"
        ws.cell(row=rr, column=6, value=flag)
        ws.cell(row=rr, column=6).fill = _status_fill(flag)
        rr += 1


def _build_variance_sheet(wb: Workbook, payloads: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Variance Analysis")
    ws.sheet_view.showGridLines = False
    _title(ws, "Variance Analysis")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    _header_row(ws, 4, ["Category", "Budget", "Actual", "Variance", "Variance %", "Flag"])
    rows = payloads.get("variance", {}).get("rows", []) if isinstance(payloads.get("variance"), dict) else []
    if not isinstance(rows, list):
        rows = []
    r = 5
    for item in rows:
        if not isinstance(item, dict):
            continue
        name = str(item.get("account", "Line Item"))
        budget = _to_num(item.get("budget"))
        actual = _to_num(item.get("actual"))
        variance = actual - budget
        v_pct = (variance / budget * 100.0) if budget else 0.0
        flag = "FAV" if variance >= 0 else "ADV"
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=budget)
        ws.cell(row=r, column=3, value=actual)
        ws.cell(row=r, column=4, value=variance)
        ws.cell(row=r, column=5, value=v_pct / 100.0)
        ws.cell(row=r, column=5).number_format = "0.0%"
        ws.cell(row=r, column=6, value=flag)
        ws.cell(row=r, column=6).fill = _status_fill(flag)
        r += 1


def _build_kpi_dashboard_sheet(wb: Workbook, payloads: dict[str, dict[str, Any]], results: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("KPI Dashboard")
    ws.sheet_view.showGridLines = False
    _title(ws, "KPI Dashboard")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    kpi = payloads.get("kpi", {}).get("actuals", {}) if isinstance(payloads.get("kpi"), dict) else {}
    rev = _to_num(kpi.get("revenue"))
    cogs = _to_num(kpi.get("cogs"))
    opex = _to_num(kpi.get("opex"))
    cash = _to_num(kpi.get("cash"))
    ebitda = rev - cogs - opex
    gross_margin = ((rev - cogs) / rev * 100.0) if rev else 0.0
    ebitda_margin = (ebitda / rev * 100.0) if rev else 0.0
    net_margin = ebitda_margin
    health_score = _to_num(results.get("reports", {}).get("health_score"))
    ratios = [
        ("Revenue", rev),
        ("COGS", cogs),
        ("Gross Margin %", gross_margin),
        ("EBITDA", ebitda),
        ("EBITDA Margin %", ebitda_margin),
        ("Net Margin %", net_margin),
        ("Cash Position", cash),
        ("Health Score", health_score),
    ]
    _header_row(ws, 4, ["KPI", "Value", "Status"])
    r = 5
    for name, value in ratios:
        status = "GOOD"
        if "margin" in name.lower() and value < 0:
            status = "CRITICAL"
        if name == "Health Score" and value < 70:
            status = "WATCH"
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=status)
        ws.cell(row=r, column=3).fill = _status_fill(status)
        r += 1


def _build_balance_sheet(wb: Workbook, payloads: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    _title(ws, "Balance Sheet")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    kpi = payloads.get("kpi", {}).get("actuals", {}) if isinstance(payloads.get("kpi"), dict) else {}
    cash = _to_num(kpi.get("cash"))
    receivables = _to_num(kpi.get("receivables"))
    inventory = _to_num(kpi.get("inventory"))
    assets = cash + receivables + inventory
    liabilities = _to_num(kpi.get("liabilities"))
    equity = assets - liabilities
    _header_row(ws, 4, ["Section", "Amount"])
    rows = [
        ("Cash & Equivalents", cash),
        ("Accounts Receivable", receivables),
        ("Inventory", inventory),
        ("Total Assets", assets),
        ("Total Liabilities", liabilities),
        ("Equity", equity),
    ]
    r = 5
    for name, value in rows:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=value)
        if "Total" in name or name == "Equity":
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=2).font = Font(bold=True)
        r += 1


def _with_defaults(data: dict[str, Any], company: str, period: str, currency: str) -> dict[str, Any]:
    out = dict(data)
    out.setdefault("company", company)
    out.setdefault("period", period)
    out.setdefault("currency", currency)
    return out


@app.post(
    "/api/excel-addin/export",
    tags=_EXCEL_ADDIN_OPENAPI_TAGS,
    summary="Export all 6 Excel add-in analyses into one XLSX workbook",
)
async def excel_addin_export(body: ExcelAddinExportBody):
    module_builders: list[tuple[str, dict[str, Any], Any, Any]] = [
        ("variance", body.variance, legacy_shim.to_analyze_variance, legacy_shim.response_variance),
        ("budget", body.budget, legacy_shim.to_analyze_budget, legacy_shim.response_budget),
        ("kpi", body.kpi, legacy_shim.to_analyze_kpi, legacy_shim.response_kpi),
        ("forecast", body.forecast, legacy_shim.to_analyze_forecast, legacy_shim.response_forecast),
        ("scenarios", body.scenarios, legacy_shim.to_analyze_scenarios, legacy_shim.response_scenarios),
        ("reports", body.reports, legacy_shim.to_analyze_reports, legacy_shim.response_reports),
    ]

    payloads: dict[str, dict[str, Any]] = {}
    results: dict[str, dict[str, Any]] = {}
    for module_name, raw_data, to_analyze, to_legacy in module_builders:
        module_payload = _with_defaults(raw_data, body.company, body.period, body.currency)
        payloads[module_name] = module_payload
        try:
            analyze_payload = to_analyze(module_payload)
            out = analyze_service.run(analyze_payload)
            results[module_name] = to_legacy(out)
        except Exception as exc:  # noqa: BLE001
            results[module_name] = {"error": str(exc)}

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    _build_executive_summary_sheet(wb, payloads, results)
    _build_variance_sheet(wb, payloads)
    _build_kpi_dashboard_sheet(wb, payloads, results)
    _build_balance_sheet(wb, payloads)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Task 6 — save Excel report to S3
    try:
        from app.core.aws_config import upload_to_s3
        upload_to_s3(output.getvalue(), "finreportai_output.xlsx", folder="reports", country="UAE")
    except Exception:
        pass  # S3 save non-critical

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=finreportai_output.xlsx"},
    )


@app.api_route(
    "/api/excel-addin/variance",
    methods=["POST", "OPTIONS"],
    tags=_EXCEL_ADDIN_OPENAPI_TAGS,
    summary="[Deprecated] Variance — use /api/excel-addin/analyze",
)
async def excel_addin_variance_legacy(request: Request):
    """Deprecated: use POST /api/excel-addin/analyze with analysis_type=variance."""
    if request.method == "OPTIONS":
        return Response(status_code=204, headers=dict(_EXCEL_ADDIN_CORS))
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Request body must be valid JSON") from None
    try:
        inner = legacy_shim.to_analyze_variance(data)
        out = analyze_service.run(inner)
        legacy = legacy_shim.response_variance(out)
    except KeyError as exc:
        raise HTTPException(status_code=422, detail=f"Missing field: {exc}") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return _legacy_json_response(legacy)


@app.api_route(
    "/api/excel-addin/budget",
    methods=["POST", "OPTIONS"],
    tags=_EXCEL_ADDIN_OPENAPI_TAGS,
    summary="[Deprecated] Budget — use /api/excel-addin/analyze",
)
async def excel_addin_budget_legacy(request: Request):
    """Deprecated: use POST /api/excel-addin/analyze with analysis_type=budget."""
    if request.method == "OPTIONS":
        return Response(status_code=204, headers=dict(_EXCEL_ADDIN_CORS))
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Request body must be valid JSON") from None
    try:
        inner = legacy_shim.to_analyze_budget(data)
        out = analyze_service.run(inner)
        legacy = legacy_shim.response_budget(out)
    except KeyError as exc:
        raise HTTPException(status_code=422, detail=f"Missing field: {exc}") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return _legacy_json_response(legacy)


@app.api_route(
    "/api/excel-addin/kpi",
    methods=["POST", "OPTIONS"],
    tags=_EXCEL_ADDIN_OPENAPI_TAGS,
    summary="[Deprecated] KPI — use /api/excel-addin/analyze",
)
async def excel_addin_kpi_legacy(request: Request):
    """Deprecated: use POST /api/excel-addin/analyze with analysis_type=kpi."""
    if request.method == "OPTIONS":
        return Response(status_code=204, headers=dict(_EXCEL_ADDIN_CORS))
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Request body must be valid JSON") from None
    try:
        inner = legacy_shim.to_analyze_kpi(data)
        out = analyze_service.run(inner)
        legacy = legacy_shim.response_kpi(out)
    except KeyError as exc:
        raise HTTPException(status_code=422, detail=f"Missing field: {exc}") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return _legacy_json_response(legacy)


@app.api_route(
    "/api/excel-addin/forecast",
    methods=["POST", "OPTIONS"],
    tags=_EXCEL_ADDIN_OPENAPI_TAGS,
    summary="[Deprecated] Forecast — use /api/excel-addin/analyze",
)
async def excel_addin_forecast_legacy(request: Request):
    """Deprecated: use POST /api/excel-addin/analyze with analysis_type=forecast."""
    if request.method == "OPTIONS":
        return Response(status_code=204, headers=dict(_EXCEL_ADDIN_CORS))
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Request body must be valid JSON") from None
    try:
        inner = legacy_shim.to_analyze_forecast(data)
        out = analyze_service.run(inner)
        legacy = legacy_shim.response_forecast(out)
    except KeyError as exc:
        raise HTTPException(status_code=422, detail=f"Missing field: {exc}") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return _legacy_json_response(legacy)


@app.api_route(
    "/api/excel-addin/scenarios",
    methods=["POST", "OPTIONS"],
    tags=_EXCEL_ADDIN_OPENAPI_TAGS,
    summary="[Deprecated] Scenarios — use /api/excel-addin/analyze",
)
async def excel_addin_scenarios_legacy(request: Request):
    """Deprecated: use POST /api/excel-addin/analyze with analysis_type=scenario."""
    if request.method == "OPTIONS":
        return Response(status_code=204, headers=dict(_EXCEL_ADDIN_CORS))
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Request body must be valid JSON") from None
    try:
        inner = legacy_shim.to_analyze_scenarios(data)
        out = analyze_service.run(inner)
        legacy = legacy_shim.response_scenarios(out)
    except KeyError as exc:
        raise HTTPException(status_code=422, detail=f"Missing field: {exc}") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return _legacy_json_response(legacy)


@app.api_route(
    "/api/excel-addin/reports",
    methods=["POST", "OPTIONS"],
    tags=_EXCEL_ADDIN_OPENAPI_TAGS,
    summary="[Deprecated] Reports — use /api/excel-addin/analyze",
)
async def excel_addin_reports_legacy(request: Request):
    """Deprecated: use POST /api/excel-addin/analyze with analysis_type=report."""
    if request.method == "OPTIONS":
        return Response(status_code=204, headers=dict(_EXCEL_ADDIN_CORS))
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Request body must be valid JSON") from None
    try:
        inner = legacy_shim.to_analyze_reports(data)
        out = analyze_service.run(inner)
        legacy = legacy_shim.response_reports(out)
    except KeyError as exc:
        raise HTTPException(status_code=422, detail=f"Missing field: {exc}") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return _legacy_json_response(legacy)


class ExcelAddinChatTurn(BaseModel):
    role: str
    content: str


class ExcelAddinChatBody(BaseModel):
    message: str = Field(..., min_length=1)
    variance_context: dict | list | str = Field(default_factory=dict)
    chat_history: list[ExcelAddinChatTurn] = Field(default_factory=list)
    session_id: str = ""


@app.api_route(
    "/api/excel-addin/chat",
    methods=["POST", "OPTIONS"],
    tags=_EXCEL_ADDIN_OPENAPI_TAGS,
    summary="CFO chat (intent + context)",
)
async def excel_addin_chat(request: Request):
    """
    CFO decision-engine chat: intent detection + locked prompts + suggested actions.
    Client keeps chat_history; server does not persist (pass session_id for correlation only).
    """
    if request.method == "OPTIONS":
        return Response(status_code=204, headers=dict(_EXCEL_ADDIN_CORS))
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Request body must be valid JSON") from None
    try:
        body = ExcelAddinChatBody.model_validate(data)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    hist = [t.model_dump() for t in body.chat_history]
    intent_detected = intent_layer.detect_intent(body.message)
    result = chat_layer.handle_chat(
        body.message,
        intent_detected,
        body.variance_context,
        hist,
    )
    out: dict = {
        "reply": result["reply"],
        "intent_detected": intent_detected,
        "confidence": result["confidence"],
        "suggested_actions": chat_layer.suggested_actions(intent_detected),
        "session_id": body.session_id or "",
    }
    return JSONResponse(content=out, headers=dict(_EXCEL_ADDIN_CORS))


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "app.main:app",
        host="127.0.0.1",
        port=8000,
        reload=True,
        limit_concurrency=10,
        limit_max_requests=100,
        timeout_keep_alive=5,
    )

@app.get("/debug-env")
async def debug_env():
    import os
    from pathlib import Path
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    cwd = os.getcwd()
    env_path = Path(__file__).resolve().parent.parent / ".env"
    return {
        "key_set": bool(key),
        "key_prefix": key[:15] if key else "EMPTY",
        "cwd": cwd,
        "env_path": str(env_path),
        "env_exists": env_path.exists()
    }
