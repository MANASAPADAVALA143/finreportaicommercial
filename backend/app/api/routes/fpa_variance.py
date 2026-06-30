"""
FP&A Variance Analysis API — Budget vs Actual intelligence.
Endpoints: upload, calculate, template, ai-narrative, download-report.
"""
from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Any, Dict
import pandas as pd
import io
import re
import json
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from app.services import llm_service
from app.services.tb_variance_commentary import (
    build_tb_variance_system_prompt,
    build_variance_commentary_prompt,
)

router = APIRouter(prefix="/api/fpa/variance", tags=["FP&A Variance"])


# ---------- Request/Response models ----------

class LineItem(BaseModel):
    account: str
    department: str
    budget: float
    actual: float


class VarianceLineItem(LineItem):
    variance: float
    variance_pct: float
    status: str  # On Track | Watch | Over Budget | Under Budget
    material: bool  # >10%


class DepartmentSummary(BaseModel):
    department: str
    budget: float
    actual: float
    variance: float
    variance_pct: float
    status: str


class UploadResponse(BaseModel):
    line_items: List[Dict[str, Any]]
    departments: List[str]
    summary_stats: Dict[str, Any]
    format_detected: str


class CalculateResponse(BaseModel):
    line_items: List[Dict[str, Any]]
    department_summary: List[Dict[str, Any]]
    total_budget: float
    total_actual: float
    total_variance: float
    total_variance_pct: float
    overall_status: str


class AINarrativeResponse(BaseModel):
    executive_summary: str
    line_commentary: List[Dict[str, str]]  # [{ account, why, recommendation }]
    action_items: List[str]
    fallback_used: bool = False


class CommentaryRequest(BaseModel):
    variance_data: List[Dict[str, Any]]
    commentary_type: str = "cfo"  # executive | cfo | board | risk


class ExportExcelRequest(BaseModel):
    variance_data: List[Dict[str, Any]]


_FPA_CCY_SYMBOL: Dict[str, str] = {
    "INR": "₹",
    "USD": "$",
    "EUR": "€",
    "GBP": "£",
    "AED": "AED ",
}


def _fp_format_currency(amount: float, currency: str = "USD", fmt: str = "GLOBAL") -> str:
    """Match frontend varianceUtils.formatCurrency for LLM prompts (compact M/K or Cr/L)."""
    cy = (currency or "USD").upper().strip()
    loc = (fmt or "GLOBAL").upper().strip()
    if loc not in ("IN", "GLOBAL"):
        loc = "GLOBAL"
    sym = _FPA_CCY_SYMBOL.get(cy, f"{cy} ")
    abs_a = abs(float(amount))
    neg = float(amount) < 0
    prefix = "-" if neg else ""

    if loc == "IN" and cy == "INR":
        if abs_a >= 10_000_000:
            return f"{prefix}{sym}{abs_a / 10_000_000:.2f}Cr"
        if abs_a >= 100_000:
            return f"{prefix}{sym}{abs_a / 100_000:.2f}L"
        return f"{prefix}{sym}{abs_a:,.0f}"

    if loc == "IN":
        return f"{prefix}{sym}{abs_a:,.0f}"

    if abs_a >= 1_000_000:
        m = abs_a / 1_000_000
        decimals = 1 if m >= 100 else 2
        return f"{prefix}{sym}{m:.{decimals}f}M"
    if abs_a >= 100_000:
        k = abs_a / 1000
        return f"{prefix}{sym}{k:.1f}K"
    return f"{prefix}{sym}{abs_a:,.0f}"


def _parse_num(val: Any) -> float:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("₹", "").replace("$", "").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _detect_and_normalize(df: pd.DataFrame) -> tuple[List[Dict], str]:
    """Detect format A, B, or C and return list of { account, department, budget, actual }."""
    cols = [c.strip() for c in df.columns.astype(str)]
    col_lower = [c.lower() for c in cols]

    # Format A: Account | Department | Budget | Actual
    if any("budget" in c and "actual" in c for c in col_lower):
        pass  # might be Format B
    elif "budget" in " ".join(col_lower) and "actual" in " ".join(col_lower):
        acc_col = next((c for c in cols if "account" in c.lower() or "category" in c.lower() or "line" in c.lower()), cols[0])
        dept_col = next((c for c in cols if "department" in c.lower() or "dept" in c.lower()), None)
        budget_col = next((c for c in cols if ("budget" in c.lower() and "actual" not in c.lower()) or c.lower() == "budget"), None)
        actual_col = next((c for c in cols if ("actual" in c.lower() and "budget" not in c.lower()) or c.lower() == "actual"), None)
        if budget_col and actual_col:
            out = []
            for _, row in df.iterrows():
                budget = _parse_num(row.get(budget_col, 0))
                actual = _parse_num(row.get(actual_col, 0))
                if budget == 0 and actual == 0:
                    continue
                out.append({
                    "account": str(row.get(acc_col, "")).strip() or "Unnamed",
                    "department": str(row.get(dept_col, "")).strip() if dept_col else "All Depts",
                    "budget": budget,
                    "actual": actual,
                })
            return out, "A"

    # Format B: Jan_Budget, Jan_Actual, Feb_Budget, ...
    budget_cols = [c for c in cols if "budget" in c.lower() and "actual" not in c.lower()]
    actual_cols = [c for c in cols if "actual" in c.lower()]
    if budget_cols and actual_cols:
        acc_col = next((c for c in cols if "account" in c.lower() or "category" in c.lower()), cols[0])
        dept_col = next((c for c in cols if "department" in c.lower()), None)
        # Sum all periods
        out = []
        for _, row in df.iterrows():
            budget = sum(_parse_num(row.get(c, 0)) for c in budget_cols)
            actual = sum(_parse_num(row.get(c, 0)) for c in actual_cols)
            if budget == 0 and actual == 0:
                continue
            out.append({
                "account": str(row.get(acc_col, "")).strip() or "Unnamed",
                "department": str(row.get(dept_col, "")).strip() if dept_col else "All Depts",
                "budget": budget,
                "actual": actual,
            })
        return out, "B"

    # Format C: Account | Department | Period | Type (Budget/Actual) | Amount
    if "type" in " ".join(col_lower) and "amount" in " ".join(col_lower):
        acc_col = next((c for c in cols if "account" in c.lower()), cols[0])
        dept_col = next((c for c in cols if "department" in c.lower()), None)
        type_col = next((c for c in cols if "type" in c.lower()), None)
        amt_col = next((c for c in cols if "amount" in c.lower()), None)
        if type_col and amt_col:
            by_key: Dict[tuple, Dict] = {}
            for _, row in df.iterrows():
                key = (str(row.get(acc_col, "")).strip(), str(row.get(dept_col, "")).strip() if dept_col else "All Depts")
                if key not in by_key:
                    by_key[key] = {"account": key[0], "department": key[1], "budget": 0.0, "actual": 0.0}
                t = str(row.get(type_col, "")).lower()
                amt = _parse_num(row.get(amt_col, 0))
                if "budget" in t:
                    by_key[key]["budget"] += amt
                elif "actual" in t:
                    by_key[key]["actual"] += amt
            out = [v for v in by_key.values() if v["budget"] != 0 or v["actual"] != 0]
            return out, "C"

    # Fallback: try common column names
    acc_col = next((c for c in cols if "account" in c.lower() or "category" in c.lower() or "name" in c.lower()), cols[0])
    dept_col = next((c for c in cols if "department" in c.lower() or "dept" in c.lower()), None)
    for b in ["budget", "Budget", "Budget_Amount", "budget_amount"]:
        if b in cols:
            for a in ["actual", "Actual", "Actual_Amount", "actual_amount"]:
                if a in cols:
                    out = []
                    for _, row in df.iterrows():
                        budget = _parse_num(row.get(b, 0))
                        actual = _parse_num(row.get(a, 0))
                        if budget == 0 and actual == 0:
                            continue
                        out.append({
                            "account": str(row.get(acc_col, "")).strip() or "Unnamed",
                            "department": str(row.get(dept_col, "")).strip() if dept_col else "All Depts",
                            "budget": budget,
                            "actual": actual,
                        })
                    return out, "A"

    raise HTTPException(status_code=400, detail="Could not detect Excel/CSV format. Use columns: Account, Department, Budget, Actual (or Budget_Amount, Actual_Amount).")


def _status(variance_pct: float) -> str:
    if abs(variance_pct) < 5:
        return "On Track"
    if variance_pct > 10:
        return "Over Budget"
    if variance_pct < -10:
        return "Under Budget"
    return "Watch"


def _normalize_variance_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for r in rows or []:
        if bool(r.get("isHeader")):
            continue
        account = str(r.get("account") or r.get("category") or "").strip()
        if not account:
            continue
        budget = _parse_num(r.get("budget", 0))
        actual = _parse_num(r.get("actual", 0))
        variance = _parse_num(r.get("variance", actual - budget))
        variance_pct = _parse_num(r.get("variance_pct", r.get("variancePct", 0)))
        if variance_pct == 0 and budget != 0:
            variance_pct = (variance / budget) * 100
        out.append(
            {
                "account": account,
                "budget": budget,
                "actual": actual,
                "variance": variance,
                "variance_pct": variance_pct,
                "department": str(r.get("department") or "All Depts"),
            }
        )
    return out


def _build_commentary_prompt(commentary_type: str, variance_data: List[Dict[str, Any]]) -> str:
    variance_text = "\n".join(
        [
            f"{row['account']}: Budget {row['budget']}, Actual {row['actual']}, "
            f"Variance {row['variance']} ({row.get('variance_pct', 0):.1f}%)"
            for row in variance_data
        ]
    )
    prompts = {
        "executive": """
Write a 5-bullet executive summary of these FP&A results.
Lead with headline number. Include one risk, one positive.
Under 150 words. Plain English. No jargon.
""",
        "cfo": """
You are a CFO analyst. Write management commentary for these budget vs actual variances.
- Name top 3 revenue drivers (favourable variances)
- Name top 2 cost risks (adverse variances)
- Flag any items > 10% adverse variance
- Suggest one corrective action per risk found
Under 250 words. Confident, direct tone. No fluff.
""",
        "board": """
Turn these monthly FP&A results into a board narrative.
Structure:
1. Performance vs budget headline (1 sentence)
2. Key drivers — what went well (2-3 bullets)
3. Key risks — what needs attention (2-3 bullets)
4. Forward outlook (1 paragraph)
Board tone — strategic, not operational.
Under 300 words.
""",
        "risk": """
Review this variance data as a risk analyst.
Flag:
- Any variance > 15% adverse (Critical risk)
- Any variance 5-15% adverse (Watch item)
- Any unusual patterns (e.g. consistent adverse trend)
- Any items that could embarrass in a board meeting
Format as structured risk register with Risk | Severity | Recommended Action columns.
""",
    }
    selected = prompts.get((commentary_type or "cfo").lower(), prompts["cfo"])
    return f"{selected}\n\nVARIANCE DATA:\n{variance_text}"


def _calculate(line_items: List[Dict]) -> Dict:
    total_budget = sum(i["budget"] for i in line_items)
    total_actual = sum(i["actual"] for i in line_items)
    total_variance = total_actual - total_budget
    total_variance_pct = (total_variance / total_budget * 100) if total_budget else 0

    computed = []
    for i in line_items:
        b, a = i["budget"], i["actual"]
        var = a - b
        var_pct = (var / b * 100) if b else 0
        computed.append({
            **i,
            "variance": var,
            "variance_pct": var_pct,
            "status": _status(var_pct),
            "material": abs(var_pct) > 10,
        })

    # Department summary
    dept_agg: Dict[str, Dict] = {}
    for i in computed:
        d = i["department"]
        if d not in dept_agg:
            dept_agg[d] = {"department": d, "budget": 0, "actual": 0, "variance": 0}
        dept_agg[d]["budget"] += i["budget"]
        dept_agg[d]["actual"] += i["actual"]
        dept_agg[d]["variance"] += i["variance"]
    dept_list = []
    for d, v in dept_agg.items():
        vpct = (v["variance"] / v["budget"] * 100) if v["budget"] else 0
        dept_list.append({
            **v,
            "variance_pct": vpct,
            "status": _status(vpct),
        })

    return {
        "line_items": computed,
        "department_summary": dept_list,
        "total_budget": total_budget,
        "total_actual": total_actual,
        "total_variance": total_variance,
        "total_variance_pct": total_variance_pct,
        "overall_status": _status(total_variance_pct),
    }


@router.post("/upload", response_model=UploadResponse)
async def upload_variance(file: UploadFile = File(...)):
    """Accept Excel or CSV; auto-detect format (A, B, or C); normalise and return line_items, departments, summary_stats."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file")
    ext = file.filename.lower().split(".")[-1]
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, .csv allowed")

    contents = await file.read()
    try:
        from app.core.aws_config import upload_to_s3
        upload_to_s3(contents, file.filename, folder="uploads", country="UAE")
    except Exception:
        pass  # S3 save is non-critical — processing continues from memory
    try:
        if ext == "csv":
            df = pd.read_csv(io.StringIO(contents.decode("utf-8")))
        else:
            df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {e}")

    if df.empty or len(df.columns) < 2:
        raise HTTPException(status_code=400, detail="File has no data or too few columns")

    line_items, fmt = _detect_and_normalize(df)
    if not line_items:
        raise HTTPException(status_code=400, detail="No valid rows found. Check column names: Account, Department, Budget, Actual.")

    departments = list({i["department"] for i in line_items})
    calc = _calculate(line_items)
    summary_stats = {
        "total_budget": calc["total_budget"],
        "total_actual": calc["total_actual"],
        "total_variance": calc["total_variance"],
        "total_variance_pct": calc["total_variance_pct"],
        "line_count": len(line_items),
        "department_count": len(departments),
    }
    return UploadResponse(
        line_items=calc["line_items"],
        departments=departments,
        summary_stats=summary_stats,
        format_detected=fmt,
    )


class CalculateRequest(BaseModel):
    line_items: List[Dict[str, Any]]


@router.post("/calculate", response_model=CalculateResponse)
async def calculate_variance(body: CalculateRequest):
    """Compute variance ₹, %, status, materiality, department totals."""
    if not body.line_items:
        raise HTTPException(status_code=400, detail="line_items required")
    normalized = []
    for i in body.line_items:
        normalized.append({
            "account": str(i.get("account", "")),
            "department": str(i.get("department", "All Depts")),
            "budget": _parse_num(i.get("budget", 0)),
            "actual": _parse_num(i.get("actual", 0)),
        })
    result = _calculate(normalized)
    try:
        from app.agents.intelligence import generate_insight
        from app.agents.memory import read_agent_memory, store_agent_run, update_agent_memory
        from app.core.database import SessionLocal

        _db = SessionLocal()
        try:
            _history = await read_agent_memory("fpa_variance", _db)
            _insight = await generate_insight(
                "fpa_variance",
                {
                    "variance_result": result,
                    "source_route": "/fpa/variance",
                    "deep_link": "/fpa/variance",
                },
                _history,
            )
            _insight["source_route"] = "/fpa/variance"
            _insight["deep_link"] = "/fpa/variance"
            _insight["module_label"] = "FP&A Variance"
            _input = body.model_dump() if hasattr(body, "model_dump") else (body.dict() if hasattr(body, "dict") else {})
            await store_agent_run("fpa_variance", _input, result, _insight, _db)
            await update_agent_memory("fpa_variance", result, _db)
        finally:
            _db.close()
    except Exception as _e:
        print(f"[agent_run] fpa_variance: {_e}")
    return CalculateResponse(**result)


@router.get("/template")
async def get_variance_template():
    """Return downloadable Excel template with sample data and Instructions sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Variance Data"
    headers = ["Account_Name", "Department", "Budget_Amount", "Actual_Amount", "Notes"]
    sample = [
        ["Marketing", "Marketing Dept", 2300000, 2720000, "Campaign costs"],
        ["IT Infrastructure", "Technology", 1800000, 2016000, "Cloud migration"],
        ["Travel & Expenses", "All Depts", 680000, 530000, "Remote work"],
        ["Salaries & Wages", "All Depts", 8500000, 8500000, "On plan"],
        ["Office Rent", "Admin", 1200000, 1200000, "Fixed"],
        ["Training & Dev", "HR", 600000, 510000, "Online shift"],
        ["Legal & Compliance", "Finance", 450000, 423000, "Under retainer"],
        ["Sales Commissions", "Sales", 1100000, 1320000, "Over target"],
        ["Customer Support", "Operations", 920000, 875000, "Efficiency gain"],
        ["R&D Expenses", "Technology", 2200000, 2310000, "New hire"],
        ["Advertising", "Marketing", 850000, 940000, ""],
        ["Software Licenses", "Technology", 420000, 398000, ""],
        ["Recruitment", "HR", 380000, 510000, ""],
        ["Insurance", "Finance", 290000, 290000, ""],
        ["Utilities", "Admin", 180000, 162000, ""],
        ["Maintenance", "Operations", 240000, 228000, ""],
        ["Professional Fees", "Finance", 620000, 698000, ""],
        ["Depreciation", "Finance", 1100000, 1100000, ""],
        ["Interest Expense", "Finance", 340000, 325000, ""],
        ["Printing & Stationery", "Admin", 45000, 38000, ""],
        ["Telephone & Internet", "Admin", 120000, 114000, ""],
        ["Bank Charges", "Finance", 28000, 31000, ""],
        ["Event & Conferences", "Marketing", 350000, 420000, ""],
        ["Security Services", "Admin", 160000, 155000, ""],
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(sample, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    inst = wb.create_sheet("Instructions")
    inst["A1"] = "How to fill the Variance Data template"
    inst["A2"] = "1. Sheet 'Variance Data' must have columns: Account_Name, Department, Budget_Amount, Actual_Amount (Notes optional)."
    inst["A3"] = "2. You can use Format A: Account | Department | Budget | Actual"
    inst["A4"] = "3. Or period columns: Jan_Budget, Jan_Actual, Feb_Budget, Feb_Actual, etc."
    inst["A5"] = "4. Or long format: Account | Department | Period | Type (Budget/Actual) | Amount"
    inst["A6"] = "5. Upload your file in the Variance Analysis module; format is auto-detected."

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=FP&A_Variance_Template.xlsx"},
    )


class AINarrativeRequest(BaseModel):
    variance_analysis: Dict[str, Any]  # full object from calculate
    narrative_mode: str = "cfo"
    currency: str = "USD"
    currency_format: str = "GLOBAL"  # IN | GLOBAL — same as FP&A Settings


class TBLineCommentaryRequest(BaseModel):
    """Single-line TB YoY variance — industry-aware commentary via Claude."""

    account_name: str
    current: float
    prior: float
    variance: float
    variance_pct: Optional[float] = None
    currency: str = "INR"
    industry: str = "general"
    company_name: str = ""


class TBLineCommentaryResponse(BaseModel):
    commentary: str


@router.post("/tb-line-commentary", response_model=TBLineCommentaryResponse)
async def tb_line_commentary(body: TBLineCommentaryRequest):
    """Industry-aware trial balance variance commentary (TB Variance page)."""
    if not llm_service.is_configured():
        raise HTTPException(
            status_code=503,
            detail="LLM not configured: set ANTHROPIC_API_KEY on the server.",
        )
    prompt = build_variance_commentary_prompt(
        account_name=body.account_name,
        current=body.current,
        prior=body.prior,
        variance=body.variance,
        variance_pct=body.variance_pct,
        currency=body.currency or "INR",
        industry=body.industry or "general",
        company_name=body.company_name or "",
    )
    try:
        text = llm_service.invoke(
            prompt=prompt,
            system=build_tb_variance_system_prompt(),
            max_tokens=500,
            temperature=0.25,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Commentary generation failed: {e}") from e
    return TBLineCommentaryResponse(commentary=(text or "").strip())


def _is_revenue_item(item: Dict[str, Any]) -> bool:
    at = str(item.get("accountType") or item.get("account_type") or "").lower()
    if at in ("income", "revenue"):
        return True
    if at in ("expense", "cost"):
        return False
    name = str(item.get("account", "")).lower()
    if re.search(
        r"commission|marketing|cost|expense|cogs|payroll|rent|admin|depreciation|interest|salary",
        name,
    ):
        return False
    return bool(re.search(r"revenue|income|turnover|receipts", name)) or (
        bool(re.search(r"\bsales\b", name))
        and not re.search(r"cogs|cost of sales|cost of goods|marketing|commission", name)
    )


def _is_expense_item(item: Dict[str, Any]) -> bool:
    at = str(item.get("accountType") or item.get("account_type") or "").lower()
    if at in ("expense", "cost"):
        return True
    if at in ("income", "revenue"):
        return False
    return not _is_revenue_item(item)


def _variance_pct_decimal(item: Dict[str, Any]) -> float:
    budget = float(item.get("budget", 0) or 0)
    if not budget:
        return 0.0
    actual = float(item.get("actual", 0) or 0)
    return (actual - budget) / abs(budget)


def _get_variance_severity(item: Dict[str, Any]) -> str:
    """Revenue over-budget = favorable; expense over-budget = adverse."""
    vp = _variance_pct_decimal(item)
    if _is_revenue_item(item):
        if vp >= 0.05:
            return "favorable"
        if vp >= 0:
            return "ok"
        if vp < -0.05:
            return "urgent"
        return "monitor"
    if vp <= -0.05:
        return "favorable"
    if vp <= 0:
        return "ok"
    if vp > 0.10:
        return "urgent"
    return "monitor"


def _find_sales_marketing_row(line_items: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    for item in line_items:
        name = str(item.get("account", "")).lower()
        if re.search(r"sales\s*[&/]?\s*marketing|sales\s+and\s+marketing|^marketing\b|s\s*&\s*m", name):
            return item
    return None


def _sales_marketing_narrative(line_items: List[Dict[str, Any]], fc) -> str:
    row = _find_sales_marketing_row(line_items)
    if not row:
        return ""
    budget = float(row.get("budget", 0) or 0)
    actual = float(row.get("actual", 0) or 0)
    if not budget:
        return ""
    pct = (actual - budget) / abs(budget) * 100
    direction = "under" if actual < budget else "over"
    sentiment = "favorable" if actual <= budget else "adverse"
    return (
        f"Sales & Marketing came in {direction} budget at {fc(actual)} vs plan {fc(budget)} "
        f"({pct:+.1f}% {sentiment})."
    )


def _format_line_for_prompt(item: Dict[str, Any], fc) -> str:
    budget = float(item.get("budget", 0) or 0)
    actual = float(item.get("actual", 0) or 0)
    variance = float(item.get("variance", actual - budget) or 0)
    variance_pct = float(item.get("variance_pct", 0) or 0)
    if not variance_pct and budget:
        variance_pct = variance / abs(budget) * 100
    direction = "over" if actual > budget else "under"
    is_revenue = _is_revenue_item(item)
    if is_revenue:
        sentiment = "FAVORABLE" if actual >= budget else "ADVERSE"
    else:
        sentiment = "FAVORABLE" if actual <= budget else "ADVERSE"
    acct_type = "REVENUE" if is_revenue else "EXPENSE"
    return (
        f"- {item.get('account', '')} ({acct_type}): Budget {fc(budget)}, Actual {fc(actual)}, "
        f"Variance {direction} by {abs(variance_pct):.1f}% — {sentiment}"
    )


def _generate_action_items(line_items: List[Dict[str, Any]], fc) -> List[str]:
    """Deterministic action items with correct revenue vs expense severity."""
    candidates: List[tuple[int, str]] = []
    for row in line_items:
        budget = float(row.get("budget", 0) or 0)
        if not budget:
            continue
        severity = _get_variance_severity(row)
        if severity == "ok":
            continue
        is_revenue = _is_revenue_item(row)
        account = str(row.get("account", "Line item"))
        variance_pct = _variance_pct_decimal(row) * 100
        variance_amt = abs(float(row.get("actual", 0) or 0) - budget)
        amt_label = fc(variance_amt)

        if severity == "urgent":
            if is_revenue:
                text = (
                    f"Revenue missed budget by {amt_label} ({variance_pct:+.1f}%) "
                    f"— investigate pipeline gaps"
                )
            else:
                text = (
                    f"{account} exceeded budget by {amt_label} ({variance_pct:+.1f}%) "
                    f"— review and contain"
                )
            candidates.append((1, f"🔴 URGENT: {text}"))
        elif severity == "monitor":
            if is_revenue:
                text = (
                    f"{account} below budget by {amt_label} ({variance_pct:+.1f}%) "
                    f"— validate drivers, update forecast"
                )
            else:
                text = (
                    f"{account} over budget by {amt_label} ({variance_pct:+.1f}%) "
                    f"— monitor closely"
                )
            candidates.append((2, f"🟡 MONITOR: {text}"))
        elif severity == "favorable":
            if is_revenue:
                text = (
                    f"{account} beat budget by {amt_label} (+{abs(variance_pct):.1f}%) "
                    f"— roll into forecast"
                )
            else:
                text = (
                    f"{account} under budget by {amt_label} ({variance_pct:+.1f}%) "
                    f"— sustain cost discipline"
                )
            candidates.append((3, f"🟢 FAVORABLE: {text}"))

    candidates.sort(key=lambda x: x[0])
    return [f"{i + 1}. {text}" for i, (_, text) in enumerate(candidates[:5])]


def _strip_embedded_action_items(text: str) -> str:
    """Remove action-item blocks accidentally included in executive summary."""
    if not text:
        return text
    for marker in (
        "---ACTION_ITEMS---",
        "---LINE_COMMENTARY---",
        "\nACTION ITEMS:",
        "\n3) ACTION ITEMS",
        "\n3. ACTION ITEMS",
    ):
        idx = text.find(marker)
        if idx >= 0:
            text = text[:idx]
    lines = text.strip().split("\n")
    while lines:
        last = lines[-1].strip()
        if re.match(r"^\d+[\.\)]\s*[🔴🟡🟢]", last) or re.match(r"^[🔴🟡🟢]\s*(URGENT|MONITOR|FAVORABLE|OPTIMISE)", last, re.I):
            lines.pop()
        else:
            break
    return "\n".join(lines).strip()


def _line_variance_commentary(item: Dict[str, Any], fc) -> Dict[str, str]:
    """Rule-based line commentary when Claude is unavailable."""
    account = str(item.get("account", "Line item"))
    var = float(item.get("variance", 0) or 0)
    var_pct = float(item.get("variance_pct", 0) or 0)
    is_revenue = _is_revenue_item(item)
    favorable = (is_revenue and var > 0) or (not is_revenue and var < 0)
    adverse = (is_revenue and var < 0) or (not is_revenue and var > 0)

    if favorable:
        why = f"Actual {fc(item.get('actual', 0))} beat budget {fc(item.get('budget', 0))} ({var_pct:+.1f}%)."
        rec = "Validate drivers and roll the uplift into the next forecast cycle."
    elif adverse:
        why = f"Actual {fc(item.get('actual', 0))} missed budget {fc(item.get('budget', 0))} ({var_pct:+.1f}%)."
        rec = "Investigate root cause with the department owner and set a recovery plan."
    else:
        why = f"Variance of {fc(var)} ({var_pct:+.1f}%) is within normal tolerance."
        rec = "Continue monitoring; no immediate action required."

    return {"account": account, "why": why, "recommendation": rec}


def _build_template_narrative(
    *,
    fc,
    narrative_mode: str,
    total_budget: float,
    total_actual: float,
    total_variance: float,
    total_variance_pct: float,
    revenue_budget: float,
    revenue_actual: float,
    revenue_growth_pct: float,
    cost_budget: float,
    cost_actual: float,
    cost_growth_pct: float,
    net_profit_budget: float,
    net_profit_actual: float,
    net_profit_variance: float,
    net_profit_growth_pct: float,
    marketing_growth_pct: float,
    budget_cogs_pct: float,
    actual_cogs_pct: float,
    cogs_growth_pct: float,
    status_label: str,
    material: List[Dict[str, Any]],
    dept_summary: List[Dict[str, Any]],
    line_items: List[Dict[str, Any]],
) -> AINarrativeResponse:
    """Deterministic CFO narrative when the LLM is unavailable."""
    profit_word = "increased" if net_profit_variance >= 0 else "decreased"
    rev_word = "above" if revenue_growth_pct >= 0 else "below"
    cost_word = "above" if cost_growth_pct > 0 else "below"

    if narrative_mode == "board":
        executive_summary = (
            f"• Net profit {profit_word} to {fc(net_profit_actual)} vs budget {fc(net_profit_budget)} "
            f"({net_profit_growth_pct:+.1f}%).\n"
            f"• Revenue {rev_word} plan at {fc(revenue_actual)} ({revenue_growth_pct:+.1f}% vs budget).\n"
            f"• Overall: {status_label}. Monitor working capital as revenue scales."
        )
    elif narrative_mode == "investor":
        executive_summary = (
            f"Revenue reached {fc(revenue_actual)} against a {fc(revenue_budget)} plan ({revenue_growth_pct:+.1f}%). "
            f"Net profit is {fc(net_profit_actual)} vs budget {fc(net_profit_budget)}. "
            f"COGS moved from {budget_cogs_pct:.1f}% to {actual_cogs_pct:.1f}% of revenue. "
            f"Key risk: cost growth at {cost_growth_pct:+.1f}% while revenue grows at {revenue_growth_pct:+.1f}%."
        )
    else:
        executive_summary = (
            f"Net profit {profit_word} to {fc(net_profit_actual)} from a budget of {fc(net_profit_budget)}, "
            f"resulting in a {'favorable' if net_profit_variance >= 0 else 'adverse'} variance of "
            f"{fc(net_profit_variance)} ({net_profit_growth_pct:+.1f}%). "
            f"Revenue is {fc(revenue_actual)} vs budget {fc(revenue_budget)} ({revenue_growth_pct:+.1f}%); "
            f"costs are {fc(cost_actual)} vs budget {fc(cost_budget)} ({cost_growth_pct:+.1f}%). "
            f"COGS increased by {cogs_growth_pct:+.1f}% relative to {revenue_growth_pct:+.1f}% revenue growth. "
        )
        sm_line = _sales_marketing_narrative(line_items, fc)
        if sm_line:
            executive_summary += f" {sm_line}"
        elif marketing_growth_pct:
            executive_summary += f" Marketing spend variance was {marketing_growth_pct:+.1f}% vs budget."
        executive_summary += (
            f" Overall assessment: {status_label}. "
            f"Working capital and liquidity should be monitored as the business scales."
        )

    line_commentary = [_line_variance_commentary(i, fc) for i in material[:10]]
    action_items = _generate_action_items(line_items, fc)
    if not action_items and revenue_growth_pct > cost_growth_pct:
        action_items = [
            "1. 🟢 FAVORABLE: Operating leverage is positive — consider reinvesting savings into growth."
        ]
    elif not action_items and cost_growth_pct > revenue_growth_pct:
        action_items = [
            "1. 🟡 MONITOR: Cost growth is outpacing revenue — review discretionary spend."
        ]

    return AINarrativeResponse(
        executive_summary=executive_summary,
        line_commentary=line_commentary,
        action_items=action_items[:5],
        fallback_used=True,
    )


@router.post("/ai-narrative", response_model=AINarrativeResponse)
async def generate_ai_narrative(body: AINarrativeRequest):
    """Call Claude to produce executive summary, line commentary, and action items."""
    va = body.variance_analysis
    line_items = va.get("line_items") or []
    dept_summary = va.get("department_summary") or []
    total_budget = va.get("total_budget", 0)
    total_actual = va.get("total_actual", 0)
    total_variance = va.get("total_variance", 0)
    total_variance_pct = va.get("total_variance_pct", 0)
    narrative_mode = str(getattr(body, "narrative_mode", "cfo") or "cfo").lower().strip()
    if narrative_mode not in {"cfo", "board", "investor"}:
        narrative_mode = "cfo"
    currency = str(getattr(body, "currency", None) or "USD").upper().strip()
    if currency not in _FPA_CCY_SYMBOL:
        currency = "USD"
    currency_fmt = str(getattr(body, "currency_format", None) or "GLOBAL").upper().strip()
    if currency_fmt not in ("IN", "GLOBAL"):
        currency_fmt = "GLOBAL"
    fc = lambda a: _fp_format_currency(float(a), currency, currency_fmt)

    def _name(item: Dict[str, Any]) -> str:
        return str(item.get("account", "")).lower()

    def _is_revenue(item: Dict[str, Any]) -> bool:
        return _is_revenue_item(item)

    def _is_expense(item: Dict[str, Any]) -> bool:
        return _is_expense_item(item)

    def _is_marketing(item: Dict[str, Any]) -> bool:
        return bool(re.search(r"marketing|advertis|sales\s*[&/]?\s*marketing", _name(item)))

    def _is_cogs(item: Dict[str, Any]) -> bool:
        return bool(re.search(r"cogs|cost of sales|cost of goods", _name(item)))

    revenue_rows = [i for i in line_items if _is_revenue(i)]
    expense_rows = [i for i in line_items if _is_expense(i)]
    marketing_rows = [i for i in line_items if _is_marketing(i)]
    cogs_rows = [i for i in line_items if _is_cogs(i)]

    revenue_budget = sum(float(i.get("budget", 0) or 0) for i in revenue_rows)
    revenue_actual = sum(float(i.get("actual", 0) or 0) for i in revenue_rows)
    cost_budget = sum(float(i.get("budget", 0) or 0) for i in expense_rows)
    cost_actual = sum(float(i.get("actual", 0) or 0) for i in expense_rows)
    net_profit_budget = revenue_budget - cost_budget
    net_profit_actual = revenue_actual - cost_actual
    net_profit_variance = net_profit_actual - net_profit_budget
    revenue_growth_pct = ((revenue_actual - revenue_budget) / revenue_budget * 100) if revenue_budget else 0
    cost_growth_pct = ((cost_actual - cost_budget) / cost_budget * 100) if cost_budget else 0
    net_profit_growth_pct = ((net_profit_variance / abs(net_profit_budget)) * 100) if net_profit_budget else 0

    marketing_budget = sum(float(i.get("budget", 0) or 0) for i in marketing_rows)
    marketing_actual = sum(float(i.get("actual", 0) or 0) for i in marketing_rows)
    marketing_growth_pct = ((marketing_actual - marketing_budget) / marketing_budget * 100) if marketing_budget else 0
    budget_cogs = sum(float(i.get("budget", 0) or 0) for i in cogs_rows)
    actual_cogs = sum(float(i.get("actual", 0) or 0) for i in cogs_rows)
    budget_cogs_pct = (budget_cogs / revenue_budget * 100) if revenue_budget else 0
    actual_cogs_pct = (actual_cogs / revenue_actual * 100) if revenue_actual else 0
    cogs_growth_pct = (
        ((actual_cogs - budget_cogs) / budget_cogs * 100) if budget_cogs else 0.0
    )

    status_label = (
        "High Growth with Margin Expansion"
        if revenue_growth_pct > cost_growth_pct and net_profit_actual > net_profit_budget
        else "Growth with Margin Pressure"
        if revenue_growth_pct > 0 and cost_growth_pct > revenue_growth_pct
        else "Underperforming"
        if revenue_growth_pct < 0 and cost_growth_pct > 0
        else "Mixed performance"
    )

    material = [i for i in line_items if i.get("material") or abs(i.get("variance_pct", 0)) > 10]
    material_text = "\n".join(_format_line_for_prompt(i, fc) for i in material[:15])
    dept_text = "\n".join(
        f"- {d.get('department', '')}: Budget {fc(d.get('budget', 0))}, Actual {fc(d.get('actual', 0))}, "
        f"Variance {d.get('variance_pct', 0):.1f}%"
        for d in dept_summary
    )

    mode_instructions = (
        "NARRATIVE MODE: CFO Summary. Keep detailed commentary with precise numbers and specific operational actions."
        if narrative_mode == "cfo"
        else "NARRATIVE MODE: Board Presentation. Keep commentary concise and visual-first. Max 3 bullets in insights."
        if narrative_mode == "board"
        else "NARRATIVE MODE: Investor Update. Emphasize growth story with TAM context, while explicitly covering key risks."
    )

    sm_narrative_hint = _sales_marketing_narrative(line_items, fc) or (
        f"Marketing spend variance was {marketing_growth_pct:+.1f}% vs budget."
        if marketing_budget
        else ""
    )

    prompt = f"""You are a CFO advisor. Analyse these budget variances and provide:
{mode_instructions}
Currency for all amounts in your answer: {currency} ({currency_fmt} compact style as in DATA — do not mix lakh/crore wording with M/K).

CRITICAL RULES:
- Revenue ABOVE budget = FAVORABLE (🟢), never flag as urgent or overspend
- Revenue BELOW budget = ADVERSE (🔴)
- Expense ABOVE budget = ADVERSE (🔴 or 🟡 by magnitude)
- Expense BELOW budget = FAVORABLE (🟢)
- Use exact variance % from DATA — never say "+0.0%" if the real figure is non-zero
- For Sales & Marketing / OpEx lines, use the computed variance from DATA, not category labels

1) EXECUTIVE SUMMARY: One paragraph (3-5 sentences) for the board.
   Start with this exact logic (use these formatted figures):
   "Net profit increased to {fc(net_profit_actual)} from a budget of {fc(net_profit_budget)}, resulting in a favorable variance of {fc(net_profit_variance)} ({net_profit_growth_pct:.1f}%)."
   IMPORTANT: Net profit is NEVER "spent".
   Include revenue performance: {fc(revenue_actual)} vs budget {fc(revenue_budget)} ({revenue_growth_pct:+.1f}%).
   {f'Include this Sales & Marketing line if relevant: "{sm_narrative_hint}"' if sm_narrative_hint else ''}
   Include COGS insight when relevant:
   "COGS increased by {cogs_growth_pct:+.1f}% relative to {revenue_growth_pct:+.1f}% revenue growth."
   Do NOT embed numbered action items in the executive summary.

2) LINE BY LINE COMMENTARY: For each material variance listed below, give:
   - WHY: One sentence on likely cause.
   - RECOMMENDATION: One sentence on what the CFO should do.
Format each as: "Account Name — WHY: ... RECOMMENDATION: ..."

3) ACTION ITEMS: Exactly 3-5 numbered action items in ---ACTION_ITEMS--- section only (not in summary).
   Mix: URGENT (adverse only), MONITOR (watch trend), FAVORABLE (beat plan).
   Format: "1. 🔴 URGENT: ..." or "2. 🟡 MONITOR: ..." or "3. 🟢 FAVORABLE: ..."
   Never mark revenue above budget as URGENT.

DATA:
Total Budget: {fc(total_budget)}
Total Actual: {fc(total_actual)}
Total Variance: {fc(total_variance)} ({total_variance_pct:+.1f}%)
Revenue: {fc(revenue_actual)} vs budget {fc(revenue_budget)} ({revenue_growth_pct:+.1f}%)
Costs: {fc(cost_actual)} vs budget {fc(cost_budget)} ({cost_growth_pct:+.1f}%)
Net Profit: {fc(net_profit_actual)} vs budget {fc(net_profit_budget)} ({net_profit_growth_pct:+.1f}%)
Marketing: {fc(marketing_actual)} vs budget {fc(marketing_budget)} ({marketing_growth_pct:+.1f}%)
Budget COGS%: {budget_cogs_pct:.1f}% | Actual COGS%: {actual_cogs_pct:.1f}%
COGS YoY vs budget %: {cogs_growth_pct:.1f}% | Revenue growth %: {revenue_growth_pct:.1f}%
Overall status anchor: {status_label}

Material line items:
{material_text}

Department summary:
{dept_text}

Output in this exact structure (so we can parse):
---EXECUTIVE_SUMMARY---
(one paragraph)
---LINE_COMMENTARY---
(one block per material item: "Account — WHY: ... RECOMMENDATION: ...")
---ACTION_ITEMS---
(one numbered item per line)
"""

    try:
        text = llm_service.invoke(prompt=prompt, max_tokens=2000, temperature=0.3)
    except Exception:
        return _build_template_narrative(
            fc=fc,
            narrative_mode=narrative_mode,
            total_budget=total_budget,
            total_actual=total_actual,
            total_variance=total_variance,
            total_variance_pct=total_variance_pct,
            revenue_budget=revenue_budget,
            revenue_actual=revenue_actual,
            revenue_growth_pct=revenue_growth_pct,
            cost_budget=cost_budget,
            cost_actual=cost_actual,
            cost_growth_pct=cost_growth_pct,
            net_profit_budget=net_profit_budget,
            net_profit_actual=net_profit_actual,
            net_profit_variance=net_profit_variance,
            net_profit_growth_pct=net_profit_growth_pct,
            marketing_growth_pct=marketing_growth_pct,
            budget_cogs_pct=budget_cogs_pct,
            actual_cogs_pct=actual_cogs_pct,
            cogs_growth_pct=cogs_growth_pct,
            status_label=status_label,
            material=material,
            dept_summary=dept_summary,
            line_items=line_items,
        )

    executive_summary = ""
    line_commentary: List[Dict[str, str]] = []
    action_items: List[str] = []

    if "---EXECUTIVE_SUMMARY---" in text:
        parts = text.split("---EXECUTIVE_SUMMARY---", 1)[1]
        if "---LINE_COMMENTARY---" in parts:
            exec_part, rest = parts.split("---LINE_COMMENTARY---", 1)
            executive_summary = exec_part.strip()
            if "---ACTION_ITEMS---" in rest:
                comm_part, act_part = rest.split("---ACTION_ITEMS---", 1)
                for line in comm_part.strip().split("\n"):
                    line = line.strip()
                    if not line:
                        continue
                    if " — WHY:" in line or " - WHY:" in line:
                        sep = " — WHY:" if " — WHY:" in line else " - WHY:"
                        a, b = line.split(sep, 1)
                        account = a.strip().strip("- ")
                        if "RECOMMENDATION:" in b:
                            why, rec = b.split("RECOMMENDATION:", 1)
                            line_commentary.append({
                                "account": account,
                                "why": why.replace("WHY:", "").strip(),
                                "recommendation": rec.strip(),
                            })
                        else:
                            line_commentary.append({"account": account, "why": b.strip(), "recommendation": ""})
                for line in act_part.strip().split("\n"):
                    line = line.strip()
                    if line and (line[0].isdigit() or line.startswith("🔴") or line.startswith("🟡") or line.startswith("🟢")):
                        action_items.append(line)
        else:
            executive_summary = parts.strip()[:2000]
    else:
        executive_summary = text.strip()[:2000]

    if not executive_summary:
        executive_summary = text.strip()[:1500]

    executive_summary = _strip_embedded_action_items(executive_summary)
    action_items = _generate_action_items(line_items, fc) or action_items

    return AINarrativeResponse(
        executive_summary=executive_summary,
        line_commentary=line_commentary,
        action_items=action_items[:5],
    )


class DownloadReportRequest(BaseModel):
    variance_analysis: Dict[str, Any]
    ai_narrative: Optional[Dict[str, Any]] = None


@router.post("/download-report")
async def download_report(body: DownloadReportRequest):
    """Generate Excel report: Executive Summary, Full Variance Table, Department Summary, AI Narrative."""
    va = body.variance_analysis
    ai = body.ai_narrative or {}
    line_items = va.get("line_items") or []
    dept_summary = va.get("department_summary") or []

    wb = Workbook()
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Sheet 1: Executive Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1["A1"] = "Variance Analysis Report — Executive Summary"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}"
    row = 4
    if ai.get("executive_summary"):
        ws1[f"A{row}"] = ai["executive_summary"]
        ws1[f"A{row}"].alignment = Alignment(wrap_text=True)
        row += 4
    ws1[f"A{row}"] = "Key metrics"
    row += 1
    ws1[f"A{row}"] = f"Total Budget: {va.get('total_budget', 0):,.0f}"
    ws1[f"B{row}"] = f"Total Actual: {va.get('total_actual', 0):,.0f}"
    row += 1
    ws1[f"A{row}"] = f"Total Variance: {va.get('total_variance', 0):,.0f} ({va.get('total_variance_pct', 0):.1f}%)"
    row += 2
    if ai.get("action_items"):
        ws1[f"A{row}"] = "Action Items"
        row += 1
        for item in ai["action_items"]:
            ws1[f"A{row}"] = item
            row += 1

    # Sheet 2: Full Variance Table
    ws2 = wb.create_sheet("Full Variance Table")
    ws2.append(["Account", "Department", "Budget", "Actual", "Variance", "Variance %", "Status"])
    for i in line_items:
        ws2.append([
            i.get("account", ""),
            i.get("department", ""),
            i.get("budget", 0),
            i.get("actual", 0),
            i.get("variance", 0),
            f"{i.get('variance_pct', 0):.1f}%",
            i.get("status", ""),
        ])

    # Sheet 3: Department Summary
    ws3 = wb.create_sheet("Department Summary")
    ws3.append(["Department", "Budget", "Actual", "Variance", "Variance %", "Status"])
    for d in dept_summary:
        ws3.append([
            d.get("department", ""),
            d.get("budget", 0),
            d.get("actual", 0),
            d.get("variance", 0),
            f"{d.get('variance_pct', 0):.1f}%",
            d.get("status", ""),
        ])

    # Sheet 4: AI Narrative
    ws4 = wb.create_sheet("AI Narrative")
    ws4["A1"] = "Executive Summary"
    ws4["A1"].font = Font(bold=True)
    ws4["A2"] = ai.get("executive_summary", "")
    ws4["A2"].alignment = Alignment(wrap_text=True)
    r = 4
    if ai.get("line_commentary"):
        ws4[f"A{r}"] = "Line commentary"
        ws4[f"A{r}"].font = Font(bold=True)
        r += 1
        for c in ai["line_commentary"]:
            ws4[f"A{r}"] = c.get("account", "")
            ws4[f"B{r}"] = f"WHY: {c.get('why', '')} RECOMMENDATION: {c.get('recommendation', '')}"
            r += 1
        r += 1
    if ai.get("action_items"):
        ws4[f"A{r}"] = "Action Items"
        ws4[f"A{r}"].font = Font(bold=True)
        r += 1
        for item in ai["action_items"]:
            ws4[f"A{r}"] = item
            r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Variance_Analysis_Report.xlsx"},
    )


@router.post("/commentary")
async def generate_fpa_commentary(body: CommentaryRequest):
    """Generate streaming FP&A commentary from variance table rows."""
    variance_data = _normalize_variance_rows(body.variance_data)
    if not variance_data:
        raise HTTPException(status_code=400, detail="variance_data must include at least one row.")
    if not llm_service.is_configured():
        raise HTTPException(status_code=503, detail="LLM not configured: set ANTHROPIC_API_KEY on the server.")
    prompt = _build_commentary_prompt(body.commentary_type, variance_data)
    client = anthropic.Anthropic()
    system_text = (
        "You are a senior FP&A analyst at a Big 4 firm. Write sharp, insight-driven financial commentary. "
        "Never write generic statements. Always reference specific numbers from the data provided."
    )

    def stream_commentary():
        with client.messages.stream(
            model="claude-sonnet-4-20250514",
            max_tokens=1000,
            temperature=0.2,
            system=system_text,
            messages=[{"role": "user", "content": prompt}],
        ) as stream:
            for text in stream.text_stream:
                safe = str(text or "").replace("\n", " ")
                if safe:
                    yield f"data: {safe}\n\n"
            yield "event: done\ndata: [DONE]\n\n"

    return StreamingResponse(stream_commentary(), media_type="text/event-stream")


@router.post("/export-excel")
async def export_fpa_excel(body: ExportExcelRequest):
    """Generate AI-enriched Excel (variance + commentary + risk register)."""
    variance_data = _normalize_variance_rows(body.variance_data)
    if not variance_data:
        raise HTTPException(status_code=400, detail="variance_data must include at least one row.")
    if not llm_service.is_configured():
        raise HTTPException(status_code=503, detail="LLM not configured: set ANTHROPIC_API_KEY on the server.")

    commentary: Dict[str, str] = {}
    for ctype in ["executive", "cfo", "board", "risk"]:
        try:
            commentary[ctype] = llm_service.invoke(
                prompt=_build_commentary_prompt(ctype, variance_data),
                max_tokens=800,
                temperature=0.2,
                system=(
                    "You are a senior FP&A analyst. Write sharp, specific commentary referencing actual numbers."
                ),
                model_id="claude-sonnet-4-20250514",
            )
        except Exception as exc:  # noqa: BLE001
            commentary[ctype] = f"AI generation failed for {ctype}: {exc}"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Variance Analysis"

    DARK_BG = "0A0F1E"
    GOLD = "F5A623"
    GREEN_FILL = "1A4731"
    RED_FILL = "4A1919"
    AMBER_FILL = "4A3A19"
    NEUTRAL_FILL = "1A1F2E"
    WHITE = "FFFFFF"

    headers = ["Account", "Budget", "Actual", "Variance (£)", "Variance (%)", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor=DARK_BG)
        cell.font = Font(bold=True, color=GOLD, size=11)
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(variance_data, 2):
        variance_pct = float(item.get("variance_pct", 0))
        if variance_pct >= 5:
            status = "Favourable"
            row_fill = PatternFill("solid", fgColor=GREEN_FILL)
        elif variance_pct <= -10:
            status = "Critical"
            row_fill = PatternFill("solid", fgColor=RED_FILL)
        elif variance_pct < 0:
            status = "Watch"
            row_fill = PatternFill("solid", fgColor=AMBER_FILL)
        else:
            status = "On Track"
            row_fill = PatternFill("solid", fgColor=NEUTRAL_FILL)

        values = [
            item.get("account", ""),
            float(item.get("budget", 0)),
            float(item.get("actual", 0)),
            float(item.get("variance", 0)),
            float(item.get("variance_pct", 0)),
            status,
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.font = Font(color=WHITE, size=10)
            cell.alignment = Alignment(horizontal="center")
            if col == 5:
                cell.number_format = "0.0%"
                cell.value = float(item.get("variance_pct", 0)) / 100.0

    for i, w in enumerate([30, 15, 15, 15, 15, 18], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    chart = BarChart()
    chart.type = "col"
    chart.title = "Budget vs Actual"
    chart.y_axis.title = "Amount"
    chart.x_axis.title = "Account"
    max_row = len(variance_data) + 1
    data_ref = Reference(ws1, min_col=2, max_col=3, min_row=1, max_row=max_row)
    cats_ref = Reference(ws1, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws1.add_chart(chart, "H2")

    ws2 = wb.create_sheet("AI Commentary")
    ws2.sheet_view.showGridLines = False
    sections = [
        ("EXECUTIVE SUMMARY", commentary.get("executive", "")),
        ("CFO COMMENTARY", commentary.get("cfo", "")),
        ("BOARD NARRATIVE", commentary.get("board", "")),
        ("RISK FLAGS", commentary.get("risk", "")),
    ]
    current_row = 1
    for title, text in sections:
        head = ws2.cell(row=current_row, column=1, value=title)
        head.fill = PatternFill("solid", fgColor=DARK_BG)
        head.font = Font(bold=True, color=GOLD, size=13)
        ws2.row_dimensions[current_row].height = 30
        current_row += 1
        body_cell = ws2.cell(row=current_row, column=1, value=text)
        body_cell.alignment = Alignment(wrap_text=True, vertical="top")
        body_cell.font = Font(color="CCCCCC", size=10)
        body_cell.fill = PatternFill("solid", fgColor="141B2D")
        lines = max(len(str(text)) // 90, 6)
        ws2.row_dimensions[current_row].height = lines * 14
        current_row += 2
    ws2.column_dimensions["A"].width = 110

    ws3 = wb.create_sheet("Risk Register")
    ws3.sheet_view.showGridLines = False
    risk_headers = ["Risk Item", "Severity", "Recommended Action"]
    for col, h in enumerate(risk_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor=DARK_BG)
        cell.font = Font(bold=True, color=GOLD, size=11)
        cell.alignment = Alignment(horizontal="center")

    risk_text = commentary.get("risk", "")
    ws3.cell(row=2, column=1, value=risk_text).alignment = Alignment(wrap_text=True, vertical="top")
    ws3.cell(row=2, column=1).font = Font(color="CCCCCC", size=10)
    ws3.cell(row=2, column=1).fill = PatternFill("solid", fgColor="141B2D")
    ws3.row_dimensions[2].height = 220
    for i, w in enumerate([40, 20, 50], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fpa_report_ai.xlsx"},
    )
