"""
India Invoice — Excel Export & Google Sheets Sync
==================================================
Endpoints:
  POST /api/india/invoice/extract-to-excel   — single extracted invoice → xlsx download
  GET  /api/india/full/purchase-invoices/export-excel — all purchase invoices → xlsx
  POST /api/india/settings/google-sheets     — save / fetch Google Sheets config
  POST /api/india/invoice/sync-to-sheets     — append one invoice row to Google Sheet
  POST /api/india/demo/seed                  — seed 5 realistic demo invoices
"""
from __future__ import annotations

import io
import json
import os
import uuid
from datetime import date, datetime
from typing import Any, Optional

import base64
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, File, Header, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.india_accounting import (
    IndiaCustomer, IndiaJournalEntry, IndiaJournalLine,
    IndiaPurchaseInvoice, IndiaPurchaseInvoiceLine,
    IndiaVendor,
)
from app.services.india_gst_service import calc_gst

router = APIRouter(tags=["India Excel & Sheets"])

NAVY = "0A1628"
WHITE = "FFFFFF"
LIGHT_BLUE = "EBF4FF"
GREEN = "E6F4EA"

INR_FORMAT = '₹#,##0.00'

COLUMNS = [
    ("Invoice #",         18),
    ("Vendor Name",       28),
    ("Vendor GSTIN",      18),
    ("Invoice Date",      14),
    ("Due Date",          14),
    ("HSN / SAC",         12),
    ("Supply Type",       12),
    ("Taxable Amt (₹)",   18),
    ("CGST (₹)",          14),
    ("SGST (₹)",          14),
    ("IGST (₹)",          14),
    ("Total GST (₹)",     14),
    ("Total Amount (₹)",  18),
    ("ITC Eligible",      12),
    ("Status",            10),
]


def _uuid() -> str:
    return str(uuid.uuid4())


def tenant_header(x_tenant_id: str = Header(default="demo")) -> str:
    return x_tenant_id


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _make_workbook(rows: list[dict[str, Any]], sheet_title: str = "GST Invoice Register") -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    navy_fill  = PatternFill("solid", fgColor=NAVY)
    header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    alt_fill   = PatternFill("solid", fgColor=LIGHT_BLUE)
    money_cols = {8, 9, 10, 11, 12, 13}   # 1-indexed columns with INR amounts

    # Header row
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill  = navy_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, inv in enumerate(rows, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        data = [
            inv.get("invoice_number", ""),
            inv.get("vendor_name", ""),
            inv.get("vendor_gstin", ""),
            inv.get("invoice_date", ""),
            inv.get("due_date", ""),
            inv.get("hsn_sac", ""),
            "Inter-state" if inv.get("supply_type") == "inter" else "Intra-state",
            float(inv.get("subtotal", 0)),
            float(inv.get("cgst_amount", 0)),
            float(inv.get("sgst_amount", 0)),
            float(inv.get("igst_amount", 0)),
            float(inv.get("cgst_amount", 0)) + float(inv.get("sgst_amount", 0)) + float(inv.get("igst_amount", 0)),
            float(inv.get("total_amount", 0)),
            "Yes" if inv.get("itc_eligible", True) else "No",
            inv.get("status", "draft").capitalize(),
        ]
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="right" if col_idx in money_cols else "left", vertical="center")
            if fill:
                cell.fill = fill
            if col_idx in money_cols and isinstance(value, float):
                cell.number_format = INR_FORMAT

    # Totals row
    if rows:
        total_row = len(rows) + 2
        total_fill = PatternFill("solid", fgColor="1E3A5F")
        total_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.fill = total_fill
            cell.font = total_font
            cell.alignment = Alignment(horizontal="right" if col_idx in money_cols else "left", vertical="center")
            if col_idx == 1:
                cell.value = f"TOTAL  ({len(rows)} invoices)"
            elif col_idx in money_cols:
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
                cell.number_format = INR_FORMAT

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    return wb


def _wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Enhancement 1: Single extracted invoice → Excel download ────────────────

class ExtractedInvoice(BaseModel):
    invoice_number:  Optional[str] = "INV-0000"
    vendor_name:     Optional[str] = ""
    vendor_gstin:    Optional[str] = ""
    invoice_date:    Optional[str] = ""
    due_date:        Optional[str] = ""
    hsn_sac:         Optional[str] = ""
    supply_type:     Optional[str] = "intra"
    subtotal:        float = 0.0
    cgst_amount:     float = 0.0
    sgst_amount:     float = 0.0
    igst_amount:     float = 0.0
    total_amount:    float = 0.0
    itc_eligible:    bool  = True
    status:          str   = "draft"


@router.post("/api/india/invoice/extract-to-excel")
def extract_to_excel(body: ExtractedInvoice):
    """Convert a single Claude-extracted invoice to an Excel file for download."""
    today_str = date.today().strftime("%Y%m%d")
    inv_no_clean = (body.invoice_number or "INV").replace("/", "-").replace(" ", "_")
    filename = f"Invoice_Extract_{inv_no_clean}_{today_str}.xlsx"

    row = body.model_dump()
    wb = _make_workbook([row], sheet_title="Extracted Invoice")

    content = _wb_to_bytes(wb)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── OCR Extraction endpoint (Claude AI) ─────────────────────────────────────

INDIA_EXTRACT_PROMPT = """Extract fields from this Indian GST invoice and return JSON only (no markdown):
{
  "invoice_number": "",
  "vendor_name": "",
  "vendor_gstin": "",
  "invoice_date": "YYYY-MM-DD",
  "due_date": "YYYY-MM-DD or null",
  "hsn_sac": "",
  "supply_type": "intra or inter",
  "subtotal": 0.0,
  "cgst_amount": 0.0,
  "sgst_amount": 0.0,
  "igst_amount": 0.0,
  "total_amount": 0.0,
  "itc_eligible": true
}
Rules:
- supply_type = "inter" if IGST > 0, else "intra"
- If CGST and SGST both present → intra-state
- If only IGST present → inter-state
- subtotal = taxable value before GST
- All amounts as plain numbers (no ₹ symbol)
- Return JSON only."""


def _extract_json_safe(text: str) -> dict:
    text = text.strip()
    for wrapper in ("```json", "```"):
        if wrapper in text:
            text = text.split(wrapper, 1)[-1].split("```", 1)[0].strip()
            break
    try:
        return json.loads(text)
    except Exception:
        # Try to find a JSON object in the text
        m = re.search(r"\{[\s\S]+\}", text)
        if m:
            try:
                return json.loads(m.group())
            except Exception:
                pass
    return {}


@router.post("/api/india/invoice/ocr-extract")
async def ocr_extract_invoice(file: UploadFile = File(...)):
    """
    Accept a PDF or image of an Indian GST invoice.
    Use Claude to extract all fields including CGST/SGST/IGST.
    Returns structured JSON.
    """
    content = await file.read()
    filename = file.filename or "invoice"
    lower = filename.lower()

    # Determine media type
    if lower.endswith(".pdf"):
        # Convert first page to image via pdf2image if available, else use base64 directly
        media_type = "application/pdf"
        encoded = base64.standard_b64encode(content).decode()
        # Claude supports PDF via document type
        message_content = [
            {
                "type": "document",
                "source": {"type": "base64", "media_type": "application/pdf", "data": encoded},
            },
            {"type": "text", "text": INDIA_EXTRACT_PROMPT},
        ]
    elif any(lower.endswith(ext) for ext in (".jpg", ".jpeg", ".png", ".webp", ".gif")):
        mt_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
                  ".webp": "image/webp", ".gif": "image/gif"}
        ext = "." + lower.rsplit(".", 1)[-1]
        media_type = mt_map.get(ext, "image/jpeg")
        encoded = base64.standard_b64encode(content).decode()
        message_content = [
            {
                "type": "image",
                "source": {"type": "base64", "media_type": media_type, "data": encoded},
            },
            {"type": "text", "text": INDIA_EXTRACT_PROMPT},
        ]
    else:
        raise HTTPException(400, "Unsupported file type. Upload PDF, JPG, PNG, or WEBP.")

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise HTTPException(500, "ANTHROPIC_API_KEY not configured on server.")

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-opus-4-5",
            max_tokens=1024,
            messages=[{"role": "user", "content": message_content}],
        )
        raw_text = response.content[0].text if response.content else "{}"
        data = _extract_json_safe(raw_text)
    except Exception as exc:
        raise HTTPException(500, f"Claude extraction failed: {exc}")

    # Ensure numeric fields are float
    for field in ("subtotal", "cgst_amount", "sgst_amount", "igst_amount", "total_amount"):
        try:
            data[field] = float(data.get(field, 0) or 0)
        except (TypeError, ValueError):
            data[field] = 0.0

    # Derive supply_type from GST amounts if Claude didn't detect it
    if not data.get("supply_type"):
        data["supply_type"] = "inter" if data.get("igst_amount", 0) > 0 else "intra"

    # Fill missing total if needed
    if not data.get("total_amount") and data.get("subtotal"):
        data["total_amount"] = (
            data["subtotal"]
            + data.get("cgst_amount", 0)
            + data.get("sgst_amount", 0)
            + data.get("igst_amount", 0)
        )

    return data


# ─── FIX 1 / Enhancement 2: Batch Excel — all purchase invoices for a period ──

@router.get("/api/india/full/purchase-invoices/export/excel")
@router.get("/api/india/full/purchase-invoices/export-excel")  # legacy alias
def export_purchase_invoices_excel(
    period: Optional[str] = None,
    tenant: str = Depends(tenant_header),
    db: Session = Depends(get_db),
):
    """Export all India purchase invoices (optionally filtered by period) to Excel."""
    q = db.query(IndiaPurchaseInvoice, IndiaVendor).outerjoin(
        IndiaVendor, IndiaPurchaseInvoice.vendor_id == IndiaVendor.id
    ).filter(IndiaPurchaseInvoice.tenant_id == tenant)

    if period:
        q = q.filter(IndiaPurchaseInvoice.invoice_date.like(f"{period}%"))

    results = q.order_by(IndiaPurchaseInvoice.invoice_date.desc()).limit(1000).all()

    rows = []
    for inv, vendor in results:
        # Get first line's HSN for register (representative)
        first_line = (
            db.query(IndiaPurchaseInvoiceLine)
            .filter_by(invoice_id=inv.id)
            .first()
        )
        rows.append({
            "invoice_number":  inv.invoice_number,
            "vendor_name":     vendor.name if vendor else "",
            "vendor_gstin":    vendor.gstin if vendor else "",
            "invoice_date":    str(inv.invoice_date) if inv.invoice_date else "",
            "due_date":        str(inv.due_date) if inv.due_date else "",
            "hsn_sac":         first_line.hsn_sac if first_line else "",
            "supply_type":     inv.supply_type or "intra",
            "subtotal":        float(inv.subtotal or 0),
            "cgst_amount":     float(inv.cgst_amount or 0),
            "sgst_amount":     float(inv.sgst_amount or 0),
            "igst_amount":     float(inv.igst_amount or 0),
            "total_amount":    float(inv.total_amount or 0),
            "itc_eligible":    bool(inv.itc_eligible),
            "status":          inv.status or "draft",
        })

    period_label = period or date.today().strftime("%Y-%m")
    filename = f"GST_Purchase_Register_{period_label}.xlsx"

    wb = _make_workbook(rows, sheet_title=f"GST Register {period_label}")

    # Footer row — Gnanova branding
    ws = wb.active
    footer_row = ws.max_row + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(COLUMNS))
    footer_cell = ws.cell(row=footer_row, column=1,
                           value="Generated by FinReportAI by Gnanova Pro AI Technologies")
    footer_cell.font = Font(name="Calibri", italic=True, size=9, color="6B7280")
    footer_cell.alignment = Alignment(horizontal="center")

    content = _wb_to_bytes(wb)

    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── FIX 2: PDF Export for Purchase Invoices ─────────────────────────────────

@router.get("/api/india/full/purchase-invoices/export/pdf")
def export_purchase_invoices_pdf(
    period: Optional[str] = None,
    company_name: Optional[str] = "Demo Company",
    company_gstin: Optional[str] = "",
    tenant: str = Depends(tenant_header),
    db: Session = Depends(get_db),
):
    """Export all India purchase invoices to a branded PDF — Gnanova footer."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    q = (
        db.query(IndiaPurchaseInvoice, IndiaVendor)
        .outerjoin(IndiaVendor, IndiaPurchaseInvoice.vendor_id == IndiaVendor.id)
        .filter(IndiaPurchaseInvoice.tenant_id == tenant)
    )
    if period:
        q = q.filter(IndiaPurchaseInvoice.invoice_date.like(f"{period}%"))
    results = q.order_by(IndiaPurchaseInvoice.invoice_date.desc()).limit(1000).all()

    period_label = period or date.today().strftime("%Y-%m")
    filename = f"GST_Purchase_Register_{period_label}.pdf"

    NAVY   = colors.HexColor("#0A1628")
    TEAL   = colors.HexColor("#0E7490")
    LIGHT  = colors.HexColor("#F0F9FF")
    WHITE  = colors.white
    GREY   = colors.HexColor("#F3F4F6")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            topMargin=15*mm, bottomMargin=20*mm,
                            leftMargin=12*mm, rightMargin=12*mm)
    styles = getSampleStyleSheet()
    story = []

    # ── Header ──
    story.append(Paragraph(
        "<font color='#0A1628' size=16><b>FinReportAI</b></font>"
        "<font color='#0E7490' size=10>  by Gnanova Pro AI Technologies</font>",
        styles["Normal"],
    ))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph(
        f"<b>GST Purchase Invoice Register — {period_label}</b>",
        ParagraphStyle("title", fontSize=13, textColor=NAVY, spaceAfter=2),
    ))
    if company_name:
        story.append(Paragraph(
            f"<b>{company_name}</b>"
            + (f"  |  GSTIN: {company_gstin}" if company_gstin else ""),
            ParagraphStyle("sub", fontSize=9, textColor=colors.grey, spaceAfter=4),
        ))
    story.append(Spacer(1, 5*mm))

    # ── Table ──
    headers = ["Invoice #", "Vendor", "GSTIN", "Date", "HSN/SAC",
               "Supply", "Taxable (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)",
               "Total GST (₹)", "Total (₹)", "ITC"]
    col_widths = [30*mm, 42*mm, 32*mm, 20*mm, 18*mm,
                  16*mm, 24*mm, 20*mm, 20*mm, 20*mm,
                  24*mm, 24*mm, 12*mm]

    data = [headers]
    tot_taxable = tot_cgst = tot_sgst = tot_igst = tot_gst = tot_total = 0.0

    for inv, vendor in results:
        fl = (db.query(IndiaPurchaseInvoiceLine).filter_by(invoice_id=inv.id).first())
        taxable = float(inv.subtotal or 0)
        cgst = float(inv.cgst_amount or 0)
        sgst = float(inv.sgst_amount or 0)
        igst = float(inv.igst_amount or 0)
        total = float(inv.total_amount or 0)
        tot_taxable += taxable; tot_cgst += cgst; tot_sgst += sgst
        tot_igst += igst; tot_gst += cgst + sgst + igst; tot_total += total

        def inr(v: float) -> str:
            return f"₹{v:,.2f}"

        data.append([
            inv.invoice_number or "",
            (vendor.name if vendor else "")[:28],
            (vendor.gstin if vendor else "") or "",
            str(inv.invoice_date) if inv.invoice_date else "",
            (fl.hsn_sac if fl else "") or "",
            "Inter" if inv.supply_type == "inter" else "Intra",
            inr(taxable), inr(cgst), inr(sgst), inr(igst),
            inr(cgst + sgst + igst), inr(total),
            "Y" if inv.itc_eligible else "N",
        ])

    # Totals row
    data.append([
        "TOTAL", f"{len(results)} invoices", "", "", "", "",
        f"₹{tot_taxable:,.2f}", f"₹{tot_cgst:,.2f}",
        f"₹{tot_sgst:,.2f}", f"₹{tot_igst:,.2f}",
        f"₹{tot_gst:,.2f}", f"₹{tot_total:,.2f}", "",
    ])

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), WHITE),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 7),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        # Totals row
        ("BACKGROUND", (0, -1), (-1, -1), TEAL),
        ("TEXTCOLOR",  (0, -1), (-1, -1), WHITE),
        ("FONTNAME",   (0, -1), (-1, -1), "Helvetica-Bold"),
        # Data rows alternating
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [WHITE, GREY]),
        ("FONTSIZE",   (0, 1), (-1, -1), 7),
        # Right-align numeric columns (6-11)
        ("ALIGN",      (6, 1), (11, -1), "RIGHT"),
        ("ALIGN",      (0, 1), (5, -1), "LEFT"),
        ("ALIGN",      (12, 1), (12, -1), "CENTER"),
        # Grid
        ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
        ("LINEBELOW",  (0, 0), (-1, 0), 1, NAVY),
        ("LINEABOVE",  (0, -1), (-1, -1), 1, TEAL),
        # Padding
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 6*mm))

    # ── GST Summary box ──
    summary_data = [
        ["GST Summary", ""],
        ["Total Taxable Amount", f"₹{tot_taxable:,.2f}"],
        ["Total CGST",           f"₹{tot_cgst:,.2f}"],
        ["Total SGST",           f"₹{tot_sgst:,.2f}"],
        ["Total IGST",           f"₹{tot_igst:,.2f}"],
        ["Total GST",            f"₹{tot_gst:,.2f}"],
        ["Total Invoice Value",  f"₹{tot_total:,.2f}"],
    ]
    summary_tbl = Table(summary_data, colWidths=[60*mm, 40*mm])
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), TEAL),
        ("TEXTCOLOR",   (0, 0), (-1, 0), WHITE),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("SPAN",        (0, 0), (-1, 0)),
        ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
        ("FONTSIZE",    (0, 0), (-1, -1), 8),
        ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [WHITE, LIGHT]),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
        ("ALIGN",       (1, 1), (1, -1), "RIGHT"),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_tbl)
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph(
        "For GST reconciliation purposes only. "
        "Generated by FinReportAI by Gnanova Pro AI Technologies.",
        ParagraphStyle("footer", fontSize=7, textColor=colors.grey, alignment=TA_CENTER),
    ))

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── FIX 3: GSTR-3B Excel export ─────────────────────────────────────────────

@router.get("/api/india/full/gst-returns/{return_id}/export/excel")
def export_gstr3b_excel(
    return_id: str,
    tenant: str = Depends(tenant_header),
    db: Session = Depends(get_db),
):
    """Export a GSTR-3B return + underlying invoice details as a two-sheet Excel."""
    from app.models.india_accounting import IndiaGSTReturn

    ret = db.query(IndiaGSTReturn).filter_by(id=return_id, tenant_id=tenant).first()
    if not ret:
        raise HTTPException(404, "GST Return not found")

    period = ret.period or date.today().strftime("%Y-%m")
    company = tenant[:20].upper()
    filename = f"GSTR3B_{period}_{company}.xlsx"

    wb = openpyxl.Workbook()

    # ── Sheet 1: GSTR-3B Summary ──
    ws1 = wb.active
    ws1.title = "GSTR-3B Summary"
    navy_fill  = PatternFill("solid", fgColor=NAVY)
    teal_fill  = PatternFill("solid", fgColor="0E7490")
    white_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    bold_font  = Font(name="Calibri", bold=True, size=11)

    def _h(ws, row: int, col: int, val, fill=None, font=None, fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        if fill: c.fill = fill
        if font: c.font = font
        if fmt:  c.number_format = fmt
        c.alignment = Alignment(vertical="center")
        return c

    # Title
    ws1.merge_cells("A1:D1")
    t = ws1.cell(row=1, column=1, value=f"GSTR-3B Summary — {period}")
    t.fill = navy_fill; t.font = white_font
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    headers = ["Section", "CGST (₹)", "SGST (₹)", "IGST (₹)"]
    for ci, h in enumerate(headers, 1):
        c = ws1.cell(row=2, column=ci, value=h)
        c.fill = teal_fill; c.font = white_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    rows_data = [
        ("3.1 Outward Supplies (Output Tax)",
         float(ret.total_cgst or 0), float(ret.total_sgst or 0), float(ret.total_igst or 0)),
        ("4 ITC Available",
         -float(ret.itc_cgst or 0), -float(ret.itc_sgst or 0), -float(ret.itc_igst or 0)),
        ("Net GST Payable",
         float(ret.net_cgst_payable or 0), float(ret.net_sgst_payable or 0), float(ret.net_igst_payable or 0)),
    ]
    alt = PatternFill("solid", fgColor=LIGHT_BLUE)
    for ri, (label, cgst, sgst, igst) in enumerate(rows_data, 3):
        fill_row = alt if ri % 2 == 0 else None
        ws1.cell(row=ri, column=1, value=label).font = Font(name="Calibri", bold=(ri == 5), size=10)
        for ci, val in enumerate([cgst, sgst, igst], 2):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.number_format = INR_FORMAT
            c.alignment = Alignment(horizontal="right", vertical="center")
            if ri == 5: c.font = bold_font
            if fill_row: c.fill = fill_row

    # Total payable callout
    ws1.merge_cells("A6:C6")
    lbl = ws1.cell(row=6, column=1, value="Total GST Payable (all heads)")
    lbl.fill = navy_fill; lbl.font = white_font
    total_c = ws1.cell(row=6, column=4, value=float(ret.total_payable or 0))
    total_c.fill = navy_fill; total_c.font = white_font
    total_c.number_format = INR_FORMAT
    total_c.alignment = Alignment(horizontal="right", vertical="center")

    if ret.arn:
        ws1.cell(row=8, column=1, value=f"ARN: {ret.arn}").font = Font(color="6B7280", size=9)
    if ret.ai_summary:
        ws1.cell(row=9, column=1, value="AI Summary:").font = Font(bold=True, size=9)
        ws1.cell(row=10, column=1, value=ret.ai_summary).font = Font(size=9)

    for ci, w in enumerate([38, 18, 18, 18], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Invoice details that fed the return ──
    ws2 = wb.create_sheet("Invoice Details")
    inv_headers = ["Invoice #", "Vendor", "Date", "Supply", "Taxable (₹)",
                   "CGST (₹)", "SGST (₹)", "IGST (₹)", "Total (₹)", "ITC Eligible"]
    for ci, h in enumerate(inv_headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.fill = navy_fill; c.font = white_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    purchases = (
        db.query(IndiaPurchaseInvoice, IndiaVendor)
        .outerjoin(IndiaVendor, IndiaPurchaseInvoice.vendor_id == IndiaVendor.id)
        .filter(
            IndiaPurchaseInvoice.tenant_id == tenant,
            IndiaPurchaseInvoice.status == "posted",
            IndiaPurchaseInvoice.invoice_date.like(f"{period}%"),
        )
        .all()
    )
    alt2 = PatternFill("solid", fgColor=LIGHT_BLUE)
    for ri, (inv, vendor) in enumerate(purchases, 2):
        fill = alt2 if ri % 2 == 0 else None
        row_vals = [
            inv.invoice_number,
            vendor.name if vendor else "",
            str(inv.invoice_date) if inv.invoice_date else "",
            "Inter-state" if inv.supply_type == "inter" else "Intra-state",
            float(inv.subtotal or 0),
            float(inv.cgst_amount or 0),
            float(inv.sgst_amount or 0),
            float(inv.igst_amount or 0),
            float(inv.total_amount or 0),
            "Yes" if inv.itc_eligible else "No",
        ]
        money_cols2 = {5, 6, 7, 8, 9}
        for ci, val in enumerate(row_vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(
                horizontal="right" if ci in money_cols2 else "left", vertical="center"
            )
            if ci in money_cols2 and isinstance(val, float):
                c.number_format = INR_FORMAT
            if fill: c.fill = fill

    for ci, w in enumerate([22, 30, 14, 14, 18, 14, 14, 14, 18, 12], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(inv_headers))}1"
    ws2.freeze_panes = "A2"

    content = _wb_to_bytes(wb)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Enhancement 3: Google Sheets settings & sync ────────────────────────────

# In-memory store (replace with DB column or env-based config in production)
_sheets_config: dict[str, dict] = {}


class GoogleSheetsConfig(BaseModel):
    sheet_url:   str
    enabled:     bool = True
    sheet_name:  Optional[str] = "India Invoices"


@router.post("/api/india/settings/google-sheets")
def save_google_sheets_config(
    body: GoogleSheetsConfig,
    tenant: str = Depends(tenant_header),
):
    """Save Google Sheets URL for auto-sync."""
    _sheets_config[tenant] = body.model_dump()
    return {"saved": True, "tenant": tenant, "sheet_url": body.sheet_url}


@router.get("/api/india/settings/google-sheets")
def get_google_sheets_config(tenant: str = Depends(tenant_header)):
    """Fetch current Google Sheets config."""
    cfg = _sheets_config.get(tenant)
    return cfg or {"enabled": False, "sheet_url": "", "sheet_name": "India Invoices"}


@router.post("/api/india/invoice/sync-to-sheets")
def sync_invoice_to_sheets(
    body: ExtractedInvoice,
    tenant: str = Depends(tenant_header),
):
    """
    Append one invoice row to the tenant's configured Google Sheet.
    Requires GOOGLE_SERVICE_ACCOUNT_JSON env var with service account credentials JSON.
    Falls back gracefully if not configured.
    """
    cfg = _sheets_config.get(tenant, {})
    if not cfg.get("enabled") or not cfg.get("sheet_url"):
        return {
            "synced": False,
            "message": "Google Sheets not configured. Enable it in India Accounting → Settings.",
        }

    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
    if not sa_json:
        return {
            "synced": False,
            "message": "Google Sheets credentials not configured on the server. "
                       "Add GOOGLE_SERVICE_ACCOUNT_JSON env var to enable sync.",
        }

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_info(json.loads(sa_json), scopes=scopes)
        gc = gspread.authorize(creds)

        # Extract spreadsheet ID from URL
        url = cfg["sheet_url"]
        # Handle both /d/{id}/edit and /d/{id}/ formats
        parts = url.split("/d/")
        if len(parts) < 2:
            raise ValueError("Invalid Google Sheets URL")
        sheet_id = parts[1].split("/")[0]

        sh = gc.open_by_key(sheet_id)
        sheet_name = cfg.get("sheet_name") or "India Invoices"
        try:
            ws = sh.worksheet(sheet_name)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=sheet_name, rows=1000, cols=20)
            # Write header if new sheet
            ws.append_row([col for col, _ in COLUMNS])

        total_gst = body.cgst_amount + body.sgst_amount + body.igst_amount
        row_data = [
            body.invoice_number or "",
            body.vendor_name or "",
            body.vendor_gstin or "",
            body.invoice_date or "",
            body.due_date or "",
            body.hsn_sac or "",
            "Inter-state" if body.supply_type == "inter" else "Intra-state",
            body.subtotal,
            body.cgst_amount,
            body.sgst_amount,
            body.igst_amount,
            total_gst,
            body.total_amount,
            "Yes" if body.itc_eligible else "No",
            body.status.capitalize(),
        ]
        ws.append_row(row_data)
        return {"synced": True, "sheet_url": cfg["sheet_url"]}

    except ImportError:
        return {
            "synced": False,
            "message": "gspread library not installed. Add 'gspread google-auth' to requirements.",
        }
    except Exception as exc:
        return {"synced": False, "message": str(exc)}


# ─── Enhancement 4: Demo seed data ───────────────────────────────────────────

DEMO_VENDORS = [
    {"name": "Tata Consultancy Services Ltd", "gstin": "27AAACT2727Q1ZW", "state_code": "27", "state_name": "Maharashtra"},
    {"name": "Reliance Industries Ltd",        "gstin": "27AAACR5055K1Z4", "state_code": "27", "state_name": "Maharashtra"},
    {"name": "Infosys Ltd",                    "gstin": "29AABCI1681B1ZN", "state_code": "29", "state_name": "Karnataka"},
    {"name": "Amazon India",                   "gstin": "29AAGCS8989F1Z9", "state_code": "27", "state_name": "Maharashtra"},
    {"name": "HDFC Bank Ltd",                  "gstin": "27AAACH2702H1ZC", "state_code": "27", "state_name": "Maharashtra"},
]

DEMO_INVOICES = [
    # (vendor_idx, supply_type, gst_rate, subtotal, hsn, due_days, desc)
    (0, "intra", 18.0, 100000.0, "998314", 30, "IT Consulting Services"),        # TCS MH→MH  → CGST 9,000 + SGST 9,000 = ₹1,18,000
    (1, "inter", 18.0, 200000.0, "271000", 45, "Petroleum Products Supply"),     # Reliance MH→GJ → IGST 36,000 = ₹2,36,000
    (2, "intra", 18.0,  50000.0, "998313", 30, "Software Development Services"), # Infosys KA→KA  → CGST 4,500 + SGST 4,500 = ₹59,000
    (3, "inter", 18.0,  40000.0, "996111", 15, "Cloud Hosting & Marketplace"),   # Amazon MH→DL   → IGST 7,200 = ₹47,200
    (4, "intra", 18.0,  20000.0, "997120", 30, "Banking & Processing Fees"),     # HDFC MH→MH     → CGST 1,800 + SGST 1,800 = ₹23,600
]


@router.post("/api/india/demo/seed")
def seed_demo_invoices(
    tenant: str = Depends(tenant_header),
    db: Session = Depends(get_db),
):
    """Seed 5 realistic India purchase invoices for demo. Idempotent — skips if already seeded."""
    existing = db.query(IndiaPurchaseInvoice).filter_by(tenant_id=tenant).count()
    if existing >= 5:
        return {"seeded": 0, "message": "Demo data already present"}

    # Upsert vendors
    vendor_ids: list[str] = []
    for v in DEMO_VENDORS:
        existing_v = db.query(IndiaVendor).filter_by(tenant_id=tenant, gstin=v["gstin"]).first()
        if existing_v:
            vendor_ids.append(existing_v.id)
        else:
            nv = IndiaVendor(
                id=_uuid(), tenant_id=tenant,
                name=v["name"], gstin=v["gstin"],
                state_code=v["state_code"], state_name=v["state_name"],
                tds_applicable=False, payment_terms_days=30,
            )
            db.add(nv)
            db.flush()
            vendor_ids.append(nv.id)

    today = date.today()
    period = today.strftime("%Y-%m")
    seeded = 0
    seeded_detail: list[dict[str, Any]] = []

    for i, (vendor_idx, supply_type, gst_rate, subtotal, hsn, due_days, desc) in enumerate(DEMO_INVOICES):
        inv_date = date(today.year, today.month, max(1, today.day - (i * 5 + 3)))
        due_date = date(today.year, today.month, min(28, inv_date.day + due_days))

        gst = calc_gst(subtotal, gst_rate, supply_type)
        total = subtotal + gst["total_tax"]

        inv_no = f"DEMO-{today.strftime('%Y%m')}-{i+1:03d}"
        inv = IndiaPurchaseInvoice(
            id=_uuid(), tenant_id=tenant,
            invoice_number=inv_no,
            vendor_id=vendor_ids[vendor_idx],
            invoice_date=inv_date,
            due_date=due_date,
            supply_type=supply_type,
            subtotal=subtotal,
            cgst_amount=gst["cgst"],
            sgst_amount=gst["sgst"],
            igst_amount=gst["igst"],
            total_amount=total,
            outstanding=total,
            itc_eligible=True,
            status="posted",
        )
        db.add(inv)
        db.flush()

        db.add(IndiaPurchaseInvoiceLine(
            id=_uuid(), invoice_id=inv.id,
            description=desc, hsn_sac=hsn,
            quantity=1, unit_price=subtotal,
            gst_rate=gst_rate,
            line_subtotal=subtotal,
            line_cgst=gst["cgst"],
            line_sgst=gst["sgst"],
            line_igst=gst["igst"],
            line_total=total,
            itc_eligible=True,
        ))
        seeded += 1
        seeded_detail.append({
            "invoice_number": inv_no,
            "vendor_name": DEMO_VENDORS[vendor_idx]["name"],
            "vendor_gstin": DEMO_VENDORS[vendor_idx]["gstin"],
            "supply_type": supply_type,
            "subtotal": subtotal,
            "cgst_amount": gst["cgst"],
            "sgst_amount": gst["sgst"],
            "igst_amount": gst["igst"],
            "total_amount": total,
        })

    db.commit()
    return {
        "seeded": seeded,
        "message": f"Seeded {seeded} demo invoices for tenant '{tenant}'",
        "period": period,
        "invoices": [d["invoice_number"] for d in seeded_detail],
        "detail": seeded_detail,
    }
