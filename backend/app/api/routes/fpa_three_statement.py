"""Simplified 3-statement financial model — POST /api/fpa/three-statement"""

from __future__ import annotations

import base64
import io
import logging
from typing import Any, List, Literal, Optional

from fastapi import APIRouter, Depends, HTTPException
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.services.fpa_commentary import fpa_commentary
from app.services.fpa_result_store import store_fpa_result

logger = logging.getLogger(__name__)
router = APIRouter(tags=["FP&A 3-Statement"])


class ThreeStatementRequest(BaseModel):
    company_name: str = "Demo Co"
    industry: str = "Software"
    projection_years: Literal[3, 5, 10] = 5
    revenue_base: float = Field(10_000_000, description="Year-0 revenue")
    revenue_growth_pct: float = Field(0.12, description="Annual revenue growth (base case)")
    gross_margin_pct: float = 0.72
    ebitda_margin_pct: float = 0.25
    net_margin_pct: float = 0.15
    starting_cash: float = 3_000_000
    total_debt: float = 5_000_000
    capex_pct_revenue: float = 0.05
    dso_days: float = 45.0
    dpo_days: float = 30.0
    dio_days: float = 20.0
    tax_rate: float = 0.25
    da_pct_revenue: float = 0.03
    scenario: Literal["base", "bull", "bear"] = "base"
    bull_growth_delta: float = 0.05
    bear_growth_delta: float = -0.06
    user_id: Optional[str] = None


def _project(req: ThreeStatementRequest) -> dict[str, Any]:
    n = int(req.projection_years)
    growth = req.revenue_growth_pct
    if req.scenario == "bull":
        growth += req.bull_growth_delta
    elif req.scenario == "bear":
        growth += req.bear_growth_delta

    years = list(range(0, n + 1))
    rev = [req.revenue_base * ((1 + growth) ** y) for y in years]

    pl_rows: list[dict[str, Any]] = []
    bs_rows: list[dict[str, Any]] = []
    cf_rows: list[dict[str, Any]] = []

    cash = req.starting_cash
    debt = req.total_debt
    ar = rev[0] * (req.dso_days / 365.0)
    inv = rev[0] * (req.dio_days / 365.0) * (1 - req.gross_margin_pct)
    ap = rev[0] * (req.dpo_days / 365.0) * (1 - req.gross_margin_pct)
    ppe = rev[0] * 1.2
    equity = None  # plug below

    flags: list[str] = []
    if growth > 0.35:
        flags.append("Revenue growth assumption is very aggressive versus typical planning ranges.")

    for i, y in enumerate(years):
        r = rev[i]
        cogs = r * (1 - req.gross_margin_pct)
        gp = r - cogs
        ebitda = r * req.ebitda_margin_pct
        da = r * req.da_pct_revenue
        ebit = ebitda - da
        tax = max(0.0, ebit) * req.tax_rate
        ni = ebit - tax

        prev_ar = ar
        prev_inv = inv
        prev_ap = ap
        ar = r * (req.dso_days / 365.0)
        inv = r * (req.dio_days / 365.0) * (1 - req.gross_margin_pct)
        ap = r * (req.dpo_days / 365.0) * (1 - req.gross_margin_pct)
        delta_wc = (ar - prev_ar) + (inv - prev_inv) - (ap - prev_ap)

        capex = r * req.capex_pct_revenue
        ppe = max(0.0, ppe + capex - da)

        cfo = ni + da - delta_wc
        cfi = -capex
        debt_pay = min(500_000, max(0.0, debt * 0.05)) if i > 0 else 0.0
        cff = -debt_pay
        debt = max(0.0, debt - debt_pay)
        net_cf = cfo + cfi + cff
        cash = cash + net_cf

        pl_rows.append(
            {
                "year_index": y,
                "revenue": r,
                "cogs": cogs,
                "gross_profit": gp,
                "ebitda": ebitda,
                "da": da,
                "ebit": ebit,
                "tax": tax,
                "net_income": ni,
            }
        )

        assets = cash + ar + inv + ppe
        if equity is None:
            equity = assets - debt - ap
        liabilities = debt + ap
        bs_rows.append(
            {
                "year_index": y,
                "cash": cash,
                "ar": ar,
                "inventory": inv,
                "ppe": ppe,
                "total_assets": assets,
                "debt": debt,
                "ap": ap,
                "total_liabilities": liabilities,
                "equity": assets - liabilities,
            }
        )

        cf_rows.append(
            {
                "year_index": y,
                "cfo": cfo,
                "cfi": cfi,
                "cff": cff,
                "net_change": net_cf,
                "ending_cash": cash,
            }
        )

        if abs(bs_rows[-1]["total_assets"] - (bs_rows[-1]["total_liabilities"] + bs_rows[-1]["equity"])) > 1.0:
            flags.append(f"Year {y}: balance sheet identity drift > £1 — review working-capital linkage.")

    return {"pl_data": pl_rows, "bs_data": bs_rows, "cf_data": cf_rows, "flags": flags}


def _workbook_b64(pl: list, bs: list, cf: list, company: str) -> str:
    wb = Workbook()
    ws_pl = wb.active
    ws_pl.title = "P_L"
    ws_pl.append(["Year", "Revenue", "COGS", "Gross Profit", "EBITDA", "D&A", "EBIT", "Tax", "Net Income"])
    for row in pl:
        ws_pl.append(
            [
                row["year_index"],
                row["revenue"],
                row["cogs"],
                row["gross_profit"],
                row["ebitda"],
                row["da"],
                row["ebit"],
                row["tax"],
                row["net_income"],
            ]
        )
    ws_bs = wb.create_sheet("Balance_Sheet")
    ws_bs.append(
        ["Year", "Cash", "AR", "Inventory", "PPE", "Total Assets", "Debt", "AP", "Total Liab", "Equity"]
    )
    for row in bs:
        ws_bs.append(
            [
                row["year_index"],
                row["cash"],
                row["ar"],
                row["inventory"],
                row["ppe"],
                row["total_assets"],
                row["debt"],
                row["ap"],
                row["total_liabilities"],
                row["equity"],
            ]
        )
    ws_cf = wb.create_sheet("Cash_Flow")
    ws_cf.append(["Year", "CFO", "CFI", "CFF", "Net", "Ending Cash"])
    for row in cf:
        ws_cf.append(
            [row["year_index"], row["cfo"], row["cfi"], row["cff"], row["net_change"], row["ending_cash"]]
        )

    if len(pl) >= 2:
        try:
            chart = LineChart()
            chart.title = "Revenue"
            chart.y_axis.title = "£"
            data = Reference(ws_pl, min_col=2, min_row=1, max_row=1 + len(pl), max_col=2)
            cats = Reference(ws_pl, min_col=1, min_row=2, max_row=1 + len(pl))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws_pl.add_chart(chart, "K2")
        except Exception:
            logger.debug("Skipping chart on 3-statement workbook", exc_info=True)

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode("ascii")


@router.post("/three-statement")
def three_statement(body: ThreeStatementRequest, db: Session = Depends(get_db)):
    try:
        model = _project(body)
        excel_b64 = _workbook_b64(model["pl_data"], model["bs_data"], model["cf_data"], body.company_name)
        commentary = fpa_commentary(
            "Review this simplified 3-statement projection for internal consistency and CFO red flags.",
            {"assumptions": body.model_dump(), **{k: v for k, v in model.items() if k != "flags"}},
        )
        out = {
            "pl_data": model["pl_data"],
            "bs_data": model["bs_data"],
            "cf_data": model["cf_data"],
            "excel_base64": excel_b64,
            "commentary": commentary,
            "flags": model["flags"],
        }
        store_fpa_result(db, "three_statement", out, user_id=body.user_id)
        return out
    except Exception as e:
        logger.exception("three-statement failed")
        raise HTTPException(status_code=500, detail=str(e)) from e
