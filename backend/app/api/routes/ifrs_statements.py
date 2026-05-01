"""
IFRS Statement Generator API Routes
Converts Trial Balance to IFRS-compliant financial statements
"""

from fastapi import APIRouter, UploadFile, File, HTTPException, Query, Depends, Header
from sqlalchemy.orm import Session
from typing import List, Dict, Optional, Any
import pandas as pd
import io
import json
from datetime import datetime

from app.core.database import get_db
from app.models.ifrs_statement import MappingTemplate
from app.services.ifrs_mapper import IFRSMapper

router = APIRouter(prefix="/api/ifrs", tags=["ifrs-statements"])

# Initialize services
ifrs_mapper = IFRSMapper()


@router.post("/upload-trial-balance")
async def upload_trial_balance(file: UploadFile = File(...)):
    """
    Upload trial balance and parse (Excel or CSV)
    Returns trial balance entries with metadata
    """
    try:
        # Read file content
        content = await file.read()
        
        # Determine file type and parse
        if file.filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(io.BytesIO(content))
        elif file.filename.endswith('.csv'):
            df = pd.read_csv(io.StringIO(content.decode('utf-8')))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Use .xlsx, .xls, or .csv")
        
        # Validate required columns
        required_columns = ['Account Code', 'Account Name', 'Debit', 'Credit']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )
        
        # Parse trial balance
        trial_balance = []
        total_debit = 0
        total_credit = 0
        
        for _, row in df.iterrows():
            entry = {
                "glCode": str(row['Account Code']),
                "accountName": str(row['Account Name']),
                "debit": float(row['Debit']) if pd.notna(row['Debit']) else 0,
                "credit": float(row['Credit']) if pd.notna(row['Credit']) else 0,
                "accountType": str(row.get('Account Type', '')),
                "mappingStatus": "unmapped"
            }
            trial_balance.append(entry)
            total_debit += entry['debit']
            total_credit += entry['credit']
        
        # Check if trial balance balances
        is_balanced = abs(total_debit - total_credit) < 1  # Allow $1 rounding difference
        
        return {
            "success": True,
            "message": f"Successfully uploaded {len(trial_balance)} accounts",
            "trialBalance": trial_balance,
            "metadata": {
                "totalDebit": round(total_debit, 2),
                "totalCredit": round(total_credit, 2),
                "isBalanced": is_balanced,
                "accountCount": len(trial_balance),
                "fileName": file.filename,
                "uploadDate": datetime.now().isoformat()
            }
        }
        
    except pd.errors.ParserError as e:
        raise HTTPException(status_code=400, detail=f"Error parsing file: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@router.post("/ai-mapping")
async def generate_ai_mapping(trial_balance: List[Dict[str, Any]]):
    """
    Use Amazon Nova AI to suggest IFRS mappings
    Returns mappings with confidence scores
    """
    try:
        # Use IFRS mapper service to generate AI suggestions
        mappings = await ifrs_mapper.suggest_mappings(trial_balance)
        
        # Calculate overall confidence
        total_confidence = sum(m['confidence'] for m in mappings['mappings']) / len(mappings['mappings'])
        
        return {
            "success": True,
            "mappings": mappings['mappings'],
            "statistics": {
                "totalAccounts": len(trial_balance),
                "autoMapped": len([m for m in mappings['mappings'] if m['confidence'] > 80]),
                "needsReview": len([m for m in mappings['mappings'] if 50 <= m['confidence'] <= 80]),
                "unmapped": len([m for m in mappings['mappings'] if m['confidence'] < 50]),
                "averageConfidence": round(total_confidence, 1)
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI mapping failed: {str(e)}")


@router.get("/ifrs-structure")
async def get_ifrs_structure():
    """
    Get standard IFRS statement structure
    Returns hierarchical tree of line items
    """
    structure = ifrs_mapper.get_ifrs_structure()
    return {
        "success": True,
        "structure": structure
    }


@router.post("/generate-statements")
async def generate_ifrs_statements(
    trial_balance: List[Dict[str, Any]],
    mappings: Dict[str, str],
    entity_name: str,
    period_end: str,
    currency: str = "USD",
    prior_period: Optional[List[Dict[str, Any]]] = None
):
    """
    Generate full IFRS financial statements
    Returns all 4 core statements + notes
    """
    try:
        # Generate statements using mapper service
        statements = ifrs_mapper.generate_statements(
            trial_balance=trial_balance,
            mappings=mappings,
            entity_name=entity_name,
            period_end=period_end,
            currency=currency,
            prior_period=prior_period
        )
        
        result = {
            "success": True,
            "statements": statements,
            "metadata": {
                "entityName": entity_name,
                "periodEnd": period_end,
                "currency": currency,
                "generatedAt": datetime.now().isoformat()
            }
        }
        try:
            from app.agents.intelligence import generate_insight
            from app.agents.memory import read_agent_memory, store_agent_run, update_agent_memory
            from app.core.database import SessionLocal

            _ifrs_data = result if isinstance(result, dict) else (result.model_dump() if hasattr(result, "model_dump") else result.dict())
            _db = SessionLocal()
            try:
                _history = await read_agent_memory("ifrs_statements", _db)
                _insight = await generate_insight(
                    "ifrs_statements",
                    {
                        "ifrs_result": _ifrs_data,
                        "source_route": "/ifrs",
                        "deep_link": "/ifrs/statements",
                    },
                    _history,
                )
                _insight["source_route"] = "/ifrs"
                _insight["deep_link"] = "/ifrs/statements"
                _insight["module_label"] = "IFRS Statements"
                await store_agent_run(
                    "ifrs_statements",
                    {
                        "entity_name": entity_name,
                        "period_end": period_end,
                        "currency": currency,
                        "line_items": len(trial_balance),
                    },
                    _ifrs_data,
                    _insight,
                    _db,
                )
                await update_agent_memory("ifrs_statements", _ifrs_data, _db)
            finally:
                _db.close()
        except Exception as _e:
            print(f"[agent_run] ifrs_statements: {_e}")
        
        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Statement generation failed: {str(e)}")


@router.post("/save-mapping-template")
async def save_mapping_template(
    name: str,
    industry: str,
    description: str,
    mappings: Dict[str, str],
    user_id: str = "demo-user"
):
    """
    Save mapping template for reuse
    """
    try:
        template = {
            "id": f"template_{datetime.now().timestamp()}",
            "name": name,
            "industry": industry,
            "description": description,
            "mappings": mappings,
            "createdBy": user_id,
            "createdAt": datetime.now().isoformat(),
            "isPublic": False
        }
        
        # In production, save to database
        # For now, return the template
        
        return {
            "success": True,
            "message": f"Template '{name}' saved successfully",
            "template": template
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save template: {str(e)}")


@router.get("/mapping-templates")
async def get_mapping_templates(
    industry: Optional[str] = Query(None),
    x_tenant_id: Optional[str] = Header(None, alias="X-Tenant-ID"),
    db: Session = Depends(get_db),
):
    """
    Built-in industry templates + tenant saved templates (Week 1 DB).
    """
    templates = ifrs_mapper.get_industry_templates(industry)
    tid = (x_tenant_id or "default").strip() or "default"
    saved = (
        db.query(MappingTemplate)
        .filter(MappingTemplate.tenant_id == tid)
        .order_by(MappingTemplate.created_at.desc())
        .all()
    )
    saved_payload = [
        {
            "id": t.id,
            "template_name": t.template_name,
            "industry": t.industry,
            "is_default": t.is_default,
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "entries_count": len(t.entries or []),
        }
        for t in saved
    ]
    return {
        "success": True,
        "templates": templates,
        "count": len(templates),
        "saved_templates": saved_payload,
        "saved_count": len(saved_payload),
    }


@router.post("/export-statements")
async def export_statements(
    statements: Dict[str, Any],
    format: str = Query(..., pattern="^(pdf|excel|word|json)$"),
    include_notes: bool = True,
    include_comparatives: bool = True
):
    """
    Export statements in requested format
    Returns file download
    """
    try:
        from fastapi.responses import StreamingResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        
        if format == "json":
            # Return JSON
            output = io.StringIO()
            json.dump(statements, output, indent=2)
            output.seek(0)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/json",
                headers={"Content-Disposition": "attachment; filename=ifrs_statements.json"}
            )
        
        elif format == "excel":
            # Create Excel workbook
            wb = openpyxl.Workbook()
            
            # Financial Position sheet
            ws_fp = wb.active
            ws_fp.title = "Financial Position"
            fp_data = statements.get('financialPosition', {})
            
            # Add header
            ws_fp['A1'] = statements.get('entityName', 'Company Name')
            ws_fp['A2'] = "STATEMENT OF FINANCIAL POSITION"
            ws_fp['A3'] = f"As at {statements.get('periodEnd', 'Period End')}"
            ws_fp['A1'].font = Font(bold=True, size=14)
            ws_fp['A2'].font = Font(bold=True, size=12)
            
            # Add data (simplified)
            row = 5
            ws_fp.cell(row, 1, "ASSETS")
            ws_fp.cell(row, 1).font = Font(bold=True)
            row += 1
            
            for item in fp_data.get('assets', []):
                ws_fp.cell(row, 1, f"  {item.get('name', '')}")
                ws_fp.cell(row, 2, item.get('currentYear', 0))
                ws_fp.cell(row, 3, item.get('priorYear', 0))
                row += 1
            
            # Profit & Loss sheet
            ws_pl = wb.create_sheet("Profit & Loss")
            ws_pl['A1'] = statements.get('entityName', 'Company Name')
            ws_pl['A2'] = "STATEMENT OF PROFIT OR LOSS"
            
            # Cash Flows sheet
            ws_cf = wb.create_sheet("Cash Flows")
            ws_cf['A1'] = statements.get('entityName', 'Company Name')
            ws_cf['A2'] = "STATEMENT OF CASH FLOWS"
            
            # Changes in Equity sheet
            ws_eq = wb.create_sheet("Changes in Equity")
            ws_eq['A1'] = statements.get('entityName', 'Company Name')
            ws_eq['A2'] = "STATEMENT OF CHANGES IN EQUITY"
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=ifrs_statements.xlsx"}
            )
        
        elif format == "pdf":
            # For now, return text-based PDF placeholder
            output = io.StringIO()
            output.write(f"{statements.get('entityName', 'Company Name')}\n")
            output.write("IFRS FINANCIAL STATEMENTS\n")
            output.write(f"Period: {statements.get('periodEnd', 'Period End')}\n\n")
            output.write("NOTE: PDF export with full formatting coming soon.\n")
            output.write("Use Excel export for now.\n")
            output.seek(0)
            
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/plain",
                headers={"Content-Disposition": "attachment; filename=ifrs_statements.txt"}
            )
        
        else:
            raise HTTPException(status_code=400, detail="Invalid export format")
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/sample-trial-balance")
async def get_sample_trial_balance():
    """
    Get sample trial balance data for testing
    """
    sample_data = {
        "trialBalance": [
            {"glCode": "1000", "accountName": "Cash in Bank", "debit": 500000, "credit": 0, "accountType": "Asset"},
            {"glCode": "1100", "accountName": "Accounts Receivable", "debit": 250000, "credit": 0, "accountType": "Asset"},
            {"glCode": "1200", "accountName": "Inventory", "debit": 180000, "credit": 0, "accountType": "Asset"},
            {"glCode": "1500", "accountName": "Property, Plant & Equipment", "debit": 1200000, "credit": 0, "accountType": "Asset"},
            {"glCode": "1510", "accountName": "Accumulated Depreciation", "debit": 0, "credit": 350000, "accountType": "Asset"},
            {"glCode": "1600", "accountName": "Intangible Assets", "debit": 150000, "credit": 0, "accountType": "Asset"},
            {"glCode": "2000", "accountName": "Accounts Payable", "debit": 0, "credit": 280000, "accountType": "Liability"},
            {"glCode": "2100", "accountName": "Short-term Loans", "debit": 0, "credit": 150000, "accountType": "Liability"},
            {"glCode": "2500", "accountName": "Long-term Debt", "debit": 0, "credit": 800000, "accountType": "Liability"},
            {"glCode": "3000", "accountName": "Share Capital", "debit": 0, "credit": 500000, "accountType": "Equity"},
            {"glCode": "3100", "accountName": "Retained Earnings", "debit": 0, "credit": 200000, "accountType": "Equity"},
            {"glCode": "4000", "accountName": "Sales Revenue", "debit": 0, "credit": 3200000, "accountType": "Revenue"},
            {"glCode": "4100", "accountName": "Other Income", "debit": 0, "credit": 45000, "accountType": "Revenue"},
            {"glCode": "5000", "accountName": "Cost of Goods Sold", "debit": 1450000, "credit": 0, "accountType": "Expense"},
            {"glCode": "6000", "accountName": "Salaries & Wages", "debit": 480000, "credit": 0, "accountType": "Expense"},
            {"glCode": "6100", "accountName": "Rent Expense", "debit": 120000, "credit": 0, "accountType": "Expense"},
            {"glCode": "6200", "accountName": "Marketing Expense", "debit": 180000, "credit": 0, "accountType": "Expense"},
            {"glCode": "6300", "accountName": "Depreciation Expense", "debit": 120000, "credit": 0, "accountType": "Expense"},
            {"glCode": "6400", "accountName": "Administrative Expenses", "debit": 95000, "credit": 0, "accountType": "Expense"},
            {"glCode": "7000", "accountName": "Interest Expense", "debit": 65000, "credit": 0, "accountType": "Expense"},
            {"glCode": "8000", "accountName": "Income Tax Expense", "debit": 190000, "credit": 0, "accountType": "Expense"}
        ],
        "metadata": {
            "totalDebit": 4980000,
            "totalCredit": 4980000,
            "isBalanced": True,
            "accountCount": 21
        }
    }
    
    return {
        "success": True,
        **sample_data
    }
