from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import JSONResponse
from typing import Optional, Dict, Any
from datetime import datetime, timedelta
import random
import os
import pandas as pd
import json
from app.services.cfo_metrics_calculator import CFOMetricsCalculator
from app.services.accounting_integrations import AccountingIntegrationFactory
from app.services.trial_balance_parser import TrialBalanceParser

router = APIRouter(prefix="/api/cfo", tags=["cfo-dashboard"])

# Path to financial data files
JOURNAL_ENTRIES_PATH = "sample_journal_entries.csv"
TRIAL_BALANCE_PATH = "trial_balance.csv"

@router.get("/dashboard")
async def get_dashboard_data(
    time_range: str = Query("month", enum=["week", "month", "quarter", "year"])
):
    """
    Get comprehensive CFO dashboard data
    
    Priority:
    1. Trial Balance (if exists)
    2. Journal Entries (if exists)
    3. Mock data (fallback)
    """
    
    try:
        # Priority 1: Try Trial Balance (most direct for CFO dashboard)
        if os.path.exists(TRIAL_BALANCE_PATH):
            parser = TrialBalanceParser(TRIAL_BALANCE_PATH)
            data = parser.calculate_dashboard_metrics()
            print(f"✅ Using Trial Balance data")
            return data
        
        # Priority 2: Try Journal Entries
        if os.path.exists(JOURNAL_ENTRIES_PATH):
            calculator = CFOMetricsCalculator(JOURNAL_ENTRIES_PATH)
            data = calculator.calculate_all_metrics(time_range)
            print(f"✅ Using Journal Entries data")
            return data
    except Exception as e:
        print(f"Error calculating metrics: {e}")
        import traceback
        traceback.print_exc()
    
    # Fallback to mock data
    data = {
        "healthScore": {
            "overall": 69,
            "trend": 2.5,
            "breakdown": {
                "liquidity": 66,
                "profitability": 68,
                "efficiency": 62,
                "stability": 71
            }
        },
        "cash": {
            "current": 562000,
            "trend": 6.8,
            "runway": 18.5,
            "history": [520000, 535000, 548000, 562000, 575000, 580000, 562000]
        },
        "revenue": {
            "monthly": 328000,
            "arr": 3936000,
            "growth": 25,
            "history": [245000, 268000, 291000, 308000, 315000, 328000]
        },
        "expenses": {
            "monthly": 234000,
            "trend": 8.2,
            "categories": [
                {"name": "Operations", "value": 105000, "percentage": 45},
                {"name": "Marketing", "value": 58000, "percentage": 25},
                {"name": "Sales", "value": 47000, "percentage": 20},
                {"name": "R&D", "value": 24000, "percentage": 10}
            ]
        },
        "insights": [
            {"icon": "💰", "text": "Cash runway trending down - consider cost optimization", "severity": "warning"},
            {"icon": "📊", "text": "AR aging increased 8 days - follow up with top 3 customers", "severity": "medium"},
            {"icon": "📈", "text": "Marketing ROI at 340% - recommend budget increase", "severity": "info"},
            {"icon": "📦", "text": "Inventory turnover slowing - review stock levels", "severity": "medium"}
        ],
        "alerts": [
            {"severity": "critical", "message": "Cash forecast shows potential shortfall in Week 11", "time": "2 hours ago", "action": "Review"},
            {"severity": "warning", "message": "Customer XYZ payment overdue by 15 days", "time": "5 hours ago", "action": "Follow up"},
            {"severity": "info", "message": "Monthly financial close completed", "time": "1 day ago", "action": "View"}
        ],
        "recentActivity": [
            {"icon": "📄", "action": "Board report generated", "time": "2 hours ago"},
            {"icon": "💰", "action": "Cash flow forecast updated", "time": "5 hours ago"},
            {"icon": "📊", "action": "P&L analysis completed", "time": "1 day ago"},
            {"icon": "🔍", "action": "Fraud detection scan finished", "time": "1 day ago"},
            {"icon": "📈", "action": "Q4 variance report created", "time": "2 days ago"}
        ],
        "recommendations": [
            {"priority": "high", "text": "Reduce marketing spend by 15% to extend runway", "impact": "+2.5 months runway"},
            {"priority": "medium", "text": "Negotiate payment terms with top 3 vendors", "impact": "$50K cash flow improvement"},
            {"priority": "low", "text": "Review SaaS subscriptions for optimization", "impact": "$5K/month savings"}
        ],
        "ratios": {
            "currentRatio": 2.4,
            "quickRatio": 1.8,
            "debtToEquity": 0.3,
            "roe": 18,
            "operatingMargin": 28.7
        }
    }
    
    return data

@router.post("/export")
async def export_dashboard(
    format: str = Query(..., enum=["pdf", "excel", "csv"]),
    time_range: str = Query("month")
):
    """
    Export dashboard data in specified format
    """
    
    try:
        # Get dashboard data
        try:
            if os.path.exists(JOURNAL_ENTRIES_PATH):
                calculator = CFOMetricsCalculator(JOURNAL_ENTRIES_PATH)
                dashboard_data = calculator.calculate_all_metrics(time_range)
            else:
                # Use mock data
                dashboard_data = _get_mock_data()
        except Exception as e:
            print(f"Error getting data: {e}")
            dashboard_data = _get_mock_data()
        
        # Generate filename
        filename = f"cfo-dashboard-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        
        if format == "csv":
            # Generate CSV
            return _export_csv(dashboard_data, filename)
        
        elif format == "excel":
            # Generate Excel
            return _export_excel(dashboard_data, filename)
        
        elif format == "pdf":
            # Generate PDF (simplified - just text for now)
            return _export_pdf(dashboard_data, filename)
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

def _export_csv(data: Dict[str, Any], filename: str):
    """Export dashboard data as CSV"""
    from fastapi.responses import StreamingResponse
    import io
    
    # Create CSV content
    output = io.StringIO()
    output.write("CFO Dashboard Export\n")
    output.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
    
    # Health Score
    output.write("Financial Health Score\n")
    output.write(f"Overall Score,{data['healthScore']['overall']}\n")
    output.write(f"Liquidity,{data['healthScore']['breakdown']['liquidity']}\n")
    output.write(f"Profitability,{data['healthScore']['breakdown']['profitability']}\n")
    output.write(f"Efficiency,{data['healthScore']['breakdown']['efficiency']}\n")
    output.write(f"Stability,{data['healthScore']['breakdown']['stability']}\n\n")
    
    # Cash
    output.write("Cash Position\n")
    output.write(f"Current Balance,${data['cash']['current']:,}\n")
    output.write(f"Runway (months),{data['cash']['runway']}\n")
    output.write(f"Trend,{data['cash']['trend']}%\n\n")
    
    # Revenue
    output.write("Revenue\n")
    output.write(f"Monthly Revenue,${data['revenue']['monthly']:,}\n")
    output.write(f"Annual Run Rate,${data['revenue']['arr']:,}\n")
    output.write(f"Growth,{data['revenue']['growth']}%\n\n")
    
    # Expenses
    output.write("Expenses\n")
    output.write(f"Monthly Expenses,${data['expenses']['monthly']:,}\n")
    output.write(f"Trend,{data['expenses']['trend']}%\n\n")
    
    output.write("Expense Categories\n")
    for cat in data['expenses']['categories']:
        output.write(f"{cat['name']},${cat['value']:,},{cat['percentage']}%\n")
    
    output.write("\n")
    
    # Financial Ratios
    output.write("Financial Ratios\n")
    output.write(f"Current Ratio,{data['ratios']['currentRatio']}\n")
    output.write(f"Quick Ratio,{data['ratios']['quickRatio']}\n")
    output.write(f"Debt-to-Equity,{data['ratios']['debtToEquity']}\n")
    output.write(f"ROE,{data['ratios']['roe']}%\n")
    output.write(f"Operating Margin,{data['ratios']['operatingMargin']}%\n")
    
    # Create response
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
    )

def _export_excel(data: Dict[str, Any], filename: str):
    """Export dashboard data as Excel"""
    from fastapi.responses import StreamingResponse
    import io
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CFO Dashboard"
        
        # Title
        ws['A1'] = "CFO Dashboard Export"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        row = 4
        
        # Health Score
        ws[f'A{row}'] = "Financial Health Score"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        ws[f'A{row}'] = "Overall Score"
        ws[f'B{row}'] = data['healthScore']['overall']
        row += 1
        ws[f'A{row}'] = "Liquidity"
        ws[f'B{row}'] = data['healthScore']['breakdown']['liquidity']
        row += 1
        ws[f'A{row}'] = "Profitability"
        ws[f'B{row}'] = data['healthScore']['breakdown']['profitability']
        row += 1
        ws[f'A{row}'] = "Efficiency"
        ws[f'B{row}'] = data['healthScore']['breakdown']['efficiency']
        row += 1
        ws[f'A{row}'] = "Stability"
        ws[f'B{row}'] = data['healthScore']['breakdown']['stability']
        row += 2
        
        # Cash
        ws[f'A{row}'] = "Cash Position"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        ws[f'A{row}'] = "Current Balance"
        ws[f'B{row}'] = f"${data['cash']['current']:,}"
        row += 1
        ws[f'A{row}'] = "Runway (months)"
        ws[f'B{row}'] = data['cash']['runway']
        row += 2
        
        # Revenue
        ws[f'A{row}'] = "Revenue"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        ws[f'A{row}'] = "Monthly Revenue"
        ws[f'B{row}'] = f"${data['revenue']['monthly']:,}"
        row += 1
        ws[f'A{row}'] = "Annual Run Rate"
        ws[f'B{row}'] = f"${data['revenue']['arr']:,}"
        row += 1
        ws[f'A{row}'] = "Growth"
        ws[f'B{row}'] = f"{data['revenue']['growth']}%"
        row += 2
        
        # Expenses
        ws[f'A{row}'] = "Expenses"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        ws[f'A{row}'] = "Monthly Expenses"
        ws[f'B{row}'] = f"${data['expenses']['monthly']:,}"
        row += 2
        
        ws[f'A{row}'] = "Category"
        ws[f'B{row}'] = "Amount"
        ws[f'C{row}'] = "Percentage"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'].font = Font(bold=True)
        row += 1
        
        for cat in data['expenses']['categories']:
            ws[f'A{row}'] = cat['name']
            ws[f'B{row}'] = f"${cat['value']:,}"
            ws[f'C{row}'] = f"{cat['percentage']}%"
            row += 1
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )
    
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        raise HTTPException(status_code=500, detail="Excel export requires openpyxl library. Please install: pip install openpyxl")

def _export_pdf(data: Dict[str, Any], filename: str):
    """Export dashboard data as PDF (simplified text-based)"""
    from fastapi.responses import Response
    
    # Simple text-based PDF alternative (plain text)
    content = f"""CFO DASHBOARD EXPORT
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

FINANCIAL HEALTH SCORE: {data['healthScore']['overall']}/100
  - Liquidity: {data['healthScore']['breakdown']['liquidity']}
  - Profitability: {data['healthScore']['breakdown']['profitability']}
  - Efficiency: {data['healthScore']['breakdown']['efficiency']}
  - Stability: {data['healthScore']['breakdown']['stability']}

CASH POSITION: ${data['cash']['current']:,}
  - Runway: {data['cash']['runway']} months
  - Trend: {data['cash']['trend']}%

REVENUE
  - Monthly: ${data['revenue']['monthly']:,}
  - Annual Run Rate: ${data['revenue']['arr']:,}
  - Growth: {data['revenue']['growth']}% YoY

EXPENSES
  - Monthly: ${data['expenses']['monthly']:,}
  - Trend: {data['expenses']['trend']}%

EXPENSE CATEGORIES:
"""
    
    for cat in data['expenses']['categories']:
        content += f"  - {cat['name']}: ${cat['value']:,} ({cat['percentage']}%)\n"
    
    content += f"""
FINANCIAL RATIOS
  - Current Ratio: {data['ratios']['currentRatio']}
  - Quick Ratio: {data['ratios']['quickRatio']}
  - Debt-to-Equity: {data['ratios']['debtToEquity']}
  - ROE: {data['ratios']['roe']}%
  - Operating Margin: {data['ratios']['operatingMargin']}%
"""
    
    return Response(
        content=content.encode('utf-8'),
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={filename}.txt"}
    )

def _get_mock_data():
    """Get mock dashboard data"""
    return {
        "healthScore": {
            "overall": 69,
            "trend": 2.5,
            "breakdown": {
                "liquidity": 66,
                "profitability": 68,
                "efficiency": 62,
                "stability": 71
            }
        },
        "cash": {
            "current": 562000,
            "trend": 6.8,
            "runway": 18.5,
            "history": [520000, 535000, 548000, 562000, 575000, 580000, 562000]
        },
        "revenue": {
            "monthly": 328000,
            "arr": 3936000,
            "growth": 25,
            "history": [245000, 268000, 291000, 308000, 315000, 328000]
        },
        "expenses": {
            "monthly": 234000,
            "trend": 8.2,
            "categories": [
                {"name": "Operations", "value": 105000, "percentage": 45},
                {"name": "Marketing", "value": 58000, "percentage": 25},
                {"name": "Sales", "value": 47000, "percentage": 20},
                {"name": "R&D", "value": 24000, "percentage": 10}
            ]
        },
        "insights": [],
        "alerts": [],
        "recentActivity": [],
        "recommendations": [],
        "ratios": {
            "currentRatio": 2.4,
            "quickRatio": 1.8,
            "debtToEquity": 0.3,
            "roe": 18,
            "operatingMargin": 28.7
        }
    }

@router.get("/insights")
async def get_ai_insights():
    """
    Get AI-generated insights using Amazon Nova
    
    This endpoint would integrate with the existing nova_service
    to generate personalized financial insights
    """
    
    insights = [
        {
            "category": "cash_flow",
            "severity": "warning",
            "title": "Cash Runway Concern",
            "description": "Current burn rate of $234K/month gives 18.5 months runway. Consider cost optimization.",
            "recommendations": [
                "Reduce marketing spend by 15%",
                "Negotiate better payment terms with vendors",
                "Review and cancel unused SaaS subscriptions"
            ],
            "impact": "+2.5 months runway extension"
        },
        {
            "category": "revenue",
            "severity": "info",
            "title": "Strong Growth Trajectory",
            "description": "Revenue growing at 25% YoY. Marketing ROI is 340%.",
            "recommendations": [
                "Consider increasing marketing budget",
                "Expand to new customer segments",
                "Launch referral program"
            ],
            "impact": "Potential +15% revenue increase"
        },
        {
            "category": "efficiency",
            "severity": "medium",
            "title": "AR Collection Slowing",
            "description": "Accounts receivable aging increased by 8 days.",
            "recommendations": [
                "Follow up with top 3 customers",
                "Implement automated payment reminders",
                "Review payment terms for new contracts"
            ],
            "impact": "$75K cash flow improvement"
        }
    ]
    
    return {"insights": insights, "generated_at": datetime.now().isoformat()}

@router.post("/forecast")
async def generate_forecast(
    period: str = Query("quarter", enum=["month", "quarter", "year"]),
    type: str = Query("cash_flow", enum=["cash_flow", "revenue", "expenses"])
):
    """
    Generate financial forecast using AI/ML models
    
    This would use:
    - Historical data analysis
    - Time series forecasting (ARIMA, Prophet)
    - Amazon Nova for insights
    """
    
    # Generate mock forecast data
    base_value = 328000 if type == "revenue" else 234000
    forecast_periods = {
        "month": 1,
        "quarter": 3,
        "year": 12
    }
    
    periods = forecast_periods[period]
    forecast_data = []
    
    for i in range(periods):
        # Simulate trend
        growth_rate = random.uniform(0.02, 0.08)
        value = base_value * (1 + growth_rate * (i + 1))
        
        forecast_data.append({
            "period": i + 1,
            "value": round(value, 2),
            "lower_bound": round(value * 0.9, 2),
            "upper_bound": round(value * 1.1, 2),
            "confidence": random.uniform(0.75, 0.95)
        })
    
    return {
        "success": True,
        "forecast_type": type,
        "period": period,
        "data": forecast_data,
        "overall_confidence": 0.85,
        "trend": "increasing" if type == "revenue" else "stable",
        "generated_at": datetime.now().isoformat()
    }

@router.get("/metrics")
async def get_key_metrics():
    """
    Get real-time key financial metrics
    """
    
    return {
        "cash": {
            "balance": 562000,
            "change_24h": 12500,
            "change_percent": 2.3
        },
        "revenue": {
            "mtd": 185000,
            "target": 328000,
            "progress": 56.4
        },
        "expenses": {
            "mtd": 142000,
            "budget": 234000,
            "utilization": 60.7
        },
        "runway": {
            "months": 18.5,
            "status": "warning",
            "trend": "declining"
        }
    }

@router.post("/chat")
async def cfo_ai_chat(query: str):
    """
    Chat with CFO AI assistant powered by Amazon Nova
    
    This would integrate with the existing Nova service
    for conversational financial insights
    """
    
    # Mock response - replace with actual Nova integration
    responses = {
        "cash": "Based on current data, your cash position is $562K with an 18.5 month runway. I recommend focusing on extending this by optimizing marketing spend.",
        "revenue": "Your revenue is growing at 25% YoY, which is strong. Marketing ROI is 340%, suggesting potential for increased investment.",
        "default": "I'm here to help with financial insights. You can ask about cash flow, revenue, expenses, forecasts, or strategic recommendations."
    }
    
    # Simple keyword matching - replace with Nova
    response_text = responses.get("default")
    if "cash" in query.lower():
        response_text = responses["cash"]
    elif "revenue" in query.lower() or "sales" in query.lower():
        response_text = responses["revenue"]
    
    return {
        "response": response_text,
        "confidence": 0.92,
        "sources": ["Cash Flow Analysis", "Revenue Report", "AI Insights"],
        "timestamp": datetime.now().isoformat()
    }

@router.post("/upload/trial-balance")
async def upload_trial_balance(file: UploadFile = File(...)):
    """
    Upload Trial Balance (CSV or Excel) to calculate CFO dashboard metrics
    
    Accepts: CSV, XLSX, XLS files
    Returns: Calculated dashboard metrics
    
    Expected columns:
    - Account Code (or Account_Code)
    - Account Name (or Account_Name)
    - Account Type (or Account_Type): Asset, Liability, Equity, Revenue, Expense
    - Debit
    - Credit
    """
    
    try:
        # Save uploaded file temporarily
        temp_path = f"temp_{file.filename}"
        
        with open(temp_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        # Calculate metrics from uploaded trial balance
        parser = TrialBalanceParser(temp_path)
        metrics = parser.calculate_dashboard_metrics()
        
        # Save as the main trial balance file
        if temp_path.endswith('.csv'):
            import shutil
            shutil.copy(temp_path, TRIAL_BALANCE_PATH)
        else:
            # Convert Excel to CSV
            df = pd.read_excel(temp_path)
            df.to_csv(TRIAL_BALANCE_PATH, index=False)
        
        # Clean up temp file
        os.remove(temp_path)
        
        return {
            "success": True,
            "message": f"Successfully processed Trial Balance: {file.filename}",
            "data": metrics,
            "file_type": "trial_balance"
        }
    
    except Exception as e:
        # Clean up temp file if it exists
        if 'temp_path' in locals() and os.path.exists(temp_path):
            os.remove(temp_path)
        
        raise HTTPException(status_code=400, detail=f"Error processing Trial Balance: {str(e)}")

@router.post("/upload/journal-entries")
async def upload_journal_entries(file: UploadFile = File(...)):
    """
    Upload journal entries (CSV or Excel) to calculate CFO dashboard metrics
    
    Accepts: CSV, XLSX, XLS files
    Returns: Calculated dashboard metrics
    """
    
    try:
        # Save uploaded file temporarily
        temp_path = f"temp_{file.filename}"
        
        with open(temp_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        # Calculate metrics from uploaded file
        calculator = CFOMetricsCalculator(temp_path)
        metrics = calculator.calculate_all_metrics(time_range="month")
        
        # Clean up temp file
        os.remove(temp_path)
        
        return {
            "success": True,
            "message": f"Successfully processed {file.filename}",
            "data": metrics,
            "file_type": "journal_entries"
        }
    
    except Exception as e:
        # Clean up temp file if it exists
        if 'temp_path' in locals() and os.path.exists(temp_path):
            os.remove(temp_path)
        
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

@router.post("/upload/financial-summary")
async def upload_financial_summary(file: UploadFile = File(...)):
    """
    Upload financial summary (JSON, CSV, or Excel)
    
    Expected format:
    - JSON: Complete dashboard data structure
    - CSV/Excel: Financial summary with columns: metric, value, date
    """
    
    try:
        content = await file.read()
        
        if file.filename.endswith('.json'):
            # Parse JSON directly
            data = json.loads(content)
            return {
                "success": True,
                "message": "Financial summary uploaded successfully",
                "data": data
            }
        
        elif file.filename.endswith(('.csv', '.xlsx', '.xls')):
            # Save and parse CSV/Excel
            temp_path = f"temp_{file.filename}"
            with open(temp_path, "wb") as f:
                f.write(content)
            
            # Read with pandas
            if file.filename.endswith('.csv'):
                df = pd.read_csv(temp_path)
            else:
                df = pd.read_excel(temp_path)
            
            # Convert to dashboard format
            # Expected columns: metric, value, date
            dashboard_data = self._convert_summary_to_dashboard(df)
            
            os.remove(temp_path)
            
            return {
                "success": True,
                "message": "Financial summary processed successfully",
                "data": dashboard_data
            }
        
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Use JSON, CSV, or Excel")
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

@router.post("/connect/accounting-software")
async def connect_accounting_software(
    platform: str = Query(..., enum=["quickbooks", "xero"]),
    access_token: str = Query(...),
    company_id: Optional[str] = Query(None)
):
    """
    Connect to accounting software (QuickBooks, Xero, etc.)
    
    Parameters:
    - platform: "quickbooks" or "xero"
    - access_token: OAuth access token
    - company_id: Company/Tenant ID (required for QuickBooks)
    """
    
    try:
        # Set environment variables for the integration
        if platform == "quickbooks":
            os.environ["QUICKBOOKS_ACCESS_TOKEN"] = access_token
            os.environ["QUICKBOOKS_COMPANY_ID"] = company_id or ""
        elif platform == "xero":
            os.environ["XERO_ACCESS_TOKEN"] = access_token
            os.environ["XERO_TENANT_ID"] = company_id or ""
        
        # Get integration
        integration = AccountingIntegrationFactory.get_integration(platform)
        
        # Fetch data
        dashboard_data = integration.get_dashboard_data(time_range="month")
        
        return {
            "success": True,
            "message": f"Successfully connected to {platform.title()}",
            "data": dashboard_data
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Connection failed: {str(e)}")

@router.get("/data-sources")
async def get_available_data_sources():
    """
    Get list of available data sources and their status
    """
    
    sources = []
    
    # Check journal entries file
    if os.path.exists(JOURNAL_ENTRIES_PATH):
        sources.append({
            "type": "journal_entries",
            "name": "Journal Entries CSV",
            "status": "connected",
            "file": JOURNAL_ENTRIES_PATH,
            "last_updated": datetime.fromtimestamp(os.path.getmtime(JOURNAL_ENTRIES_PATH)).isoformat()
        })
    
    # Check QuickBooks connection
    if os.getenv("QUICKBOOKS_ACCESS_TOKEN"):
        sources.append({
            "type": "quickbooks",
            "name": "QuickBooks Online",
            "status": "connected",
            "company_id": os.getenv("QUICKBOOKS_COMPANY_ID")
        })
    
    # Check Xero connection
    if os.getenv("XERO_ACCESS_TOKEN"):
        sources.append({
            "type": "xero",
            "name": "Xero",
            "status": "connected",
            "tenant_id": os.getenv("XERO_TENANT_ID")
        })
    
    if not sources:
        sources.append({
            "type": "mock",
            "name": "Mock Data",
            "status": "active",
            "description": "Using sample data for demonstration"
        })
    
    return {
        "sources": sources,
        "primary_source": sources[0]["type"] if sources else "mock"
    }

def _convert_summary_to_dashboard(df: pd.DataFrame) -> Dict[str, Any]:
    """Convert financial summary DataFrame to dashboard format"""
    
    # This is a helper function to convert uploaded CSV/Excel to dashboard format
    # Expected columns: metric, value, date (optional)
    
    dashboard_data = {
        "healthScore": {"overall": 69, "trend": 2.5, "breakdown": {}},
        "cash": {"current": 0, "trend": 0, "runway": 0, "history": []},
        "revenue": {"monthly": 0, "arr": 0, "growth": 0, "history": []},
        "expenses": {"monthly": 0, "trend": 0, "categories": []},
        "insights": [],
        "alerts": [],
        "recentActivity": [],
        "recommendations": [],
        "ratios": {}
    }
    
    # Parse DataFrame and populate dashboard_data
    for _, row in df.iterrows():
        metric = row.get('metric', '').lower()
        value = row.get('value', 0)
        
        if 'cash' in metric:
            dashboard_data['cash']['current'] = float(value)
        elif 'revenue' in metric:
            dashboard_data['revenue']['monthly'] = float(value)
        elif 'expense' in metric:
            dashboard_data['expenses']['monthly'] = float(value)
    
    return dashboard_data
