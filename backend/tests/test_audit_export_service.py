"""Unit tests for audit export service."""
from __future__ import annotations

import hashlib
import io
import json
import sys
import unittest
import uuid
import zipfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

_PORTED = Path(__file__).resolve().parents[1] / "app" / "modules" / "gulftax" / "ported"
if str(_PORTED) not in sys.path:
    sys.path.insert(0, str(_PORTED))

from app.core.database import Base
from app.models.client_data import (
    BadDebtReliefClaim,
    CtReturn,
    DesignatedZoneTransaction,
    GulftaxTransaction,
    PartialExemptionCalculation,
)
from app.models.workspace_audit import WorkspaceAuditLog
from models import Company, ReconciliationResult
from app.services.audit_export_service import generate_period_audit_pack


def _mock_vat_boxes() -> dict:
    return {
        "box1_standard_rated_sales_net": 10000.0,
        "box1_standard_rated_sales_vat": 500.0,
        "box2_tourist_refunds": 0.0,
        "box2_advance_payment_vat": 0.0,
        "box3_reverse_charge_supplies_net": 0.0,
        "box3_reverse_charge_supplies_vat": 0.0,
        "box4_zero_rated_supplies": 0.0,
        "box5_exempt_supplies": 0.0,
        "box6_imports_vat": 0.0,
        "box7_output_adjustments": 0.0,
        "box8_total_output_vat": 500.0,
        "box9_standard_rated_expenses": 2000.0,
        "box10_reverse_charge_expenses": 0.0,
        "box11_total_input_vat": 100.0,
        "box11_total_input_vat_raw": 100.0,
        "box12_net_vat_payable_or_refundable": 400.0,
        "partial_exemption_applied": False,
        "recovery_percentage": 100.0,
        "bad_debt_relief_applied": 0.0,
        "sales_invoice_count": 1,
        "purchase_entry_count": 1,
        "source": "gulftax_transactions",
    }


class AuditExportServiceTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(
            cls.engine,
            tables=[
                GulftaxTransaction.__table__,
                CtReturn.__table__,
                WorkspaceAuditLog.__table__,
                PartialExemptionCalculation.__table__,
                BadDebtReliefClaim.__table__,
                DesignatedZoneTransaction.__table__,
                Company.__table__,
                ReconciliationResult.__table__,
            ],
        )
        cls.Session = sessionmaker(bind=cls.engine)

    def setUp(self):
        self.db = self.Session()
        self.ported_db = self.Session()
        self.tenant = "ws-audit-export"
        self.company_id = str(uuid.uuid4())
        self.period = "2026-Q1"
        self.period_start = date(2026, 1, 1)
        self.period_end = date(2026, 3, 31)

        self.ported_db.add(
            Company(
                id=self.company_id,
                name="Audit Export Test Co",
                trade_license_number="TL-AUDIT",
                trn="100000000000003",
                entity_type="mainland",
            )
        )
        self.ported_db.commit()

        self.db.add(
            GulftaxTransaction(
                id=str(uuid.uuid4()),
                tenant_id=self.tenant,
                company_id=self.company_id,
                source="ap_invoiceflow",
                ap_invoice_id=str(uuid.uuid4()),
                tax_period=self.period,
                transaction_date=date(2026, 2, 15),
                vendor_name="Vendor A",
                vendor_trn="100000000000003",
                invoice_number="INV-001",
                gross_amount=Decimal("1050.00"),
                vat_amount=Decimal("50.00"),
                vat_category="standard",
                fta_box="box9",
                direction="input",
                status="posted",
            )
        )
        self.db.add(
            CtReturn(
                tenant_id=self.tenant,
                company_id=self.company_id,
                period_start=self.period_start,
                period_end=self.period_end,
                revenue=Decimal("500000"),
                accounting_profit=Decimal("50000"),
                taxable_income=Decimal("48000"),
                ct_payable_aed=Decimal("4050"),
                sbr_eligible=False,
                qfzp_eligible=False,
                free_zone_status="mainland",
                status="approved",
                approved_at=datetime.utcnow(),
            )
        )
        self.db.add(
            WorkspaceAuditLog(
                workspace_id=self.tenant,
                company_id=self.company_id,
                action="je_posted",
                entity_type="journal_entry",
                entity_id="je-1",
                user_email="auditor@test.com",
                details={"note": "test"},
                created_at=datetime(2026, 2, 1, 10, 0, 0),
            )
        )
        self.db.commit()

        self.ported_db.add(
            ReconciliationResult(
                company_id=self.company_id,
                tax_period=self.period,
                period_start=self.period_start,
                period_end=self.period_end,
                difference_aed=0.0,
                status="matched",
                source="gulftax_transactions",
                box_breakdown={"box8_total_output_vat": 500.0},
            )
        )
        self.ported_db.commit()

    def tearDown(self):
        self.db.rollback()
        self.ported_db.rollback()
        self.db.close()
        self.ported_db.close()

    @patch("app.services.audit_export_service._fetch_company_config")
    @patch("app.services.audit_export_service.fetch_all_vat_return_boxes")
    def test_generate_period_audit_pack(self, mock_boxes, mock_company):
        mock_boxes.return_value = _mock_vat_boxes()
        mock_company.return_value = {"name": "Audit Export Test Co", "trn": "100000000000003"}

        result = generate_period_audit_pack(
            self.db,
            self.ported_db,
            tenant_id=self.tenant,
            company_id=self.company_id,
            tax_period=self.period,
            generated_by="auditor@test.com",
        )

        self.assertIn("excel_bytes", result)
        self.assertIn("zip_bytes", result)
        self.assertIn("manifest", result)

        manifest = result["manifest"]
        self.assertEqual(manifest["tax_period"], self.period)
        self.assertEqual(manifest["company_name"], "Audit Export Test Co")
        self.assertEqual(manifest["trn"], "100000000000003")
        self.assertEqual(manifest["generated_by"], "auditor@test.com")
        self.assertIn("excel_sha256", manifest)

        excel_hash = hashlib.sha256(result["excel_bytes"]).hexdigest()
        self.assertEqual(manifest["excel_sha256"], excel_hash)

        wb = load_workbook(io.BytesIO(result["excel_bytes"]))
        expected_sheets = {
            "Cover",
            "VAT Return Summary",
            "Transaction Listing",
            "VAT Reconciliation",
            "CT Return",
            "Advanced VAT",
            "Audit Trail",
        }
        self.assertEqual(set(wb.sheetnames), expected_sheets)

        cover = wb["Cover"]
        self.assertEqual(cover["B3"].value, "Audit Export Test Co")
        self.assertEqual(cover["B4"].value, "100000000000003")
        self.assertEqual(cover["B5"].value, self.period)
        self.assertEqual(cover["B9"].value, "auditor@test.com")

        self.assertGreaterEqual(result["sheet_row_counts"]["Transaction Listing"], 1)
        self.assertGreaterEqual(result["sheet_row_counts"]["Audit Trail"], 1)

        with zipfile.ZipFile(io.BytesIO(result["zip_bytes"])) as zf:
            names = set(zf.namelist())
            self.assertIn("manifest.json", names)
            xlsx_files = [n for n in names if n.endswith(".xlsx")]
            self.assertEqual(len(xlsx_files), 1)
            manifest_in_zip = json.loads(zf.read("manifest.json"))
            self.assertEqual(manifest_in_zip["excel_sha256"], excel_hash)


if __name__ == "__main__":
    unittest.main()
