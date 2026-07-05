#!/usr/bin/env python3
"""Live E2E — audit-ready period export on RDS."""
from __future__ import annotations

import hashlib
import io
import json
import os
import sys
import uuid
import zipfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

def _resolve_backend() -> Path:
    here = Path(__file__).resolve()
    for candidate in (here.parent, *here.parents):
        if (candidate / "app" / "services").is_dir():
            return candidate
    return Path("/app")


_BACKEND = _resolve_backend()
sys.path.insert(0, str(_BACKEND))
_PORTED = _BACKEND / "app" / "modules" / "gulftax" / "ported"
if str(_PORTED) not in sys.path:
    sys.path.insert(0, str(_PORTED))

from dotenv import load_dotenv

load_dotenv(_BACKEND / ".env", override=True)

from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from app.models.client_data import ApCompany, CtReturn, GulftaxTransaction
from app.models.workspace_audit import WorkspaceAuditLog
from models import Company, ReconciliationResult
from app.services.audit_export_service import generate_period_audit_pack

E2E_PREFIX = "E2E-AUDIT-"
TENANT = "59818b25-a981-4fe4-9a1f-7ffaafecef13"
COMPANY_ID = "e26d6523-d86b-4e77-8e16-23f251304480"
TAX_PERIOD = "2026-Q2"
PERIOD_START = date(2026, 4, 1)
PERIOD_END = date(2026, 6, 30)


def _db_url() -> str:
    url = (os.getenv("DATABASE_URL") or os.getenv("SUPABASE_DB_URL") or "").strip()
    if not url.startswith("postgresql"):
        print("ERROR: DATABASE_URL not set. Run on EC2.")
        return ""
    return url


def _ensure_company(db, ported_db) -> None:
    if not db.query(ApCompany).filter_by(id=COMPANY_ID).first():
        db.add(
            ApCompany(
                id=COMPANY_ID,
                tenant_id=TENANT,
                name="ABC TRADING LLC",
                slug="abc-trading-e2e-audit",
            )
        )
        db.commit()

    if not ported_db.query(Company).filter(Company.id == COMPANY_ID).first():
        ported_db.add(
            Company(
                id=COMPANY_ID,
                name="ABC TRADING LLC",
                trade_license_number="TL-E2E-AUDIT",
                trn="100000000000003",
                entity_type="mainland",
            )
        )
        ported_db.commit()


def _seed(db, ported_db) -> str:
    tx_id = str(uuid.uuid4())
    inv_id = str(uuid.uuid4())

    db.query(GulftaxTransaction).filter(
        GulftaxTransaction.tenant_id == TENANT,
        GulftaxTransaction.company_id == COMPANY_ID,
        GulftaxTransaction.tax_period == TAX_PERIOD,
        GulftaxTransaction.invoice_number.like(f"{E2E_PREFIX}%"),
    ).delete(synchronize_session=False)

    ported_db.query(ReconciliationResult).filter(
        ReconciliationResult.company_id == COMPANY_ID,
        ReconciliationResult.tax_period == TAX_PERIOD,
        ReconciliationResult.source == "e2e_audit_export",
    ).delete(synchronize_session=False)

    db.query(CtReturn).filter(
        CtReturn.tenant_id == TENANT,
        CtReturn.company_id == COMPANY_ID,
        CtReturn.period_start == PERIOD_START,
        CtReturn.period_end == PERIOD_END,
    ).delete(synchronize_session=False)

    try:
        db.query(WorkspaceAuditLog).filter(
            WorkspaceAuditLog.workspace_id == TENANT,
            WorkspaceAuditLog.entity_id == f"{E2E_PREFIX}je",
        ).delete(synchronize_session=False)
    except Exception:
        db.rollback()

    db.add(
        GulftaxTransaction(
            id=tx_id,
            tenant_id=TENANT,
            company_id=COMPANY_ID,
            source="ap_invoiceflow",
            ap_invoice_id=inv_id,
            tax_period=TAX_PERIOD,
            transaction_date=date(2026, 5, 10),
            vendor_name=f"{E2E_PREFIX}Vendor",
            vendor_trn="100000000000003",
            invoice_number=f"{E2E_PREFIX}INV-001",
            gross_amount=Decimal("1050.00"),
            vat_amount=Decimal("50.00"),
            vat_category="standard",
            fta_box="box9",
            direction="input",
            status="posted",
        )
    )
    db.add(
        CtReturn(
            tenant_id=TENANT,
            company_id=COMPANY_ID,
            period_start=PERIOD_START,
            period_end=PERIOD_END,
            revenue=Decimal("1200000"),
            accounting_profit=Decimal("80000"),
            taxable_income=Decimal("75000"),
            ct_payable_aed=Decimal("6750"),
            sbr_eligible=False,
            qfzp_eligible=False,
            free_zone_status="mainland",
            status="approved",
            approved_at=datetime.utcnow(),
        )
    )
    db.commit()

    ported_db.add(
        ReconciliationResult(
            company_id=COMPANY_ID,
            tax_period=TAX_PERIOD,
            period_start=PERIOD_START,
            period_end=PERIOD_END,
            difference_aed=0.0,
            status="matched",
            source="e2e_audit_export",
            box_breakdown={"box8_total_output_vat": 100.0, "box11_total_input_vat": 50.0},
        )
    )
    ported_db.commit()
    return tx_id


def main() -> int:
    url = _db_url()
    if not url:
        return 1

    engine = create_engine(url)
    Session = sessionmaker(bind=engine)
    db = Session()
    ported_db = Session()

    try:
        _ensure_company(db, ported_db)
        _seed(db, ported_db)

        result = generate_period_audit_pack(
            db,
            ported_db,
            tenant_id=TENANT,
            company_id=COMPANY_ID,
            tax_period=TAX_PERIOD,
            generated_by="e2e@finreportai.com",
        )

        manifest = result["manifest"]
        excel_bytes = result["excel_bytes"]
        excel_hash = hashlib.sha256(excel_bytes).hexdigest()

        wb = load_workbook(io.BytesIO(excel_bytes))
        sheet_counts = {name: wb[name].max_row for name in wb.sheetnames}

        with zipfile.ZipFile(io.BytesIO(result["zip_bytes"])) as zf:
            zip_manifest = json.loads(zf.read("manifest.json"))

        checks = {
            "manifest_hash_matches": manifest.get("excel_sha256") == excel_hash,
            "zip_manifest_hash_matches": zip_manifest.get("excel_sha256") == excel_hash,
            "seven_sheets": len(wb.sheetnames) == 7,
            "transaction_listing_has_data": sheet_counts.get("Transaction Listing", 0) >= 2,
            "audit_trail_sheet_present": "Audit Trail" in wb.sheetnames,
            "cover_has_company": wb["Cover"]["B3"].value is not None,
        }
        passed = all(checks.values())

        report = {
            "pass": passed,
            "tax_period": TAX_PERIOD,
            "excel_sha256": excel_hash,
            "sheet_row_counts": result["sheet_row_counts"],
            "workbook_max_rows": sheet_counts,
            "manifest_artifacts": manifest.get("artifacts"),
            "checks": checks,
        }
        print(json.dumps(report, indent=2))
        return 0 if passed else 1
    except Exception as exc:
        print(json.dumps({"pass": False, "error": str(exc)}))
        return 1
    finally:
        db.close()
        ported_db.close()


if __name__ == "__main__":
    raise SystemExit(main())
